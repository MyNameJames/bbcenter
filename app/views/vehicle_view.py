from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, session, current_app
from flask_login import login_required, current_user
from models import db, get_bkk_time, User, Vehicle, VehicleBooking, Driver, VehicleMileage, SystemConfig, VehicleBudget, VehicleBudgetLog, VehicleDepartment, BudgetType, Notification, DeptApprover, OTRateConfig, DriverOT, DriverOTSlot, FuelPrice, FuelBill, RepairTicket, MaintenanceTicket, RoomBooking
from sqlalchemy import and_, extract, or_, func
from datetime import datetime, date, timedelta
from views.telegram_service import (notify_approved, notify_forwarded_to_approver, notify_approver_approved, notify_rejected,
                                    notify_cancelled            as tg_notify_cancelled)
from views.notification_service import (
    notify_booking_created      as _n_booking_created,
    notify_admin_assigned       as _n_admin_assigned,
    notify_admin_approved       as _n_admin_approved,
    notify_forwarded_to_approver as _n_forwarded,
    notify_approver_approved    as _n_approver_approved,
    notify_rejected             as _n_rejected,
    notify_merged_into_group    as _n_merged,
    notify_mileage_started      as _n_mileage_start,
    notify_mileage_ended        as _n_mileage_end,
    notify_budget_deducted      as _n_budget,
    notify_payment_required     as _n_payment_required,
    notify_admin_deleted        as _n_admin_deleted,
    notify_payment_confirmed    as _n_payment_confirmed,
    notify_user_cancelled       as _n_user_cancelled,
)
from services import budget_service as budget_svc
import os, time
from werkzeug.utils import secure_filename

vehicle_bp    = Blueprint('vehicle', __name__)
adminfleet_bp = Blueprint('adminfleet', __name__)
admincost_bp  = Blueprint('admincost', __name__)
driver_bp     = Blueprint('driver', __name__)

EXPENSE_CATEGORIES = {
    # ── ส่วนกลาง ──────────────────────────────────────────────
    # เพิ่ม/ลบ หมวดย่อยที่นี่ → จะขึ้นใน dropdown อัตโนมัติ
    "central": [
        {"key": "medical",       "label": "ค่ารักษาพยาบาล"},
        {"key": "training",      "label": "ค่าอบรม / สัมมนา"},
        {"key": "religious",     "label": "งานกิจนิมนต์ / ศาสนา"},
        {"key": "official",      "label": "ราชการ / ติดต่อหน่วยงาน"},
        {"key": "welfare",       "label": "สวัสดิการ / เยี่ยมไข้"},
        {"key": "procurement",   "label": "จัดซื้อจัดจ้าง"},
        {"key": "other_central", "label": "อื่น ๆ (ส่วนกลาง)"},
    ],
    # ── งานกอง (department) ───────────────────────────────────
    # แต่ละ key = ชื่อกอง, label = ชื่อที่แสดง
    "department": [
        {"key": "กองสนับสนุนและบริการ", "label": "กองสนับสนุนและบริการ"},
        {"key": "กองวิชาการ",           "label": "กองวิชาการ"},
        {"key": "กองDOU",               "label": "กองDOU"},
        {"key": "กองบริหาร",            "label": "กองบริหาร"},
        {"key": "กองเลขานุการ",         "label": "กองเลขานุการ"},
        {"key": "กองกิจการนิสิต",       "label": "กองกิจการนิสิต"},
        {"key": "กองพระไตรปิฏก",        "label": "กองพระไตรปิฏก"},
    ],
}


def is_vehicle_admin():
    return current_user.role_vehicle == 'admin' or current_user.is_superadmin


# ─────────────────────────────────────────────
# หน้าหลัก
# ─────────────────────────────────────────────
@vehicle_bp.route('/vehicle')
@login_required
def index():
    # ad-hoc (driver-created off-the-books) ซ่อนจากปฏิทินผู้ใช้
    bookings  = (VehicleBooking.query
                 .filter(VehicleBooking.is_ad_hoc == False)
                 .order_by(VehicleBooking.created_at.desc()).all())
    vehicles  = Vehicle.query.filter_by(status='active').order_by(Vehicle.id).all()
    drivers   = Driver.query.filter_by(is_active=True).order_by(Driver.id).all()
    return render_template(
        'vehicle/vehicle.html',
        bookings=bookings,
        vehicles=vehicles,
        drivers=drivers,
        expense_categories=EXPENSE_CATEGORIES,
        total_vehicles=len(vehicles),
        page_section='บริการ',
        page_title='ปฏิทินการจองรถ',
        # Phase 9 (2026-05-22) — `canCancel` gating needs admin + now
        now=datetime.now(),
        is_vehicle_admin=is_vehicle_admin(),
    )


# ─────────────────────────────────────────────
# จองรถแบบใหม่ — ไม่ต้องเลือกรถ admin กำหนดให้
# ─────────────────────────────────────────────
@vehicle_bp.route('/vehicle/book', methods=['POST'])
@login_required
def book_vehicle_simple():
    try:
        start_str = request.form.get('start_datetime', '').strip()
        end_str   = request.form.get('end_datetime',   '').strip()

        if not start_str or not end_str:
            flash('กรุณาเลือกวัน-เวลาไปและกลับให้ครบ', 'warning')
            return redirect(url_for('vehicle.index'))

        start_dt        = datetime.strptime(start_str, '%Y-%m-%dT%H:%M')
        end_dt          = datetime.strptime(end_str,   '%Y-%m-%dT%H:%M')
        destination      = request.form.get('destination')
        purpose          = request.form.get('purpose')
        passenger_count  = int(request.form.get('passenger_count', 1))
        need_driver      = request.form.get('need_driver') == 'on'
        pickup_location  = request.form.get('pickup_location', '').strip()

        if start_dt < datetime.now():
            flash('ไม่สามารถจองย้อนหลังได้ กรุณาเลือกวันเวลาในอนาคต', 'warning')
            return redirect(url_for('vehicle.index'))

        if start_dt >= end_dt:
            flash('เวลากลับต้องมากกว่าเวลาไป', 'danger')
            return redirect(url_for('vehicle.index'))

        # ── ข้อ 5: ห้ามจองข้ามวัน ─────────────────────────────
        if start_dt.date() != end_dt.date():
            flash('ไม่สามารถจองข้ามวันได้ กรุณาเลือกวันเริ่มและสิ้นสุดเป็นวันเดียวกัน', 'warning')
            return redirect(url_for('vehicle.index'))

        new_booking = VehicleBooking(
            user_id         = current_user.id,
            start_datetime  = start_dt,
            end_datetime    = end_dt,
            destination     = destination,
            purpose         = purpose,
            passenger_count = passenger_count,
            need_driver     = need_driver,
            status          = 'pending',
            pickup_location = pickup_location or None,
        )
        db.session.add(new_booking)
        db.session.flush()
        _n_booking_created(new_booking)
        db.session.commit()
        flash(f'ส่งคำขอจองรถเรียบร้อยแล้ว (#{ new_booking.id }) รอ Admin พิจารณา', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'เกิดข้อผิดพลาด: {str(e)}', 'danger')

    return redirect(url_for('vehicle.index'))



# ─────────────────────────────────────────────
# แก้ไขการจอง
# ─────────────────────────────────────────────
@vehicle_bp.route('/vehicle/edit/<int:booking_id>', methods=['GET', 'POST'])
@login_required
def edit_booking(booking_id):
    booking = VehicleBooking.query.get_or_404(booking_id)

    if current_user.id != booking.user_id:
        flash('คุณไม่มีสิทธิ์แก้ไขรายการนี้', 'danger')
        return redirect(url_for('vehicle.index'))

    # อนุญาตแก้ไขได้เฉพาะสถานะ pending เท่านั้น
    if booking.status != 'pending':
        flash('ไม่สามารถแก้ไขได้ เนื่องจากคำขอนี้ถูกดำเนินการแล้ว', 'warning')
        return redirect(url_for('vehicle.index'))

    if request.method == 'POST':
        try:
            booking.start_datetime  = datetime.strptime(request.form.get('start_datetime'), '%Y-%m-%dT%H:%M')
            booking.end_datetime    = datetime.strptime(request.form.get('end_datetime'),   '%Y-%m-%dT%H:%M')
            booking.destination     = request.form.get('destination')
            booking.purpose         = request.form.get('purpose')
            booking.passenger_count = int(request.form.get('passenger_count', 1))
            booking.need_driver     = True if request.form.get('need_driver') else False
            booking.pickup_location = request.form.get('pickup_location', '').strip() or None

            db.session.commit()
            flash('อัปเดตข้อมูลการจองเรียบร้อยแล้ว', 'success')
            return redirect(url_for('vehicle.index'))
        except Exception as e:
            db.session.rollback()
            flash(f'เกิดข้อผิดพลาด: {str(e)}', 'danger')

    start_str = booking.start_datetime.strftime('%Y-%m-%dT%H:%M')
    end_str   = booking.end_datetime.strftime('%Y-%m-%dT%H:%M')
    return render_template('vehicle/vehicle_edit.html', booking=booking, start_str=start_str, end_str=end_str)


# ─────────────────────────────────────────────
# ลบการจอง
# ─────────────────────────────────────────────
@vehicle_bp.route('/vehicle/delete/<int:booking_id>', methods=['POST'])
@login_required
def delete_booking(booking_id):
    booking = VehicleBooking.query.get_or_404(booking_id)

    # Admin ลบได้ทุกสถานะ แต่ User ทั่วไปลบได้เฉพาะ pending และ rejected
    if not is_vehicle_admin() and current_user.id != booking.user_id:
        flash('คุณไม่มีสิทธิ์ลบรายการนี้', 'danger')
        return redirect(url_for('vehicle.index'))

    if not is_vehicle_admin() and booking.status in ('approved', 'waiting_approver'):
        flash('ไม่สามารถลบคำขอที่อยู่ในระหว่างดำเนินการหรืออนุมัติแล้วได้ กรุณาติดต่อ Admin', 'warning')
        return redirect(url_for('vehicle.index'))

    try:
        # ถ้า admin ลบของคนอื่น → แจ้งเตือน user (Event #15)
        should_notify = is_vehicle_admin() and current_user.id != booking.user_id
        snap = (booking.id, booking.user_id, booking.destination)

        # คืนงบก่อนลบ (อ่าน booking.mileage ได้ก่อน cascade) — no-op ถ้ายังไม่เคยหัก
        budget_svc.refund_for_booking(
            booking,
            note=f'delete booking #{booking.id} by {current_user.username}',
        )
        db.session.flush()

        db.session.delete(booking)
        db.session.flush()
        if should_notify:
            _n_admin_deleted(snap[0], snap[1], snap[2], current_user)
        db.session.commit()
        flash('ยกเลิกและลบรายการจองเรียบร้อยแล้ว', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'เกิดข้อผิดพลาดในการลบ: {str(e)}', 'danger')

    return redirect(url_for('vehicle.index'))


# ─────────────────────────────────────────────
# Phase 9 (2026-05-22) — Cancel booking (soft, status='cancelled')
# C1: User+Admin can cancel pending/waiting_approver/approved bookings
# Time guard: must be BEFORE booking.start_datetime
# ─────────────────────────────────────────────
@vehicle_bp.route('/vehicle/cancel/<int:booking_id>', methods=['POST'])
@login_required
def cancel_booking(booking_id):
    """C1 — Soft cancel approved/pending/waiting_approver booking. Row kept for audit."""
    booking = VehicleBooking.query.get_or_404(booking_id)

    # Permission: owner OR admin
    is_owner = (current_user.id == booking.user_id)
    is_admin = is_vehicle_admin()
    if not (is_owner or is_admin):
        flash('คุณไม่มีสิทธิ์ยกเลิกการจองนี้', 'danger')
        return redirect(url_for('vehicle.index'))

    # Status guard
    if booking.status not in ('pending', 'waiting_approver', 'approved'):
        flash(f'ยกเลิกไม่ได้ — สถานะปัจจุบันคือ {booking.status}', 'warning')
        return redirect(url_for('vehicle.detail_booking', booking_id=booking_id))

    # Time guard — block หลัง trip start
    if datetime.now() >= booking.start_datetime:
        flash('ทริปเริ่มแล้ว ไม่สามารถยกเลิกได้ — ติดต่อ Admin หากจำเป็น', 'warning')
        return redirect(url_for('vehicle.detail_booking', booking_id=booking_id))

    try:
        prev_status = booking.status

        # ── Refund budget (idempotent — no-op ถ้ายังไม่เคยหัก)
        refunds = budget_svc.refund_for_booking(
            booking,
            note=f'cancel booking #{booking.id} by {current_user.username}',
        )
        db.session.flush()

        # ── Build recipient sets BEFORE status flip
        # Priority cascade (highest → lowest): owner > admin > approver > driver > mate
        # ทุก lower-priority set จะ discard user_id ที่อยู่ใน higher-priority set แล้ว
        # → ทุก user_id ได้รับ notification เพียง 1 ใบเท่านั้น (role_label = highest priority)
        already_notified = set()
        already_notified.add(current_user.id)  # canceler never notifies themselves

        # 1) Owner (priority #1) — only if admin cancels someone else's booking
        owner_notify_id = booking.user_id if (is_admin and not is_owner) else None
        if owner_notify_id and owner_notify_id not in already_notified:
            already_notified.add(owner_notify_id)
        else:
            owner_notify_id = None  # ป้องกัน double-notify ถ้า owner = canceler

        # 2) Admin user_ids (priority #2)
        admin_user_ids = {u.id for u in User.query.filter(
            or_(User.role_vehicle == 'admin', User.is_superadmin.is_(True))
        ).all()}
        admin_user_ids -= already_notified
        already_notified |= admin_user_ids

        # 3) Approver user_ids (priority #3) — only if booking went through approver flow
        approver_user_ids = set()
        if booking.trip_department_id and prev_status in ('waiting_approver', 'approved'):
            apv_rows = DeptApprover.query.filter_by(
                dept_id=booking.trip_department_id).all()
            approver_user_ids = {r.user_id for r in apv_rows} - already_notified
            already_notified |= approver_user_ids

        # 4) Driver-as-user (priority #4) — if driver.user_id linked
        driver_user_id = None
        if booking.driver_id and booking.driver and booking.driver.user_id:
            cand = booking.driver.user_id
            if cand not in already_notified:
                driver_user_id = cand
                already_notified.add(cand)

        # 5) Trip mates (priority #5) — other user_ids in same trip_group
        trip_mate_user_ids = set()
        if booking.trip_group:
            mate_rows = VehicleBooking.query.filter(
                VehicleBooking.trip_group == booking.trip_group,
                VehicleBooking.id != booking.id,
            ).all()
            trip_mate_user_ids = {m.user_id for m in mate_rows if m.user_id} - already_notified
            already_notified |= trip_mate_user_ids

        # ── Notify (in-app) — ลำดับตาม priority
        if owner_notify_id:
            _n_user_cancelled(user_id=owner_notify_id, booking=booking,
                              cancelled_by=current_user, role_label='owner')

        for uid in admin_user_ids:
            _n_user_cancelled(user_id=uid, booking=booking,
                              cancelled_by=current_user, role_label='admin')

        for uid in approver_user_ids:
            _n_user_cancelled(user_id=uid, booking=booking,
                              cancelled_by=current_user, role_label='approver')

        if driver_user_id:
            _n_user_cancelled(user_id=driver_user_id, booking=booking,
                              cancelled_by=current_user, role_label='driver')

        for uid in trip_mate_user_ids:
            _n_user_cancelled(user_id=uid, booking=booking,
                              cancelled_by=current_user, role_label='mate')

        # ── Soft cancel — flip status, row kept
        booking.status = 'cancelled'
        booking.updated_by = current_user.id

        # ── Telegram — delete old (approved/assigned msg) + send cancel
        # Pattern matches approve_booking — let exception abort txn (consistent w/ project convention)
        tg_notify_cancelled(booking, current_user)

        db.session.commit()

        if refunds:
            flash(f'ยกเลิกการจอง #{booking_id} เรียบร้อย · คืนงบ {len(refunds)} รายการ', 'success')
        else:
            flash(f'ยกเลิกการจอง #{booking_id} เรียบร้อย', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'ยกเลิกไม่สำเร็จ: {e}', 'danger')

    return redirect(url_for('vehicle.index'))


# ─────────────────────────────────────────────
# รายละเอียดทริป
# ─────────────────────────────────────────────
@vehicle_bp.route('/vehicle/detail/<int:booking_id>')
@login_required
def detail_booking(booking_id):
    booking = VehicleBooking.query.get_or_404(booking_id)

    _is_dept_approver = DeptApprover.query.filter_by(
        user_id=current_user.id, dept_id=booking.trip_department_id
    ).first() is not None
    if (not is_vehicle_admin()
            and not _is_dept_approver
            and not current_user.is_superadmin
            and current_user.id != booking.user_id):
        flash('คุณไม่มีสิทธิ์เข้าถึงข้อมูลนี้', 'danger')
        return redirect(url_for('vehicle.index'))

    drivers = Driver.query.filter_by(is_active=True).all()
    return render_template('vehicle/vehicle_detail.html',
                           booking=booking, drivers=drivers,
                           is_dept_approver=_is_dept_approver)


# ─────────────────────────────────────────────
# Approver Inbox
# ─────────────────────────────────────────────
@vehicle_bp.route('/vehicle/approver')
@login_required
def approver_inbox():
    my_rows = DeptApprover.query.filter_by(user_id=current_user.id).all()
    if not my_rows and not current_user.is_superadmin:
        flash('คุณไม่มีสิทธิ์เข้าถึงหน้านี้', 'danger')
        return redirect(url_for('vehicle.index'))

    my_dept_ids = [r.dept_id for r in my_rows]

    pending = (VehicleBooking.query
               .filter(
                   VehicleBooking.status == 'waiting_approver',
                   VehicleBooking.trip_department_id.in_(my_dept_ids)
               )
               .order_by(VehicleBooking.start_datetime.asc())
               .all())

    history = (VehicleBooking.query
               .filter(
                   VehicleBooking.status.in_(['approved', 'rejected']),
                   VehicleBooking.trip_department_id.in_(my_dept_ids),
                   VehicleBooking.updated_by == current_user.id
               )
               .order_by(VehicleBooking.updated_at.desc())
               .limit(20)
               .all())

    from models import get_bkk_time
    today = get_bkk_time().date()
    budgets = (VehicleBudget.query
               .filter(
                   VehicleBudget.approver_id == current_user.id,
                   VehicleBudget.year == today.year,
                   VehicleBudget.month == today.month
               )
               .all())

    return render_template('vehicle/approver_inbox.html',
                           pending=pending, history=history,
                           budgets=budgets,
                           active_menu='approver')


# ─────────────────────────────────────────────
# API ปฏิทิน
# ─────────────────────────────────────────────
@vehicle_bp.route('/api/vehicle/bookings')
@login_required
def api_bookings():
    bookings = VehicleBooking.query.filter(
        VehicleBooking.status.in_(['pending', 'waiting_approver', 'approved']),
        VehicleBooking.is_ad_hoc == False
    ).all()

    events = []
    for b in bookings:
        color = '#198754' if b.status == 'approved' else ('#0dcaf0' if b.status == 'waiting_approver' else '#ffc107')
        vehicle_label = f"{b.assigned_vehicle.brand} {b.assigned_vehicle.model}" if b.assigned_vehicle else "รอกำหนดรถ"
        events.append({
            'id':    b.id,
            'title': f"{b.destination} ({b.user.full_name or b.user.username})",
            'start': b.start_datetime.isoformat() if b.start_datetime else None,
            'end':   b.end_datetime.isoformat()   if b.end_datetime   else None,
            'color': color,
            'url':   url_for('vehicle.detail_booking', booking_id=b.id)
        })
    return jsonify(events)


@vehicle_bp.route('/api/custom-bookings')
@login_required
def custom_bookings():
    bookings = VehicleBooking.query.filter(
        VehicleBooking.status.in_(['pending', 'waiting_approver', 'approved']),
        VehicleBooking.is_ad_hoc == False
    ).all()

    events = []
    for b in bookings:
        color = '#198754' if b.status == 'approved' else ('#0dcaf0' if b.status == 'waiting_approver' else '#ffc107')
        vehicle_label = f"{b.assigned_vehicle.brand} {b.assigned_vehicle.model}" if b.assigned_vehicle else "รอกำหนดรถ"
        events.append({
            'id':       b.id,
            'booker':   b.user.full_name or b.user.username,
            'title':    b.destination,
            'start':    b.start_datetime.isoformat() if b.start_datetime else None,
            'end':      b.end_datetime.isoformat()   if b.end_datetime   else None,
            'dest':     b.destination,
            'pax':      b.passenger_count,
            'booker':   b.user.full_name or b.user.username,
            'status':   b.status,
            'url':   url_for('vehicle.detail_booking', booking_id=b.id),
        })
    return jsonify(events)
# ─────────────────────────────────────────────
# Budget lookup helper — ใช้ใน approve_booking เพื่อ check is_active
# ─────────────────────────────────────────────
def _lookup_budget_for_booking(booking):
    """หา VehicleBudget row ที่ booking นี้จะหักงบจริง (ใช้ start_datetime month).
    คืน (budget, key_label) — budget=None ถ้าไม่พบ"""
    if booking.expense_type not in ('central', 'department'):
        return None, None
    if not booking.start_datetime:
        return None, None
    bt = BudgetType.query.filter_by(name=booking.expense_type).first()
    if not bt:
        return None, booking.expense_type

    if booking.expense_type == 'central':
        key_label = booking.central_category
        dept_obj = VehicleDepartment.query.filter_by(name=key_label).first() if key_label else None
    else:
        key_label = booking.trip_department or (booking.user.department if booking.user else None)
        if booking.trip_department_id:
            dept_obj = VehicleDepartment.query.get(booking.trip_department_id)
        elif key_label:
            dept_obj = VehicleDepartment.query.filter_by(name=key_label).first()
        else:
            dept_obj = None
    if not dept_obj:
        return None, key_label

    budget = VehicleBudget.query.filter_by(
        department_id=dept_obj.id,
        year=booking.start_datetime.year,
        month=booking.start_datetime.month,
        budget_type_id=bt.id,
    ).first()
    return budget, key_label


# ─────────────────────────────────────────────
# อนุมัติ / ปฏิเสธ
# ─────────────────────────────────────────────
@vehicle_bp.route('/vehicle/approve/<int:booking_id>', methods=['POST'])
@login_required
def approve_booking(booking_id):
    booking   = VehicleBooking.query.get_or_404(booking_id)
    action    = request.form.get('action')
    driver_id = request.form.get('driver_id')

    acted_as_approver = False
    try:
        if is_vehicle_admin() and booking.status == 'pending':
            if action == 'approve':
                # Block ถ้า target budget ถูกปิดใช้งาน
                _bgt, _kl = _lookup_budget_for_booking(booking)
                if _bgt and not _bgt.is_active:
                    flash(
                        f'อนุมัติไม่ได้ — งบ "{_bgt.department.name}" '
                        f'เดือน {_bgt.month}/{_bgt.year + 543} ถูกปิดใช้งานอยู่',
                        'danger'
                    )
                    return redirect(url_for('vehicle.detail_booking', booking_id=booking.id))
                if driver_id: booking.driver_id = driver_id
                if booking.expense_type == 'department':
                    if booking.trip_department_id is None:
                        dept_name = booking.trip_department or booking.user.department
                        if dept_name:
                            dept = VehicleDepartment.query.filter_by(name=dept_name).first()
                            if dept:
                                booking.trip_department_id = dept.id
                    booking.status = 'waiting_approver'
                    db.session.flush()
                    notify_forwarded_to_approver(booking)    # Telegram
                    _n_forwarded(booking)                    # In-app Event #4
                    db.session.commit()
                    flash(f'อนุมัติแล้ว — รอผู้ประสานงานแผนก {booking.user.department} ยืนยัน', 'info')
                else:
                    booking.status = 'approved'
                    db.session.flush()
                    notify_approved(booking)                 # Telegram
                    _n_admin_approved(booking)               # In-app Event #3
                    db.session.commit()
                    flash('อนุมัติการจองรถเรียบร้อย', 'success')
            elif action == 'reject':
                booking.status = 'rejected'
                booking.reject_reason = request.form.get('reject_reason', '').strip() or None
                db.session.flush()
                # คืนงบถ้า mileage เคยปิด+หักไปแล้ว (no-op ถ้ายังไม่เคยหัก)
                budget_svc.refund_for_booking(
                    booking,
                    note=f'reject by admin {current_user.username}: {booking.reject_reason or "—"}',
                )
                notify_rejected(booking, current_user)       # Telegram
                _n_rejected(booking, current_user, by_approver=False)  # In-app Event #6
                db.session.commit()
                flash('ไม่อนุมัติการจองรถ', 'danger')

        elif DeptApprover.query.filter_by(user_id=current_user.id).first():
            acted_as_approver = True
            if booking.status != 'waiting_approver':
                flash('รายการนี้ไม่ได้อยู่ในสถานะรอคุณอนุมัติ', 'warning')
                return redirect(url_for('vehicle.approver_inbox'))

            my_dept_ids = [r.dept_id for r in DeptApprover.query.filter_by(user_id=current_user.id).all()]
            if booking.trip_department_id not in my_dept_ids:
                flash('คุณสามารถอนุมัติได้เฉพาะแผนกที่คุณรับผิดชอบเท่านั้น', 'danger')
                return redirect(url_for('vehicle.approver_inbox'))

            if action == 'approve':
                # Block ถ้า target budget ถูกปิดใช้งาน (approver path)
                _bgt, _kl = _lookup_budget_for_booking(booking)
                if _bgt and not _bgt.is_active:
                    flash(
                        f'อนุมัติไม่ได้ — งบ "{_bgt.department.name}" '
                        f'เดือน {_bgt.month}/{_bgt.year + 543} ถูกปิดใช้งานอยู่',
                        'danger'
                    )
                    return redirect(url_for('vehicle.approver_inbox'))
                booking.status = 'approved'
                booking.updated_by = current_user.id
                db.session.flush()
                notify_approver_approved(booking, current_user)   # Telegram
                _n_approver_approved(booking, current_user)       # In-app Event #5
                db.session.commit()
                flash('อนุมัติการเดินทางเรียบร้อยแล้ว', 'success')
            elif action == 'reject':
                booking.status = 'rejected'
                booking.updated_by = current_user.id
                booking.reject_reason = request.form.get('reject_reason', '').strip() or None
                db.session.flush()
                # คืนงบถ้า mileage เคยปิด+หักไปแล้ว (no-op ถ้ายังไม่เคยหัก)
                budget_svc.refund_for_booking(
                    booking,
                    note=f'reject by approver {current_user.username}: {booking.reject_reason or "—"}',
                )
                notify_rejected(booking, current_user)            # Telegram
                _n_rejected(booking, current_user, by_approver=True)  # In-app Event #6
                db.session.commit()
                flash('ปฏิเสธการเดินทางนี้แล้ว', 'danger')
        else:
            flash('คุณไม่มีสิทธิ์ทำรายการนี้', 'danger')
    except Exception as e:
        db.session.rollback()
        flash(f'เกิดข้อผิดพลาด: {str(e)}', 'danger')

    if acted_as_approver:
        return redirect(url_for('vehicle.approver_inbox'))
    return redirect(url_for('vehicle.detail_booking', booking_id=booking.id))


# ─────────────────────────────────────────────
# Notifications API
# ─────────────────────────────────────────────
@vehicle_bp.route('/api/notifications')
@login_required
def api_notifications():
    from models import get_bkk_time
    now = get_bkk_time()
    cutoff_90d = now - timedelta(days=90)

    # ดึง 90 วันล่าสุด (ไม่จำกัดจำนวน — frontend จัดการ pagination ด้วย since)
    notifs = Notification.query.filter(
        Notification.user_id == current_user.id,
        Notification.created_at >= cutoff_90d
    ).order_by(Notification.created_at.desc()).limit(200).all()

    # unread count: ไม่นับที่หมดอายุแล้ว (expired_at != null AND now > expired_at)
    unread = Notification.query.filter(
        Notification.user_id == current_user.id,
        Notification.is_read == False,
        or_(Notification.expired_at.is_(None), Notification.expired_at > now)
    ).count()

    # Payment unpaid (sticky) — ไม่หมดอายุ
    sticky = Notification.query.filter(
        Notification.user_id == current_user.id,
        Notification.is_sticky == True,
        Notification.is_read == False,
        Notification.category.in_(['payment', 'payment_admin'])
    ).order_by(Notification.created_at.desc()).all()

    unread_payment = sum(1 for n in sticky)

    # Helper: relative time
    def _rel_time(dt):
        if not dt: return ''
        d = now - dt
        sec = int(d.total_seconds())
        if sec < 60:   return 'เมื่อสักครู่'
        if sec < 3600: return f'{sec // 60} นาทีที่แล้ว'
        if sec < 86400: return f'{sec // 3600} ชั่วโมงที่แล้ว'
        if sec < 604800: return f'{sec // 86400} วันที่แล้ว'
        return dt.strftime('%d/%m/%Y')

    def _to_dict(n):
        b = n.booking
        return {
            'id':          n.id,
            'message':     n.message,
            'ntype':       n.ntype,
            'category':    n.category or 'status',
            'icon':        n.icon or 'fa-solid fa-circle-info',
            'is_read':     n.is_read,
            'is_sticky':   bool(n.is_sticky),
            'booking_id':  n.booking_id,
            'booking_title': (b.destination if b else None),
            'action_url':  n.action_url or (f'/vehicle/detail/{n.booking_id}' if n.booking_id else '#'),
            'created_at':  n.created_at.strftime('%d/%m/%Y %H:%M') if n.created_at else '',
            'created_rel': _rel_time(n.created_at),
        }

    # Group by booking_id (non-sticky only)
    groups_map = {}
    loose_items = []
    for n in notifs:
        if n.is_sticky and not n.is_read:
            continue  # sticky แสดงที่ sticky section แทน
        d = _to_dict(n)
        bid = n.booking_id
        if bid:
            g = groups_map.setdefault(bid, {
                'booking_id':    bid,
                'booking_title': d['booking_title'] or f'คำขอ #{bid}',
                'items':         [],
                'unread_count':  0,
                'latest':        None,
            })
            g['items'].append(d)
            if not d['is_read']:
                g['unread_count'] += 1
            if g['latest'] is None:
                g['latest'] = d
                g['_sort_dt'] = n.created_at  # raw datetime for sorting (DD/MM/YYYY string sort broken)
        else:
            loose_items.append(d)

    # ══════════════════════════════════════════════════════════
    # Stage tracker — 3 roles: user > approver > admin
    # ══════════════════════════════════════════════════════════
    # Role detection (global)
    is_admin_role = (current_user.role_vehicle == 'admin') or current_user.is_superadmin
    approver_dept_ids = {d.dept_id for d in DeptApprover.query.filter(
        DeptApprover.user_id == current_user.id
    ).all()}

    # ── Synthetic groups (approver/admin) — booking ที่ไม่มี notif target current_user
    if approver_dept_ids or is_admin_role:
        existing_bids = set(groups_map.keys())
        synth_bids = set()

        if approver_dept_ids:
            approver_synth = (VehicleBooking.query
                              .filter(VehicleBooking.trip_department_id.in_(approver_dept_ids),
                                      VehicleBooking.status.in_(['waiting_approver', 'approved', 'rejected']),
                                      VehicleBooking.created_at >= cutoff_90d,
                                      VehicleBooking.is_ad_hoc == False)
                              .order_by(VehicleBooking.updated_at.desc().nullslast(),
                                        VehicleBooking.created_at.desc())
                              .limit(50).all())
            for b in approver_synth:
                if b.id not in existing_bids:
                    synth_bids.add(b.id)

        if is_admin_role:
            admin_synth = (VehicleBooking.query
                           .filter(VehicleBooking.created_at >= (now - timedelta(days=60)),
                                   VehicleBooking.is_ad_hoc == False,
                                   VehicleBooking.status.in_(['pending', 'waiting_approver', 'approved']))
                           .order_by(VehicleBooking.updated_at.desc().nullslast(),
                                     VehicleBooking.created_at.desc())
                           .limit(50).all())
            for b in admin_synth:
                if b.id not in existing_bids:
                    synth_bids.add(b.id)

        if synth_bids:
            synth_bookings = VehicleBooking.query.filter(VehicleBooking.id.in_(synth_bids)).all()
            for b in synth_bookings:
                groups_map[b.id] = {
                    'booking_id':   b.id,
                    'booking_title': b.destination or f'คำขอ #{b.id}',
                    'items':        [],
                    'unread_count': 0,
                    'latest':       None,
                    'is_synthetic': True,
                }

    groups = list(groups_map.values())

    # ── Bulk fetch — สำหรับทุก booking ใน groups (ไม่จำกัด role) ──
    all_booking_ids = [g['booking_id'] for g in groups if g['booking_id']]
    booking_map = {}
    mileage_map = {}
    log_map = {}
    notifs_by_booking = {}
    mates_by_group = {}
    mate_users = {}
    updater_users = {}

    if all_booking_ids:
        all_bookings = VehicleBooking.query.filter(
            VehicleBooking.id.in_(all_booking_ids)
        ).all()
        booking_map = {b.id: b for b in all_bookings}

        mileages = VehicleMileage.query.filter(
            VehicleMileage.booking_id.in_(all_booking_ids)
        ).all()
        mileage_map = {m.booking_id: m for m in mileages}

        logs = (VehicleBudgetLog.query
                .filter(VehicleBudgetLog.booking_id.in_(all_booking_ids),
                        VehicleBudgetLog.event_type == 'deduct')
                .order_by(VehicleBudgetLog.created_at.desc())
                .all())
        for log in logs:
            if log.booking_id not in log_map:
                log_map[log.booking_id] = log

        # All notifications for these bookings (any user_id) — สำหรับ event timestamp
        all_notifs = (Notification.query
                      .filter(Notification.booking_id.in_(all_booking_ids))
                      .order_by(Notification.created_at.asc())
                      .all())
        for n in all_notifs:
            notifs_by_booking.setdefault(n.booking_id, []).append(n)

        # Trip mates (สำหรับ user stage)
        trip_groups_set = {b.trip_group for b in booking_map.values() if b.trip_group}
        if trip_groups_set:
            all_mates = (VehicleBooking.query
                         .filter(VehicleBooking.trip_group.in_(trip_groups_set))
                         .all())
            mate_user_ids = {m.user_id for m in all_mates if m.user_id}
            if mate_user_ids:
                mate_users = {u.id: u for u in User.query.filter(User.id.in_(mate_user_ids)).all()}
            for m in all_mates:
                mates_by_group.setdefault(m.trip_group, []).append(m)

        # Updater users (สำหรับ admin/approver stage actor names)
        updater_ids = {b.updated_by for b in booking_map.values() if b.updated_by}
        booking_owner_ids = {b.user_id for b in booking_map.values() if b.user_id}
        all_user_ids = updater_ids | booking_owner_ids
        if all_user_ids:
            updater_users = {u.id: u for u in User.query.filter(User.id.in_(all_user_ids)).all()}

    # ── Helpers ────────────────────────────────────────────────
    def _fmt_ts(dt):
        return dt.strftime('%d/%m/%Y %H:%M') if dt else ''

    def _resolve_role(booking):
        """Priority: user > approver > admin"""
        if booking.user_id == current_user.id:
            return 'user'
        if booking.trip_department_id in approver_dept_ids:
            return 'approver'
        if is_admin_role:
            return 'admin'
        return None

    def _extract_events(notifs):
        """Map event_key → notification.created_at (asc-sorted input)."""
        ev = {}
        saw_forwarded = False
        for n in notifs:
            icon = n.icon or ''
            cat  = n.category or ''
            msg  = n.message or ''
            if 'fa-calendar-plus' in icon:
                ev.setdefault('booking_created', n.created_at)
            elif 'fa-car' in icon:
                ev.setdefault('admin_assigned', n.created_at)
            elif 'fa-paper-plane' in icon:
                ev.setdefault('forwarded', n.created_at)
                saw_forwarded = True
            elif 'fa-circle-check' in icon and cat == 'status':
                if saw_forwarded:
                    ev.setdefault('approver_approved', n.created_at)
                else:
                    ev.setdefault('admin_approved', n.created_at)
            elif 'fa-circle-xmark' in icon:
                if 'หัวหน้าแผนก' in msg:
                    ev.setdefault('approver_rejected', n.created_at)
                else:
                    ev.setdefault('admin_rejected', n.created_at)
            elif 'fa-link' in icon:
                ev.setdefault('merged', n.created_at)
            elif 'fa-flag-checkered' in icon:
                ev.setdefault('mileage_end_notif', n.created_at)
            elif 'fa-flag' in icon:
                ev.setdefault('mileage_start_notif', n.created_at)
            elif 'fa-sack-dollar' in icon:
                ev.setdefault('budget_deducted', n.created_at)
            elif 'fa-credit-card' in icon and cat == 'payment':
                if 'ยืนยันแล้ว' in msg:
                    ev.setdefault('payment_confirmed', n.created_at)
                else:
                    ev.setdefault('payment_required', n.created_at)
        return ev

    def _budget_label(log):
        budget = log.budget if log else None
        if not budget:
            return ''
        bt = (budget.budget_type.name if budget.budget_type else '').lower()
        type_label = 'งบส่วนกลาง' if bt == 'central' else 'งบส่วนกอง'
        dept_name = budget.department.name if budget.department else ''
        return f'{type_label} - {dept_name}' if dept_name else type_label

    def _plate_of(booking):
        return booking.snap_vehicle_plate or (
            booking.assigned_vehicle.license_plate if booking.assigned_vehicle else ''
        )

    # ── Stage builders ─────────────────────────────────────────
    def _build_user_stages(booking, mileage, log, events):
        stages = []
        is_approved = booking.status == 'approved'
        is_rejected = bool(events.get('admin_rejected') or events.get('approver_rejected'))

        # Stage 0 (fallback): pending / forwarded — แสดงเฉพาะกรณียังไม่ approved และยังไม่ rejected
        # เพื่อให้ booking ทุกสถานะมี stage อย่างน้อย 1 อัน (ไม่ตก fallback timeline)
        if not is_approved and not is_rejected:
            if events.get('forwarded'):
                stages.append({
                    'key': 'pending_approver', 'icon': 'send',
                    'title': 'รอหัวหน้าแผนกอนุมัติ',
                    'desc_main': booking.destination or '',
                    'ts': _fmt_ts(events.get('forwarded')),
                })
            else:
                stages.append({
                    'key': 'pending', 'icon': 'clock',
                    'title': 'รอ Admin พิจารณา',
                    'desc_main': booking.destination or '',
                    'ts': _fmt_ts(events.get('booking_created') or booking.created_at),
                })
        # Stage 1: approved
        if booking.status == 'approved' and booking.updated_at:
            plate = _plate_of(booking)
            desc_main = f'อนุมัติรถ {plate}'.strip() if plate else 'อนุมัติคำขอจองรถ'
            desc_sub = ''
            if booking.trip_group:
                names = []
                for m in mates_by_group.get(booking.trip_group, []):
                    if m.id == booking.id: continue
                    u = mate_users.get(m.user_id)
                    nm = (u.full_name if u else None) or m.contact_name
                    if nm and nm not in names:
                        names.append(nm)
                if names:
                    desc_sub = f'เดินทางร่วมกับ {", ".join(names)}'
            ts = events.get('approver_approved') or events.get('admin_approved') or booking.updated_at
            stages.append({
                'key': 'approved', 'icon': 'check-circle-2',
                'title': 'ได้รับการอนุมัติแล้ว',
                'desc_main': desc_main, 'desc_sub': desc_sub,
                'ts': _fmt_ts(ts),
            })
        # Stage 2: trip_start
        if mileage and mileage.odometer_start is not None:
            stages.append({
                'key': 'trip_start', 'icon': 'play-circle',
                'title': 'เริ่มเดินทาง',
                'desc': f'เริ่มต้นที่ {mileage.odometer_start:,} กม.',
                'ts': _fmt_ts(mileage.actual_start or events.get('mileage_start_notif') or mileage.created_at),
            })
        # Stage 3: trip_end
        if mileage and mileage.odometer_end is not None:
            distance = mileage.odometer_end - (mileage.odometer_start or 0)
            stages.append({
                'key': 'trip_end', 'icon': 'flag',
                'title': 'เดินทางเสร็จสิ้น',
                'desc': f'รวมระยะทาง {distance:,} กม.',
                'ts': _fmt_ts(mileage.actual_end or events.get('mileage_end_notif')),
            })
        # Stage 4: budget
        if log:
            stages.append({
                'key': 'budget', 'icon': 'wallet',
                'title': 'ใช้งบประมาณ',
                'desc_main': f'ใช้ ฿{abs(float(log.change_amount)):,.0f}',
                'desc_sub': f'หักจาก {_budget_label(log)}' if _budget_label(log) else '',
                'ts': _fmt_ts(log.created_at),
            })
        # Stage R (terminal): rejected — แสดงเป็น stage สุดท้ายถ้าถูกปฏิเสธ
        if events.get('admin_rejected'):
            stages.append({
                'key': 'rejected', 'icon': 'x-circle',
                'title': 'ถูกปฏิเสธโดย Admin',
                'desc_main': booking.reject_reason or '',
                'ts': _fmt_ts(events.get('admin_rejected')),
            })
        elif events.get('approver_rejected'):
            stages.append({
                'key': 'rejected', 'icon': 'x-circle',
                'title': 'ถูกปฏิเสธโดยหัวหน้าแผนก',
                'desc_main': booking.reject_reason or '',
                'ts': _fmt_ts(events.get('approver_rejected')),
            })
        return stages

    def _build_admin_stages(booking, mileage, log, events):
        stages = []
        owner = updater_users.get(booking.user_id) if booking.user_id else None
        owner_name = (owner.full_name if owner else '') or 'ไม่ระบุ'

        # Stage 1: created
        ts = events.get('booking_created') or booking.created_at
        stages.append({
            'key': 'created', 'icon': 'inbox',
            'title': 'คำขอเข้ามา',
            'desc_main': f'คำขอจาก {owner_name}',
            'desc_sub': booking.destination or '',
            'ts': _fmt_ts(ts),
        })
        # Stage 2: assigned
        if events.get('admin_assigned') or booking.assigned_vehicle_id:
            plate = _plate_of(booking)
            drv = booking.snap_driver_name or (booking.driver.name if booking.driver else '')
            stages.append({
                'key': 'assigned', 'icon': 'truck',
                'title': 'มอบหมายรถ + คนขับ',
                'desc_main': plate or 'รอกำหนดรถ',
                'desc_sub': f'คนขับ: {drv}' if drv else '',
                'ts': _fmt_ts(events.get('admin_assigned')),
            })
        # Stage 3: decision (admin approve เอง OR forward)
        if events.get('forwarded'):
            stages.append({
                'key': 'forwarded', 'icon': 'send',
                'title': 'ส่งต่อหัวหน้าแผนก',
                'desc_main': booking.trip_department or booking.snap_department_name or '',
                'ts': _fmt_ts(events.get('forwarded')),
            })
        elif events.get('admin_approved'):
            updater = updater_users.get(booking.updated_by) if booking.updated_by else None
            updater_name = (updater.full_name if updater else '') or 'Admin'
            stages.append({
                'key': 'admin_approved', 'icon': 'check-circle-2',
                'title': 'อนุมัติโดย Admin',
                'desc_main': updater_name,
                'ts': _fmt_ts(events.get('admin_approved')),
            })
        elif events.get('admin_rejected'):
            stages.append({
                'key': 'admin_rejected', 'icon': 'x-circle',
                'title': 'ปฏิเสธโดย Admin',
                'desc_main': booking.reject_reason or '',
                'ts': _fmt_ts(events.get('admin_rejected')),
            })
        # Stage 4: approver decision (เฉพาะกรณี waiting_approver → ...)
        if events.get('approver_approved'):
            updater = updater_users.get(booking.updated_by) if booking.updated_by else None
            updater_name = (updater.full_name if updater else '') or 'หัวหน้าแผนก'
            stages.append({
                'key': 'approver_approved', 'icon': 'check-check',
                'title': 'หัวหน้าแผนกอนุมัติ',
                'desc_main': updater_name,
                'ts': _fmt_ts(events.get('approver_approved')),
            })
        elif events.get('approver_rejected'):
            stages.append({
                'key': 'approver_rejected', 'icon': 'x-circle',
                'title': 'หัวหน้าแผนกปฏิเสธ',
                'desc_main': booking.reject_reason or '',
                'ts': _fmt_ts(events.get('approver_rejected')),
            })
        # Stage 5: trip_done
        if mileage and mileage.odometer_end is not None:
            distance = mileage.odometer_end - (mileage.odometer_start or 0)
            fuel = float(mileage.fuel_cost or 0)
            stages.append({
                'key': 'trip_done', 'icon': 'flag',
                'title': 'ทริปเสร็จสิ้น',
                'desc_main': f'ระยะทางรวม {distance:,} กม.',
                'desc_sub': f'ค่าน้ำมัน ฿{fuel:,.0f}' if fuel else '',
                'ts': _fmt_ts(mileage.actual_end),
            })
        # Stage 6: budget
        if log:
            stages.append({
                'key': 'budget', 'icon': 'wallet',
                'title': 'หักงบเสร็จ',
                'desc_main': f'ใช้ ฿{abs(float(log.change_amount)):,.0f}',
                'desc_sub': _budget_label(log),
                'ts': _fmt_ts(log.created_at),
            })
        # Stage 7: payment received (personal เท่านั้น)
        if mileage and mileage.personal_paid_at:
            fuel = float(mileage.fuel_cost or 0)
            stages.append({
                'key': 'payment_received', 'icon': 'coins',
                'title': 'รับเงินจาก User',
                'desc_main': f'{owner_name} ฿{fuel:,.0f}',
                'desc_sub': 'ยืนยันรับเงินแล้ว',
                'ts': _fmt_ts(mileage.personal_paid_at),
            })
        return stages

    def _build_approver_stages(booking, mileage, log, events):
        stages = []
        owner = updater_users.get(booking.user_id) if booking.user_id else None
        owner_name = (owner.full_name if owner else '') or 'ไม่ระบุ'

        # Stage 1: forwarded — emit เสมอ (fallback ใช้ booking.created_at ถ้าไม่มี forwarded event)
        forwarded_ts = events.get('forwarded') or booking.created_at
        stages.append({
            'key': 'forwarded', 'icon': 'send',
            'title': 'ได้รับคำขอ' if events.get('forwarded') else 'รอ Admin ส่งต่อ',
            'desc_main': 'ส่งต่อโดย Admin' if events.get('forwarded') else '',
            'desc_sub': f'{owner_name} · {booking.destination or ""}',
            'ts': _fmt_ts(forwarded_ts),
        })
        # Stage 2: my decision
        if events.get('approver_approved'):
            stages.append({
                'key': 'my_approved', 'icon': 'check-circle-2',
                'title': 'อนุมัติแล้ว',
                'desc_main': '',
                'ts': _fmt_ts(events.get('approver_approved')),
            })
        elif events.get('approver_rejected'):
            stages.append({
                'key': 'my_rejected', 'icon': 'x-circle',
                'title': 'ปฏิเสธแล้ว',
                'desc_main': booking.reject_reason or '',
                'ts': _fmt_ts(events.get('approver_rejected')),
            })
        # Stage 3: trip_done
        if mileage and mileage.odometer_end is not None:
            distance = mileage.odometer_end - (mileage.odometer_start or 0)
            stages.append({
                'key': 'trip_done', 'icon': 'flag',
                'title': 'ทริปเสร็จ',
                'desc_main': f'ระยะทางรวม {distance:,} กม.',
                'ts': _fmt_ts(mileage.actual_end),
            })
        # Stage 4: budget (เฉพาะ dept budget)
        if log and log.budget and log.budget.budget_type \
                and log.budget.budget_type.name.lower() == 'department':
            stages.append({
                'key': 'budget', 'icon': 'wallet',
                'title': 'หักงบแผนก',
                'desc_main': f'ใช้ ฿{abs(float(log.change_amount)):,.0f}',
                'desc_sub': _budget_label(log),
                'ts': _fmt_ts(log.created_at),
            })
        return stages

    # ── Run stage builders ────────────────────────────────────
    for g in groups:
        bid = g['booking_id']
        booking = booking_map.get(bid)
        if not booking:
            continue
        role = _resolve_role(booking)
        if not role:
            continue
        mileage = mileage_map.get(bid)
        log = log_map.get(bid)
        events = _extract_events(notifs_by_booking.get(bid, []))

        if role == 'user':
            stages = _build_user_stages(booking, mileage, log, events)
        elif role == 'approver':
            stages = _build_approver_stages(booking, mileage, log, events)
        else:
            stages = _build_admin_stages(booking, mileage, log, events)

        if stages:
            g['stages'] = stages
            g['role']   = role
            # Synthetic groups → ใช้ stage สุดท้ายเป็น preview/sort key
            if g.get('is_synthetic'):
                last = stages[-1]
                preview_msg = last.get('desc_main') or last.get('desc') or last.get('title', '')
                g['latest'] = {
                    'id': None, 'message': preview_msg,
                    'ntype': 'info', 'category': 'status',
                    'icon': last.get('icon', 'info'),
                    'is_read': True, 'is_sticky': False,
                    'booking_id': bid, 'booking_title': g['booking_title'],
                    'action_url': f'/vehicle/detail/{bid}',
                    'created_at': last.get('ts', ''),
                    'created_rel': '',
                }
                # raw datetime สำหรับ sort — รวบทุก event source แล้วเลือก max
                _candidate_dts = [
                    booking.created_at, booking.updated_at,
                    events.get('booking_created'), events.get('admin_assigned'),
                    events.get('forwarded'),
                    events.get('admin_approved'), events.get('approver_approved'),
                    events.get('admin_rejected'), events.get('approver_rejected'),
                    mileage.actual_start if mileage else None,
                    mileage.actual_end if mileage else None,
                    mileage.personal_paid_at if mileage else None,
                    log.created_at if log else None,
                ]
                _valid_dts = [dt for dt in _candidate_dts if dt is not None]
                g['_sort_dt'] = max(_valid_dts) if _valid_dts else booking.created_at

    # ── Sort + drop group ไม่มี stages (frontend ไม่มี fallback timeline แล้ว) ──
    groups = [g for g in groups if g.get('stages')]
    # sort by raw datetime (latest event/stage on top) — string `DD/MM/YYYY HH:MM` ไม่ sortable lexically
    _EPOCH = datetime(1970, 1, 1)
    groups.sort(key=lambda g: g.get('_sort_dt') or _EPOCH, reverse=True)
    # strip internal raw datetime ก่อน jsonify (Flask jsonify ไม่ serialize datetime default)
    for g in groups:
        g.pop('_sort_dt', None)

    # Badge count for UI (max 30+)
    badge = unread if unread <= 30 else '30+'

    return jsonify({
        'notifications':  [_to_dict(n) for n in notifs],   # flat (backward compat)
        'groups':         groups,
        'sticky':         [_to_dict(n) for n in sticky],
        'loose':          loose_items,
        'unread':         unread,
        'unread_payment': unread_payment,
        'badge':          badge,
    })


@vehicle_bp.route('/api/notifications/read-all', methods=['POST'])
@login_required
def mark_all_read():
    Notification.query.filter_by(user_id=current_user.id, is_read=False)\
        .update({'is_read': True})
    db.session.commit()
    return jsonify({'ok': True})


@vehicle_bp.route('/api/notifications/<int:notif_id>/read', methods=['POST'])
@login_required
def mark_one_read(notif_id):
    n = Notification.query.get_or_404(notif_id)
    if n.user_id == current_user.id:
        # Sticky payment card ห้าม mark-as-read จากการคลิกเฉย ๆ — ต้อง admin confirm ก่อน
        if n.is_sticky and n.category in ('payment', 'payment_admin'):
            return jsonify({'ok': True, 'skipped': 'sticky'})
        n.is_read = True
        db.session.commit()
    return jsonify({'ok': True})


# ─────────────────────────────────────────────
# Payment — User แจ้ง "จ่ายแล้ว" (ยังไม่ใช่ยืนยันจริง)
# ─────────────────────────────────────────────
@vehicle_bp.route('/api/payment/report/<int:mileage_id>', methods=['POST'])
@login_required
def payment_report_paid(mileage_id):
    m = VehicleMileage.query.get_or_404(mileage_id)
    b = m.booking
    if b.user_id != current_user.id:
        return jsonify({'ok': False, 'msg': 'ไม่มีสิทธิ์'}), 403

    if m.personal_status == 1:
        return jsonify({'ok': False, 'msg': 'ชำระแล้ว'}), 400

    m.user_reported_paid = True
    m.user_reported_at   = datetime.now()
    db.session.commit()
    return jsonify({'ok': True, 'msg': 'แจ้งสำเร็จ — รอ Admin ยืนยัน'})


# ─────────────────────────────────────────────
# Payment — User แจ้ง "จ่ายแล้ว" จาก notification (อ้างอิง booking_id)
# ใช้โดย notification panel ที่ไม่รู้ mileage_id
# ─────────────────────────────────────────────
@vehicle_bp.route('/api/payment/report-by-booking/<int:booking_id>', methods=['POST'])
@login_required
def payment_report_paid_by_booking(booking_id):
    b = VehicleBooking.query.get_or_404(booking_id)
    if b.user_id != current_user.id:
        return jsonify({'ok': False, 'msg': 'ไม่มีสิทธิ์'}), 403

    # หา mileage ที่ end แล้ว + personal + ยังไม่ paid — เอาตัวล่าสุด
    m = (VehicleMileage.query
         .filter_by(booking_id=b.id)
         .filter(VehicleMileage.odometer_end.isnot(None))
         .filter((VehicleMileage.personal_status == 0) | (VehicleMileage.personal_status.is_(None)))
         .order_by(VehicleMileage.id.desc())
         .first())
    if not m:
        return jsonify({'ok': False, 'msg': 'ไม่พบรายการค้างชำระ'}), 404

    m.user_reported_paid = True
    m.user_reported_at   = datetime.now()
    db.session.commit()
    return jsonify({'ok': True, 'msg': 'แจ้งสำเร็จ — รอ Admin ยืนยัน', 'mileage_id': m.id})


# ─────────────────────────────────────────────
# Unified Activity History (Phase 10, 2026-05-22)
# รวม 4 service types: vehicle / repair / room / maintenance
# ─────────────────────────────────────────────

# status → (badge tone, dot color suffix, thai label)
_HIST_STATUS_META = {
    'pending':          ('warning', 'amber',  'รออนุมัติ'),
    'waiting_approver': ('warning', 'amber',  'รอหัวหน้าอนุมัติ'),
    'approved':         ('blue',    'blue',   'อนุมัติแล้ว'),
    'rejected':         ('danger',  'red',    'ปฏิเสธ'),
    'in_progress':      ('blue',    'blue',   'ดำเนินการ'),
    'done':             ('success', 'green',  'เสร็จสิ้น'),
    'cancelled':        ('neutral', 'subtle', 'ยกเลิก'),
    'confirmed':        ('blue',    'blue',   'จองแล้ว'),
}

# service_type → (lucide icon, label, create-url endpoint)
_HIST_SERVICE_META = {
    'vehicle':     ('car',       'จองรถ',          'vehicle.index'),
    'repair':      ('wrench',    'แจ้งซ่อม IT',    'repair.index'),
    'room':        ('door-open', 'จองห้องประชุม',  'room.index'),
    'maintenance': ('settings',  'แจ้งซ่อมอาคาร',  'maintenance.index'),
}


def _hist_status(status):
    return _HIST_STATUS_META.get(status, ('neutral', 'subtle', status or '—'))


def _hist_day_label(dt):
    """relative thai label: วันนี้ / เมื่อวาน / N วันก่อน / 'D MMM YYYY' (พ.ศ.)"""
    today = get_bkk_time().date()
    d     = dt.date()
    delta = (today - d).days
    if delta == 0: return 'วันนี้'
    if delta == 1: return 'เมื่อวาน'
    if 1 < delta < 7: return f'{delta} วันก่อน'
    th_months = ['ม.ค.','ก.พ.','มี.ค.','เม.ย.','พ.ค.','มิ.ย.',
                 'ก.ค.','ส.ค.','ก.ย.','ต.ค.','พ.ย.','ธ.ค.']
    return f'{d.day} {th_months[d.month-1]} {d.year + 543}'


def _hist_base_item(prefix, row, *, service_type, title, subtitle,
                    status, occurs_at, meta, detail_url, reject_reason=None):
    tone, dot, label = _hist_status(status)
    ts = row.created_at or occurs_at
    return {
        'id':            f'{prefix}-{row.id}',
        'service_type':  service_type,
        'service_icon':  _HIST_SERVICE_META[service_type][0],
        'service_label': _HIST_SERVICE_META[service_type][1],
        'title':         title,
        'subtitle':      subtitle,
        'status':        status,
        'status_label':  label,
        'status_tone':   tone,
        'status_dot':    dot,
        'timestamp':     ts,
        'occurs_at':     occurs_at,
        'meta':          meta,
        'detail_url':    detail_url,
        'reject_reason': reject_reason,
        'day_key':       ts.strftime('%Y-%m-%d') if ts else '',
        'day_label':     _hist_day_label(ts) if ts else '',
    }


def _vehicle_to_activity(b):
    veh = b.assigned_vehicle
    subtitle = (f'{veh.brand} {veh.model} · {veh.license_plate}' if veh
                else (b.snap_vehicle_plate or 'ยังไม่ได้รับรถ'))
    meta = [
        ('clock', f"{b.start_datetime.strftime('%H:%M')}–{b.end_datetime.strftime('%H:%M')}"),
        ('users', f'{b.passenger_count} คน'),
    ]
    if b.driver:               meta.append(('user-check', b.driver.name))
    elif not b.need_driver:    meta.append(('user',       'ขับเอง'))
    if b.trip_group:           meta.append(('git-branch', f'กลุ่ม {b.trip_group}'))
    return _hist_base_item(
        'veh', b, service_type='vehicle',
        title=b.destination, subtitle=subtitle, status=b.status,
        occurs_at=b.start_datetime, meta=meta,
        detail_url=url_for('vehicle.detail_booking', booking_id=b.id),
        reject_reason=(b.reject_reason if b.status == 'rejected' else None),
    )


def _repair_to_activity(t):
    meta = [('map-pin', t.location), ('tag', t.category)]
    if t.urgency: meta.append(('alert-triangle', t.urgency))
    if t.asset_tag: meta.append(('hash', t.asset_tag))
    return _hist_base_item(
        'rep', t, service_type='repair',
        title=t.subject, subtitle=f'แจ้งซ่อม IT · {t.category}',
        status=t.status, occurs_at=t.created_at, meta=meta,
        detail_url=url_for('repair.edit', id=t.id),
    )


def _maintenance_to_activity(t):
    meta = [('map-pin', t.location), ('tag', t.category)]
    if t.urgency: meta.append(('alert-triangle', t.urgency))
    if t.contact_number: meta.append(('phone', t.contact_number))
    return _hist_base_item(
        'mnt', t, service_type='maintenance',
        title=t.subject, subtitle=f'แจ้งซ่อมอาคาร · {t.category}',
        status=t.status, occurs_at=t.created_at, meta=meta,
        detail_url=url_for('maintenance.edit', id=t.id),
    )


def _room_to_activity(b):
    meta = [
        ('clock', f"{b.start_time.strftime('%d %b · %H:%M')}–{b.end_time.strftime('%H:%M')}"),
        ('map-pin', b.room_name),
    ]
    return _hist_base_item(
        'room', b, service_type='room',
        title=b.title, subtitle=f'ห้องประชุม · {b.room_name}',
        status='confirmed', occurs_at=b.start_time, meta=meta,
        detail_url=url_for('room.index'),
    )


def _collect_user_activities(user_id, *, service_type='', status='', q=''):
    """รวม activity 4 service ของ user → sorted newest-first."""
    items = []
    wants = {service_type} if service_type else {'vehicle','repair','room','maintenance'}

    if 'vehicle' in wants:
        items.extend(_vehicle_to_activity(b)
                     for b in VehicleBooking.query.filter_by(user_id=user_id).all())
    if 'repair' in wants:
        items.extend(_repair_to_activity(t)
                     for t in RepairTicket.query.filter_by(user_id=user_id).all())
    if 'maintenance' in wants:
        items.extend(_maintenance_to_activity(t)
                     for t in MaintenanceTicket.query.filter_by(user_id=user_id).all())
    if 'room' in wants:
        items.extend(_room_to_activity(b)
                     for b in RoomBooking.query.filter_by(user_id=user_id).all())

    if status:
        items = [i for i in items if i['status'] == status]
    if q:
        ql = q.lower().strip()
        items = [i for i in items
                 if ql in (i['title'] or '').lower() or ql in (i['subtitle'] or '').lower()]

    items.sort(key=lambda x: x['timestamp'] or get_bkk_time(), reverse=True)
    return items


def _hist_counts(all_items):
    return {
        'total':       len(all_items),
        'pending':     sum(1 for i in all_items if i['status'] in ('pending','waiting_approver')),
        'in_progress': sum(1 for i in all_items if i['status'] == 'in_progress'),
        'done':        sum(1 for i in all_items if i['status'] in ('approved','done','confirmed')),
        'rejected':    sum(1 for i in all_items if i['status'] in ('rejected','cancelled')),
        'by_type': {
            t: sum(1 for i in all_items if i['service_type'] == t)
            for t in ('vehicle','repair','room','maintenance')
        },
    }


@vehicle_bp.route('/vehicle/history')
@login_required
def booking_history():
    """Unified activity history — vehicle + repair + room + maintenance."""
    filters = {
        'type':   request.args.get('type', ''),
        'status': request.args.get('status', ''),
        'q':      request.args.get('q', ''),
    }
    items     = _collect_user_activities(current_user.id, service_type=filters['type'],
                                          status=filters['status'], q=filters['q'])
    all_items = _collect_user_activities(current_user.id)   # for counts (unfiltered)
    return render_template(
        'vehicle/vehicle_history.html',
        items=items,
        counts=_hist_counts(all_items),
        filters=filters,
        service_meta=_HIST_SERVICE_META,
    )


@vehicle_bp.route('/vehicle/history/feed')
@login_required
def history_feed():
    """JSON feed — client-side filter refetch (no full page reload)."""
    items = _collect_user_activities(
        current_user.id,
        service_type=request.args.get('type', ''),
        status=request.args.get('status', ''),
        q=request.args.get('q', ''),
    )
    def _ser(i):
        d = dict(i)
        d['timestamp'] = i['timestamp'].isoformat() if i['timestamp'] else None
        d['occurs_at'] = i['occurs_at'].isoformat() if i['occurs_at'] else None
        return d
    return jsonify({'items': [_ser(i) for i in items]})


# ─────────────────────────────────────────────
# Admin: จัดการรถและคนขับ
# ─────────────────────────────────────────────
@adminfleet_bp.route('/admin/manage-fleet', methods=['GET', 'POST'])
@login_required
def manage_fleet():
    if not is_vehicle_admin():
        flash('คุณไม่มีสิทธิ์เข้าหน้านี้', 'danger')
        return redirect(url_for('vehicle.index'))

    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'add_vehicle':
            new_vehicle = Vehicle(
                brand         = request.form.get('brand'),
                model         = request.form.get('model'),
                license_plate = request.form.get('license_plate'),
                capacity      = int(request.form.get('capacity')),
                fuel_rate     = float(request.form.get('fuel_rate') or 10)
            )
            db.session.add(new_vehicle)
            db.session.commit()
            flash(f"เพิ่มรถ {new_vehicle.brand} {new_vehicle.model} สำเร็จ!", 'success')

        elif action == 'add_driver':
            new_driver = Driver(
                name      = request.form.get('name'),
                phone     = request.form.get('phone'),
                is_active = bool(request.form.get('is_active')),
                user_id   = request.form.get('user_id') or None
            )
            db.session.add(new_driver)
            db.session.commit()
            flash(f"เพิ่มพนักงานขับรถ {new_driver.name} สำเร็จ!", 'success')

        elif action == 'edit_vehicle':
            vid     = int(request.form.get('vehicle_id'))
            vehicle = Vehicle.query.get_or_404(vid)
            vehicle.brand         = request.form.get('brand')
            vehicle.model         = request.form.get('model')
            vehicle.license_plate = request.form.get('license_plate')
            vehicle.capacity      = int(request.form.get('capacity'))
            vehicle.status        = request.form.get('status', 'active')
            fuel_rate_str = request.form.get('fuel_rate', '').strip()
            if fuel_rate_str:
                vehicle.fuel_rate = float(fuel_rate_str)
            svc_date_str = request.form.get('next_service_date', '').strip()
            vehicle.next_service_date = date.fromisoformat(svc_date_str) if svc_date_str else None
            svc_km_str = request.form.get('next_service_km', '').strip()
            vehicle.next_service_km = int(svc_km_str) if svc_km_str else None
            tax_date_str = request.form.get('tax_due_date', '').strip()
            vehicle.tax_due_date = date.fromisoformat(tax_date_str) if tax_date_str else None
            db.session.commit()
            flash(f"อัปเดตข้อมูลรถ {vehicle.brand} {vehicle.model} สำเร็จ!", 'success')

        elif action == 'delete_vehicle':
            vid     = int(request.form.get('vehicle_id'))
            vehicle = Vehicle.query.get_or_404(vid)
            db.session.delete(vehicle)
            db.session.commit()
            flash('ลบรถออกจากระบบแล้ว', 'success')

        elif action == 'edit_driver':
            did    = int(request.form.get('driver_id'))
            driver = Driver.query.get_or_404(did)
            driver.name      = request.form.get('name')
            driver.phone     = request.form.get('phone')
            driver.is_active = True if request.form.get('is_active') else False
            driver.user_id   = request.form.get('user_id') or None
            db.session.commit()
            flash(f"อัปเดตข้อมูลคนขับ {driver.name} สำเร็จ!", 'success')

        elif action == 'delete_driver':
            did    = int(request.form.get('driver_id'))
            driver = Driver.query.get_or_404(did)
            db.session.delete(driver)
            db.session.commit()
            flash('ลบพนักงานขับรถออกจากระบบแล้ว', 'success')

        # ── CRUD: DeptApprover ──────────────────────────────────
        elif action == 'add_approver':
            uid = int(request.form.get('approver_user_id'))
            did = int(request.form.get('approver_dept_id'))
            exists = DeptApprover.query.filter_by(user_id=uid, dept_id=did).first()
            if exists:
                flash('ผู้อนุมัติคนนี้ถูกเพิ่มในกองนั้นแล้ว', 'warning')
            else:
                db.session.add(DeptApprover(user_id=uid, dept_id=did))
                db.session.commit()
                flash('เพิ่มผู้อนุมัติเรียบร้อยแล้ว', 'success')

        elif action == 'delete_approver':
            aid = int(request.form.get('approver_id'))
            row = DeptApprover.query.get_or_404(aid)
            db.session.delete(row)
            db.session.commit()
            flash('ลบผู้อนุมัติออกจากกองแล้ว', 'success')

        return redirect(url_for('adminfleet.manage_fleet'))

    vehicles  = Vehicle.query.order_by(Vehicle.id).all()
    drivers   = Driver.query.order_by(Driver.id).all()
    users     = User.query.order_by(User.full_name).all()
    depts     = (VehicleDepartment.query
                 .filter_by(is_disable=0)
                 .order_by(VehicleDepartment.name).all())
    approvers = (DeptApprover.query
                 .join(DeptApprover.dept)
                 .order_by(VehicleDepartment.name).all())

    odo_rows = (db.session.query(
            VehicleBooking.assigned_vehicle_id,
            func.max(VehicleMileage.odometer_end))
        .join(VehicleMileage, VehicleMileage.booking_id == VehicleBooking.id)
        .filter(VehicleBooking.assigned_vehicle_id.isnot(None),
                VehicleMileage.odometer_end.isnot(None))
        .group_by(VehicleBooking.assigned_vehicle_id).all())
    vehicle_odometers = {vid: odo for vid, odo in odo_rows}

    now_dt      = datetime.now()
    month_start = now_dt.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    job_rows = (db.session.query(VehicleBooking.driver_id, func.count(VehicleBooking.id))
        .filter(VehicleBooking.driver_id.isnot(None),
                VehicleBooking.start_datetime >= month_start,
                VehicleBooking.status == 'approved')
        .group_by(VehicleBooking.driver_id).all())
    driver_jobs = {did: cnt for did, cnt in job_rows}

    return render_template('vehicle/admin/admin_manage_fleet.html',
                           vehicles=vehicles, drivers=drivers, users=users,
                           depts=depts, approvers=approvers,
                           vehicle_odometers=vehicle_odometers,
                           driver_jobs=driver_jobs,
                           now=datetime.now())


# ─────────────────────────────────────────────
# Admin: หน้าจัดการทริป (รวมทริป / เปลี่ยนรถ)
# ─────────────────────────────────────────────
@vehicle_bp.route('/vehicle/admin')
@login_required
def admin_trips():
    if not is_vehicle_admin():
        flash('คุณไม่มีสิทธิ์เข้าหน้านี้', 'danger')
        return redirect(url_for('vehicle.index'))

    bookings = VehicleBooking.query.order_by(VehicleBooking.created_at.desc()).all()
    vehicles = Vehicle.query.order_by(Vehicle.id).all()
    drivers  = Driver.query.filter_by(is_active=True).order_by(Driver.id).all()

    users_dept = [d.name for d in VehicleDepartment.query
                  .filter_by(is_disable=0).order_by(VehicleDepartment.name).all()]

    now = datetime.utcnow() + timedelta(hours=7)
    budget_rows = VehicleBudget.query.filter_by(year=now.year, month=now.month).all()
    budget_map  = {br.department_id: br for br in budget_rows}

    all_depts = VehicleDepartment.query.filter_by(is_disable=0)\
                    .order_by(VehicleDepartment.name).all()

    central_items = []
    dept_items    = []
    for dept in all_depts:
        br = budget_map.get(dept.id)
        # Match budget_manage: show only depts that have a VehicleBudget row
        # for this month — skip zero-budget placeholders.
        if br is None:
            continue
        total = float(br.budget_amount)
        used  = float(br.used_amount)
        if dept.budget_type.name == 'central':
            central_items.append({
                'key':   dept.name,
                'label': dept.name,
                'total': total,
                'used':  used,
            })
        elif dept.budget_type.name == 'department':
            approver_name = ''
            if br.approver:
                approver_name = br.approver.full_name or br.approver.username
            dept_items.append({
                'key':      dept.name,
                'label':    dept.name,
                'total':    total,
                'used':     used,
                'approver': approver_name,
            })

    fuel_price = FuelPrice.get_for_date(date.today()) or float(SystemConfig.get('fuel_price', 0) or 0)

    return render_template('vehicle/admin/vehicle_admin.html',
                           bookings=bookings,
                           vehicles=vehicles,
                           drivers=drivers,
                           users_dept=users_dept,
                           central_items=central_items,
                           dept_items=dept_items,
                           fuel_price=fuel_price,
                           now=now)


# ─────────────────────────────────────────────
# Admin: แจ้ง Telegram สำหรับ booking ที่อนุมัติแล้ว
# ─────────────────────────────────────────────
@vehicle_bp.route('/vehicle/admin/booking/<int:booking_id>/notify', methods=['POST'])
@login_required
def admin_notify_booking(booking_id):
    if not is_vehicle_admin():
        return jsonify({'ok': False, 'msg': 'ไม่มีสิทธิ์'}), 403
    b = VehicleBooking.query.get_or_404(booking_id)
    if b.status != 'approved':
        return jsonify({'ok': False, 'msg': 'booking ไม่ได้อนุมัติ'}), 400
    notify_approved(b)
    return jsonify({'ok': True})


# Admin: ย้อนสถานะ booking → pending
# ─────────────────────────────────────────────
@vehicle_bp.route('/vehicle/admin/booking/<int:booking_id>/revert', methods=['POST'])
@login_required
def admin_revert_booking(booking_id):
    if not is_vehicle_admin():
        return jsonify({'ok': False, 'msg': 'ไม่มีสิทธิ์'}), 403
    b = VehicleBooking.query.get_or_404(booking_id)
    b.status = 'pending'
    db.session.commit()
    return jsonify({'ok': True})


# ─────────────────────────────────────────────
# Admin: เปลี่ยนสถานะรถ (ส่งซ่อม / เสร็จซ่อม)
# ─────────────────────────────────────────────
@vehicle_bp.route('/vehicle/admin/vehicle/<int:vehicle_id>/repair', methods=['POST'])
@login_required
def admin_vehicle_repair(vehicle_id):
    if not is_vehicle_admin():
        return jsonify({'ok': False, 'msg': 'ไม่มีสิทธิ์'}), 403
    v = Vehicle.query.get_or_404(vehicle_id)
    v.status = 'maintenance'
    v.repair_note = request.form.get('repair_note', '')
    v.repair_started_at = datetime.utcnow() + timedelta(hours=7)
    db.session.commit()
    return jsonify({'ok': True})


@vehicle_bp.route('/vehicle/admin/vehicle/<int:vehicle_id>/fix-done', methods=['POST'])
@login_required
def admin_vehicle_fix_done(vehicle_id):
    if not is_vehicle_admin():
        return jsonify({'ok': False, 'msg': 'ไม่มีสิทธิ์'}), 403
    v = Vehicle.query.get_or_404(vehicle_id)
    v.status = 'active'
    v.repair_note = None
    v.repair_started_at = None
    db.session.commit()
    return jsonify({'ok': True, 'label': f'{v.brand} {v.model} ({v.license_plate})'})


# ─────────────────────────────────────────────
# Admin: Swap รถให้ booking
# ─────────────────────────────────────────────
@vehicle_bp.route('/vehicle/admin/booking/<int:booking_id>/swap', methods=['POST'])
@login_required
def admin_swap_vehicle(booking_id):
    if not is_vehicle_admin():
        return jsonify({'ok': False, 'msg': 'ไม่มีสิทธิ์'}), 403
    b = VehicleBooking.query.get_or_404(booking_id)
    new_vehicle_id = request.form.get('vehicle_id', type=int)
    if not new_vehicle_id:
        return jsonify({'ok': False, 'msg': 'ไม่ได้เลือกรถ'}), 400
    b.assigned_vehicle_id = new_vehicle_id
    db.session.commit()
    v = Vehicle.query.get(new_vehicle_id)
    label = f'{v.brand} {v.model} ({v.license_plate})' if v else ''
    return jsonify({'ok': True, 'label': label})


# ─────────────────────────────────────────────
# Admin: รวมทริป (Merge)
# ─────────────────────────────────────────────
@vehicle_bp.route('/vehicle/admin/merge', methods=['POST'])
@login_required
def admin_merge():
    if not is_vehicle_admin():
        flash('คุณไม่มีสิทธิ์', 'danger')
        return redirect(url_for('vehicle.index'))

    booking_ids         = request.form.getlist('booking_ids')
    assigned_vehicle_id = request.form.get('assigned_vehicle_id')
    driver_id           = request.form.get('driver_id') or None
    trip_group          = request.form.get('trip_group', '').strip()
    expense_type        = request.form.get('expense_type') or None

    import sys
    print(f'[DEBUG admin_merge] booking_ids={booking_ids} vehicle={assigned_vehicle_id} driver={driver_id} trip_group={trip_group!r} expense_type={expense_type}', file=sys.stderr, flush=True)

    if len(booking_ids) < 2:
        print(f'[DEBUG admin_merge] BLOCKED: need >= 2 booking_ids, got {len(booking_ids)}', file=sys.stderr, flush=True)
        return jsonify({'ok': False, 'msg': 'กรุณาเลือกรายการอย่างน้อย 2 รายการเพื่อรวมทริป'}), 400

    if not assigned_vehicle_id:
        print(f'[DEBUG admin_merge] BLOCKED: no assigned_vehicle_id', file=sys.stderr, flush=True)
        return jsonify({'ok': False, 'msg': 'กรุณาเลือกรถที่จะใช้สำหรับทริปนี้'}), 400

    # หมายเหตุ: ไม่บังคับเลือกคนขับตอน merge — สามารถ assign ทีหลังได้

    # สร้างชื่อกลุ่มอัตโนมัติถ้าไม่ได้กรอก
    if not trip_group:
        count  = db.session.query(VehicleBooking.trip_group)\
                           .filter(VehicleBooking.trip_group.isnot(None))\
                           .distinct().count()
        trip_group = f"TRP-{str(count + 1).zfill(3)}"

    # กำหนด status — ถ้างบกอง → waiting_approver
    new_status = 'waiting_approver' if expense_type == 'department' else 'approved'
    print(f'[DEBUG admin_merge] trip_group={trip_group!r} expense_type={expense_type} new_status={new_status} updating {len(booking_ids)} bookings', file=sys.stderr, flush=True)

    # อัปเดตทุก booking ที่เลือก
    for bid in booking_ids:
        booking = VehicleBooking.query.get(int(bid))
        if booking:
            booking.trip_group          = trip_group
            booking.assigned_vehicle_id = int(assigned_vehicle_id)
            booking.status              = new_status
            booking.expense_type        = expense_type
            if driver_id and booking.need_driver:
                booking.driver_id = int(driver_id)
            print(f'[DEBUG admin_merge]   updated booking #{booking.id} → status={new_status}', file=sys.stderr, flush=True)
        else:
            print(f'[DEBUG admin_merge]   booking #{bid} NOT FOUND', file=sys.stderr, flush=True)

    db.session.commit()
    print(f'[DEBUG admin_merge] commit OK', file=sys.stderr, flush=True)

    # แจ้งเตือน (Telegram + In-app) ทุก booking ใน group
    if new_status == 'waiting_approver':
        for bid in booking_ids:
            b = VehicleBooking.query.get(int(bid))
            if b:
                notify_forwarded_to_approver(b)        # Telegram
                _n_merged(b, trip_group)               # In-app Event #7
                _n_forwarded(b)                        # In-app Event #4
    else:
        for bid in booking_ids:
            b = VehicleBooking.query.get(int(bid))
            if b:
                notify_approved(b)                     # Telegram
                _n_merged(b, trip_group)               # In-app Event #7
                _n_admin_approved(b)                   # In-app Event #3
    db.session.commit()
    return jsonify({'ok': True, 'trip_group': trip_group})


# ─────────────────────────────────────────────
# Admin: เปลี่ยนรถ / แก้กลุ่มรายการเดี่ยว
# ─────────────────────────────────────────────
@vehicle_bp.route('/vehicle/admin/assign/<int:booking_id>', methods=['POST'])
@login_required
def admin_assign(booking_id):
    if not is_vehicle_admin():
        flash('คุณไม่มีสิทธิ์', 'danger')
        return redirect(url_for('vehicle.index'))

    booking              = VehicleBooking.query.get_or_404(booking_id)
    assigned_vehicle_id  = request.form.get('assigned_vehicle_id')
    assigned_vehicle2_id = request.form.get('assigned_vehicle2_id') or None
    driver_id            = request.form.get('driver_id') or None
    trip_group           = request.form.get('trip_group', '').strip() or None
    action               = request.form.get('action', 'assign')
    assign_action        = request.form.get('assign_action', 'approve')

    if action == 'ungroup':
        booking.trip_group           = None
        booking.assigned_vehicle_id  = None
        booking.assigned_vehicle2_id = None
        db.session.commit()
        flash(f'นำ #{booking_id} ออกจากกลุ่มทริปแล้ว', 'success')
    else:
        # ── ข้อ 1: ถ้าเป็นทริปร่วม (มี trip_group) ──────────────
        # รถและคนขับสืบทอดจากทริปหลักอัตโนมัติ ไม่ต้องกำหนดใหม่
        is_join_trip = bool(trip_group)

        if not is_join_trip:
            # validate คนขับเฉพาะกรณีไม่ใช่ทริปร่วม
            if booking.need_driver and not driver_id and not booking.driver_id:
                return jsonify({'ok': False, 'msg': f'รายการ #{booking_id} ขอคนขับ กรุณาเลือกคนขับด้วย'}), 400
            # กำหนดรถและคนขับเฉพาะทริปอิสระ
            if assigned_vehicle_id:
                booking.assigned_vehicle_id  = int(assigned_vehicle_id)
            booking.assigned_vehicle2_id = int(assigned_vehicle2_id) if assigned_vehicle2_id else None
            if driver_id:
                booking.driver_id  = int(driver_id)

        booking.trip_group       = trip_group
        booking.expense_type     = request.form.get('expense_type') or None
        booking.central_category = request.form.get('central_category') or None
        trip_dept = request.form.get('trip_department', '').strip()
        booking.trip_department  = trip_dept or booking.user.department or None

        trip_dept = request.form.get('trip_department', '').strip()
        booking.trip_department = trip_dept or booking.user.department or None
        if booking.trip_department:
            dept_obj = VehicleDepartment.query.filter_by(name=booking.trip_department).first()
            if dept_obj:
                booking.trip_department_id = dept_obj.id

        if assign_action == 'reject':
            booking.status = 'rejected'
            booking.reject_reason = request.form.get('reject_reason', '').strip() or None
            db.session.flush()
            # คืนงบถ้าเคยหักแล้ว (no-op ถ้ายังไม่เคยหัก — ปกติ reject ตอน assign จะยังไม่มี mileage)
            budget_svc.refund_for_booking(
                booking,
                note=f'reject by admin {current_user.username} (assign): {booking.reject_reason or "—"}',
            )
            notify_rejected(booking, current_user)                 # Telegram
            _n_rejected(booking, current_user, by_approver=False)  # In-app Event #6
            db.session.commit()
        else:
            # approve — ถ้างบกอง → ส่ง Approver อัตโนมัติ
            if not is_join_trip and (booking.assigned_vehicle_id or booking.driver_id):
                _n_admin_assigned(booking)                         # In-app Event #2
            if booking.expense_type == 'department':
                booking.status = 'waiting_approver'
                db.session.flush()
                notify_forwarded_to_approver(booking)              # Telegram
                _n_forwarded(booking)                              # In-app Event #4
            else:
                booking.status = 'approved'
                db.session.flush()
                notify_approved(booking)                           # Telegram
                _n_admin_approved(booking)                         # In-app Event #3
            db.session.commit()

    return jsonify({'ok': True})

# ─────────────────────────────────────────────
# กรอกไมล์ (admin + superadmin)
# ─────────────────────────────────────────────

@vehicle_bp.route('/vehicle/mileage', methods=['GET', 'POST'])
@login_required
def mileage_log():
    if not is_vehicle_admin():
        flash('คุณไม่มีสิทธิ์เข้าหน้านี้', 'danger')
        return redirect(url_for('vehicle.index'))

    if request.method == 'POST':
        booking_id = int(request.form.get('booking_id'))
        booking    = VehicleBooking.query.get_or_404(booking_id)
        entry_type = request.form.get('entry_type')

        mileage = VehicleMileage.query.filter_by(booking_id=booking_id).first()
        if not mileage:
            mileage = VehicleMileage(booking_id=booking_id, noted_by=current_user.id)
            db.session.add(mileage)

        upload_folder = os.path.join('static', 'uploads', 'mileage')
        os.makedirs(upload_folder, exist_ok=True)

        if entry_type == 'start':
            mileage.odometer_start = int(request.form.get('odometer_start', 0))
            mileage.actual_start   = datetime.strptime(request.form.get('actual_start'), '%Y-%m-%dT%H:%M')
            # รูปหน้าปัดก่อนออก
            img = request.files.get('odometer_start_img')
            if img and img.filename:
                fname = f"{int(time.time())}_start_{secure_filename(img.filename)}"
                img.save(os.path.join(upload_folder, fname))
                mileage.odometer_start_img = fname
            db.session.flush()
            _n_mileage_start(booking, mileage)   # Event #8
            flash(f'บันทึกเลขไมล์ก่อนออก #{booking_id} เรียบร้อย', 'success')

        elif entry_type == 'end':
            submitted_end_mileage = int(request.form.get('odometer_end', 0))
            # 🌟 เช็คว่าเลขไมล์ตอนจบ ต้องมากกว่าเลขไมล์ตอนเริ่ม
            # (ดักเผื่อกรณี mileage.odometer_start มีค่าอยู่แล้ว)
            if mileage.odometer_start is not None and submitted_end_mileage <= mileage.odometer_start:
                flash(f'❌ บันทึกไม่สำเร็จ! เลขไมล์ตอนจบ ({submitted_end_mileage}) ต้องมากกว่าเลขไมล์ตอนเริ่ม ({mileage.odometer_start})', 'danger')
                return redirect(url_for('vehicle.mileage_log')) # เด้งกลับไปให้กรอกใหม่
            
            # ถ้าเลขไมล์ถูกต้อง ค่อยเอาไปใส่ใน object
            mileage.odometer_end = submitted_end_mileage
            mileage.actual_end   = datetime.strptime(request.form.get('actual_end'), '%Y-%m-%dT%H:%M')
            # รูปหน้าปัดหลังกลับ
            img = request.files.get('odometer_end_img')
            if img and img.filename:
                fname = f"{int(time.time())}_end_{secure_filename(img.filename)}"
                img.save(os.path.join(upload_folder, fname))
                mileage.odometer_end_img = fname
            # เติมน้ำมันระหว่างทาง
            mileage.refuel = True if request.form.get('refuel') else False
            if mileage.refuel:
                refuel_amt = request.form.get('refuel_amount', '').strip()
                if refuel_amt:
                    mileage.refuel_amount = float(refuel_amt)
                refuel_img = request.files.get('refuel_img')
                if refuel_img and refuel_img.filename:
                    fname = f"{int(time.time())}_refuel_{secure_filename(refuel_img.filename)}"
                    refuel_img.save(os.path.join(upload_folder, fname))
                    mileage.refuel_img = fname
            # admin กรอกค่าน้ำมัน manual
            fuel = request.form.get('fuel_cost', '').strip()
            if fuel:
                mileage.fuel_cost = float(fuel)
            flash(f'บันทึกเลขไมล์หลังกลับ #{booking_id} เรียบร้อย', 'success')

        # Event #9 — แจ้งเมื่อปิดงาน
        if entry_type == 'end':
            _n_mileage_end(booking, mileage)

        db.session.commit()

        # สร้าง OT record อัตโนมัติเมื่อปิดงาน
        if entry_type == 'end':
            auto_generate_ot(booking, mileage)

        # หักงบอัตโนมัติ (รองรับทั้ง central และ department) + Events #10, #11, #12
        if entry_type == 'end':
            m2         = VehicleMileage.query.filter_by(booking_id=booking_id).first()
            distance   = (m2.odometer_end - m2.odometer_start) if (m2 and m2.odometer_end and m2.odometer_start) else None
            target_date = m2.actual_end.date() if (m2 and m2.actual_end) else date.today()
            fuel_price = FuelPrice.get_for_date(target_date) or float(SystemConfig.get('fuel_price', '40') or 40)
            trip_cost  = float(m2.fuel_cost) if (m2 and m2.fuel_cost and float(m2.fuel_cost) > 0) else \
                         (round((distance / float(booking.assigned_vehicle.fuel_rate)) * fuel_price, 2)
                          if distance and booking.assigned_vehicle and booking.assigned_vehicle.fuel_rate else 0)

            # หักงบ central/department — ผ่าน BudgetService (ledger + idempotent)
            if booking.trip_department and booking.expense_type in ['central', 'department'] and trip_cost > 0:
                # Bug fix: booking.expense_type_id เป็น NULL — ใช้ name lookup แทน
                # - department: ใช้ trip_department_id (กองตัวเอง)
                # - central:    ใช้ central_category → match vehicle_department row
                bt = BudgetType.query.filter_by(name=booking.expense_type).first()
                if booking.expense_type == 'central':
                    central_dept = VehicleDepartment.query.filter_by(name=booking.central_category).first() \
                                   if booking.central_category else None
                    target_dept_id = central_dept.id if central_dept else None
                else:
                    target_dept_id = booking.trip_department_id
                budget = (VehicleBudget.query.filter_by(
                    department_id=target_dept_id,
                    year=target_date.year, month=target_date.month,
                    budget_type_id=bt.id
                ).first()) if (bt and target_dept_id) else None
                if budget:
                    budget_svc.deduct_for_mileage(
                        m2, budget, trip_cost,
                        snap={'distance': distance,
                              'fuel_rate': float(booking.assigned_vehicle.fuel_rate) if booking.assigned_vehicle else None,
                              'fuel_price': fuel_price},
                        note=f'mileage_log booking #{booking.id}',
                    )
                else:
                    _key_label = booking.central_category if booking.expense_type == 'central' else booking.trip_department
                    current_app.logger.warning(
                        '[budget-deduct skip] booking #%s: ไม่พบ VehicleBudget '
                        '(target_dept_id=%s, year=%s, month=%s, expense_type=%s→bt_id=%s, key_label=%s, trip_cost=%s)',
                        booking.id, target_dept_id, target_date.year, target_date.month,
                        booking.expense_type, (bt.id if bt else None), _key_label, trip_cost,
                    )
                    flash(
                        f'⚠️ ปิดทริป #{booking.id} แล้ว แต่ไม่ได้หักงบ '
                        f'(ไม่พบงบ {booking.expense_type} ของ "{_key_label or "—"}" '
                        f'เดือน {target_date.month}/{target_date.year})',
                        'warning'
                    )
                _n_budget(booking, trip_cost, booking.expense_type)   # Event #10 / #11
                db.session.commit()
            elif booking.expense_type in ['central', 'department']:
                # trip_cost==0 หรือ trip_department ว่าง — log silent skip
                current_app.logger.warning(
                    '[budget-deduct skip] booking #%s: ข้ามการหักงบ '
                    '(trip_department=%s, expense_type=%s, trip_cost=%s)',
                    booking.id, booking.trip_department, booking.expense_type, trip_cost,
                )
                if trip_cost == 0:
                    flash(
                        f'⚠️ ปิดทริป #{booking.id} แล้ว แต่ไม่ได้หักงบ '
                        f'(trip_cost = 0 — ตรวจ fuel_cost หรือ vehicle.fuel_rate)',
                        'warning'
                    )

            # Event #12 — ต้องจ่ายส่วนตัว
            elif booking.expense_type == 'personal' and trip_cost > 0:
                _n_payment_required(booking, m2, trip_cost)
                db.session.commit()

        return redirect(url_for('vehicle.mileage_log'))

    # ── GET: Admin mileage dashboard ────────────────────────────
    today      = date.today()
    now        = datetime.now()
    fuel_price = FuelPrice.get_for_date(today) or float(SystemConfig.get('fuel_price', '40') or 40)

    # Filter params
    show_all     = request.args.get('show_all', '') == '1'
    f_date_start = request.args.get('date_start', '').strip()
    f_date_end   = request.args.get('date_end', '').strip()
    f_vehicle    = request.args.get('vehicle_id', type=int)
    f_driver     = request.args.get('driver_id', type=int)
    f_status     = request.args.get('status_filter', '').strip()   # complete|partial|none
    f_cost_min   = request.args.get('cost_min', type=float)
    f_cost_max   = request.args.get('cost_max', type=float)
    f_budget_type = request.args.get('budget_type', '').strip()   # central|department|personal
    f_budget_sub  = request.args.get('budget_sub', '').strip()    # central_category or trip_department
    f_booker      = request.args.get('booker_q', '').strip()

    # Default to current month when no dates given and not show_all
    if not show_all and not f_date_start and not f_date_end:
        f_date_start = today.replace(day=1).strftime('%Y-%m-%d')
        f_date_end   = today.strftime('%Y-%m-%d')

    # Base: approved bookings, past + today (ตัดอนาคต)
    cutoff = datetime.combine(today + timedelta(days=1), datetime.min.time())
    q = VehicleBooking.query.filter(
        VehicleBooking.status == 'approved',
        VehicleBooking.start_datetime < cutoff,
    )
    if f_date_start:
        try:
            q = q.filter(VehicleBooking.start_datetime >= datetime.strptime(f_date_start, '%Y-%m-%d'))
        except ValueError:
            pass
    if f_date_end:
        try:
            end_dt = datetime.strptime(f_date_end, '%Y-%m-%d') + timedelta(days=1)
            q = q.filter(VehicleBooking.start_datetime < end_dt)
        except ValueError:
            pass
    if f_vehicle:
        q = q.filter(VehicleBooking.assigned_vehicle_id == f_vehicle)
    if f_driver:
        q = q.filter(VehicleBooking.driver_id == f_driver)
    if f_budget_type in ('central', 'department', 'personal'):
        q = q.filter(VehicleBooking.expense_type == f_budget_type)
    if f_budget_sub and f_budget_type == 'central':
        q = q.filter(VehicleBooking.central_category == f_budget_sub)
    elif f_budget_sub and f_budget_type == 'department':
        q = q.filter(VehicleBooking.trip_department == f_budget_sub)
    if f_booker:
        like = f'%{f_booker}%'
        q = q.join(User, VehicleBooking.user_id == User.id).filter(
            or_(User.full_name.ilike(like), User.username.ilike(like))
        )

    bookings = q.order_by(VehicleBooking.start_datetime.desc()).all()

    def _compute_cost(b, m):
        """return (distance, fuel_cost, status_key)"""
        if not m:
            return (None, None, 'none')
        if m.odometer_start and m.odometer_end:
            d = m.odometer_end - m.odometer_start
            if m.fuel_cost and float(m.fuel_cost) > 0:
                c = float(m.fuel_cost)
            elif b.assigned_vehicle and b.assigned_vehicle.fuel_rate:
                td = m.actual_end.date() if m.actual_end else b.start_datetime.date()
                fp = FuelPrice.get_for_date(td) or fuel_price
                c = round((d / float(b.assigned_vehicle.fuel_rate)) * fp, 2)
            else:
                c = 0.0
            return (d, c, 'complete')
        if m.odometer_start:
            return (None, None, 'partial')
        return (None, None, 'none')

    # Pre-fetch all FuelBill mileages grouped by vehicle (for in-trip refuel detection)
    fuel_by_vehicle = {}
    for vid, mileage in (db.session.query(FuelBill.vehicle_id, FuelBill.mileage)
                                   .filter(FuelBill.mileage.isnot(None)).all()):
        fuel_by_vehicle.setdefault(vid, []).append(mileage)

    def _budget_info(b):
        et = (b.expense_type or '').strip()
        if et == 'central':
            return ('central', 'งบส่วนกลาง', (b.central_category or '').strip() or None)
        if et == 'department':
            sub = (b.trip_department or (b.user.department if b.user else '') or '').strip()
            return ('department', 'งบส่วนกอง', sub or None)
        if et == 'personal':
            return ('personal', 'งบส่วนตัว', None)
        return ('', '—', None)

    def _has_refuel_in_trip(b, m):
        if not (b.assigned_vehicle_id and m and m.odometer_start and m.odometer_end):
            return False
        bills = fuel_by_vehicle.get(b.assigned_vehicle_id, [])
        lo, hi = m.odometer_start, m.odometer_end
        return any(lo <= km <= hi for km in bills)

    rows = []
    for b in bookings:
        m = b.mileage[0] if b.mileage else None
        distance, fuel_cost, status_key = _compute_cost(b, m)

        # Apply status + cost filter
        if f_status and f_status != status_key:
            continue
        if f_cost_min is not None and (fuel_cost or 0) < f_cost_min:
            continue
        if f_cost_max is not None and (fuel_cost or 0) > f_cost_max:
            continue

        budget_type, budget_label, budget_sub = _budget_info(b)
        rows.append({
            'b': b, 'm': m,
            'distance': distance,
            'fuel_cost': fuel_cost,
            'status_key': status_key,
            'budget_type': budget_type,
            'budget_label': budget_label,
            'budget_sub': budget_sub,
            'has_refuel': _has_refuel_in_trip(b, m),
        })

    # ── Group rows by trip_group (representative = first row) ────
    display_rows = []
    seen_groups = set()
    for r in rows:
        tg = r['b'].trip_group
        if tg is None:
            display_rows.append({'kind': 'single', 'row': r, 'count': 1, 'members': [r]})
            continue
        if tg in seen_groups:
            continue
        seen_groups.add(tg)
        members = [x for x in rows if x['b'].trip_group == tg]
        display_rows.append({'kind': 'group', 'row': members[0], 'count': len(members), 'members': members})

    # ── Dashboard KPIs (year/month) ─────────────────────────────
    year_budgets = VehicleBudget.query.filter_by(year=now.year).all()
    total_budget    = sum(float(bu.budget_amount) for bu in year_budgets)
    total_used      = sum(float(bu.used_amount)   for bu in year_budgets)
    total_remaining = total_budget - total_used

    # Current month cost (all bookings completed in this month)
    month_trips = VehicleBooking.query.filter(
        VehicleBooking.status == 'approved',
        extract('year',  VehicleBooking.start_datetime) == now.year,
        extract('month', VehicleBooking.start_datetime) == now.month,
    ).all()
    month_total_cost = 0.0
    for b in month_trips:
        m = b.mileage[0] if b.mileage else None
        _, c, st = _compute_cost(b, m)
        if st == 'complete' and c:
            month_total_cost += c

    # Pending personal reimbursement
    pending_personal_count = VehicleMileage.query.join(VehicleBooking).filter(
        VehicleBooking.expense_type == 'personal',
        VehicleMileage.personal_status == 0,
        VehicleMileage.odometer_end.isnot(None),
    ).count()

    # Missing mileage (approved + past but no complete record)
    all_past = VehicleBooking.query.filter(
        VehicleBooking.status == 'approved',
        VehicleBooking.start_datetime < cutoff,
    ).all()
    missing_count = 0
    for b in all_past:
        m = b.mileage[0] if b.mileage else None
        _, _, st = _compute_cost(b, m)
        if st != 'complete':
            missing_count += 1

    # ── Per-vehicle × month breakdown (current year) ────────────
    vehicles_all = Vehicle.query.order_by(Vehicle.license_plate).all()
    drivers_all  = Driver.query.filter_by(is_active=True).order_by(Driver.name).all()
    booker_ids   = [uid for (uid,) in db.session.query(VehicleBooking.user_id).distinct().all()]
    bookers_all  = User.query.filter(User.id.in_(booker_ids)).order_by(User.full_name).all() if booker_ids else []

    # Budget sub-list: ดึงเฉพาะค่าที่ปรากฎจริงใน bookings ของหน้านี้
    seen_central = {b.central_category for b in bookings if b.expense_type == 'central' and b.central_category}
    seen_dept    = {b.trip_department  for b in bookings if b.expense_type == 'department' and b.trip_department}
    budget_subs = {
        'central':    [c for c in EXPENSE_CATEGORIES['central']    if c['key'] in seen_central],
        'department': [c for c in EXPENSE_CATEGORIES['department'] if c['key'] in seen_dept],
    }

    breakdown        = {v.id: [0.0]*12 for v in vehicles_all}
    breakdown_totals = [0.0]*12

    year_trips = VehicleBooking.query.filter(
        VehicleBooking.status == 'approved',
        extract('year', VehicleBooking.start_datetime) == now.year,
        VehicleBooking.assigned_vehicle_id.isnot(None),
    ).all()
    for b in year_trips:
        m = b.mileage[0] if b.mileage else None
        _, c, st = _compute_cost(b, m)
        if st != 'complete' or not c:
            continue
        mo_idx = b.start_datetime.month - 1
        if b.assigned_vehicle_id in breakdown:
            breakdown[b.assigned_vehicle_id][mo_idx] += c
            breakdown_totals[mo_idx] += c

    return render_template('vehicle/admin/mileage_admin.html',
        rows=rows,
        display_rows=display_rows,
        fuel_price=fuel_price,
        today=today,
        curr_year=now.year,
        curr_month=now.month,
        # KPIs
        month_total_cost=month_total_cost,
        total_budget=total_budget,
        total_used=total_used,
        total_remaining=total_remaining,
        pending_personal_count=pending_personal_count,
        missing_count=missing_count,
        # Breakdown
        vehicles_all=vehicles_all,
        drivers_all=drivers_all,
        bookers_all=bookers_all,
        budget_subs=budget_subs,
        breakdown=breakdown,
        breakdown_totals=breakdown_totals,
        # Filter echo
        f={'date_start': f_date_start, 'date_end': f_date_end,
           'vehicle_id': f_vehicle or '', 'driver_id': f_driver or '',
           'status_filter': f_status,
           'cost_min': f_cost_min if f_cost_min is not None else '',
           'cost_max': f_cost_max if f_cost_max is not None else '',
           'budget_type': f_budget_type,
           'budget_sub': f_budget_sub,
           'booker_q': f_booker,
           'show_all': show_all},
    )


# ─────────────────────────────────────────────
# Export Excel — mileage (admin)
# ─────────────────────────────────────────────
@vehicle_bp.route('/vehicle/mileage/export')
@login_required
def mileage_export():
    if not is_vehicle_admin():
        flash('คุณไม่มีสิทธิ์', 'danger')
        return redirect(url_for('vehicle.index'))

    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        flash('ไม่พบ openpyxl — รัน: pip install openpyxl', 'danger')
        return redirect(url_for('vehicle.mileage_log'))

    from flask import send_file
    today          = date.today()
    fallback_price = float(SystemConfig.get('fuel_price', '40') or 40)

    f_date_start = request.args.get('date_start', '').strip()
    f_date_end   = request.args.get('date_end', '').strip()
    f_vehicle    = request.args.get('vehicle_id', type=int)
    f_driver     = request.args.get('driver_id', type=int)
    f_status     = request.args.get('status_filter', '').strip()
    f_cost_min   = request.args.get('cost_min', type=float)
    f_cost_max   = request.args.get('cost_max', type=float)

    cutoff = datetime.combine(today + timedelta(days=1), datetime.min.time())
    q = VehicleBooking.query.filter(
        VehicleBooking.status == 'approved',
        VehicleBooking.start_datetime < cutoff,
    )
    if f_date_start:
        try: q = q.filter(VehicleBooking.start_datetime >= datetime.strptime(f_date_start, '%Y-%m-%d'))
        except ValueError: pass
    if f_date_end:
        try: q = q.filter(VehicleBooking.start_datetime < datetime.strptime(f_date_end, '%Y-%m-%d') + timedelta(days=1))
        except ValueError: pass
    if f_vehicle: q = q.filter(VehicleBooking.assigned_vehicle_id == f_vehicle)
    if f_driver:  q = q.filter(VehicleBooking.driver_id == f_driver)
    bookings = q.order_by(VehicleBooking.start_datetime.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Mileage {today.strftime('%Y-%m')}"

    headers = ['Booking','ผู้จอง','รถ','ทะเบียน','คนขับ','ปลายทาง','วันเดินทาง',
               'ไมล์ออก','ไมล์กลับ','ระยะทาง(กม.)','ค่าน้ำมัน(฿)','สถานะ','ประเภท']
    hdr_fill = PatternFill('solid', fgColor='4F46E5')
    hdr_font = Font(bold=True, color='FFFFFF', name='Sarabun')
    thin = Side(style='thin', color='E4E4E7')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border

    ST_LABEL  = {'complete':'ครบ','partial':'รอกลับ','none':'รอกรอก'}
    EXP_LABEL = {'central':'ส่วนกลาง','department':'หน่วยงาน','personal':'ส่วนตัว'}
    total_distance = 0.0
    total_fuel     = 0.0
    ri = 2
    for b in bookings:
        m = b.mileage[0] if b.mileage else None
        distance = fuel_cost = None
        status_key = 'none'
        if m and m.odometer_start and m.odometer_end:
            distance = m.odometer_end - m.odometer_start
            if m.fuel_cost and float(m.fuel_cost) > 0:
                fuel_cost = float(m.fuel_cost)
            elif b.assigned_vehicle and b.assigned_vehicle.fuel_rate:
                td = m.actual_end.date() if m.actual_end else b.start_datetime.date()
                fp = FuelPrice.get_for_date(td) or fallback_price
                fuel_cost = round(distance / float(b.assigned_vehicle.fuel_rate) * fp, 2)
            status_key = 'complete'
        elif m and m.odometer_start:
            status_key = 'partial'

        if f_status and f_status != status_key: continue
        if f_cost_min is not None and (fuel_cost or 0) < f_cost_min: continue
        if f_cost_max is not None and (fuel_cost or 0) > f_cost_max: continue

        if distance: total_distance += distance
        if fuel_cost: total_fuel += fuel_cost

        row_data = [
            f"BK-{b.id}",
            b.user.full_name or b.user.username,
            f"{b.assigned_vehicle.brand} {b.assigned_vehicle.model}" if b.assigned_vehicle else '-',
            b.assigned_vehicle.license_plate if b.assigned_vehicle else '-',
            b.driver.name if b.driver else '-',
            b.destination or '-',
            b.start_datetime.strftime('%d/%m/%Y'),
            m.odometer_start if m and m.odometer_start else None,
            m.odometer_end   if m and m.odometer_end   else None,
            distance,
            round(fuel_cost, 2) if fuel_cost else None,
            ST_LABEL.get(status_key, '-'),
            EXP_LABEL.get(b.expense_type or '', 'ไม่ระบุ'),
        ]
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = border
            c.alignment = Alignment(horizontal='center' if ci in [1,7,8,9,10,11,12] else 'left')
            if ri % 2 == 0:
                c.fill = PatternFill('solid', fgColor='FAFAFA')
        ri += 1

    # Totals row
    tr = ri
    ws.cell(row=tr, column=9,  value='รวม').font = Font(bold=True)
    ws.cell(row=tr, column=10, value=round(total_distance, 2)).font = Font(bold=True)
    ws.cell(row=tr, column=11, value=round(total_fuel, 2)).font = Font(bold=True)

    col_widths = [10,22,18,14,16,28,12,12,12,14,14,12,14]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"mileage_{today.strftime('%Y%m%d')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ─────────────────────────────────────────────
# สรุปค่าใช้จ่าย (admin + superadmin)
# ─────────────────────────────────────────────
def auto_generate_ot(booking, mileage):
    """Auto-generate DriverOT + DriverOTSlots เมื่อปิดงาน (entry_type='end').
    Idempotent — ถ้า DriverOT สำหรับ booking นี้มีอยู่แล้วจะ skip ทันที"""
    if not booking.need_driver or not booking.driver_id:
        return
    if not mileage or not mileage.actual_start or not mileage.actual_end:
        return
    if DriverOT.query.filter_by(booking_id=booking.id).first():
        return  # already generated — idempotent

    rate_configs = OTRateConfig.query.filter_by(is_active=True).order_by(OTRateConfig.sort_order).all()
    if not rate_configs:
        return

    # Per-weekday override: if any rate row targets booking's weekday → use only those.
    # Otherwise fall back to weekday-agnostic rows (day_of_week IS NULL).
    booking_dow = mileage.actual_end.weekday()  # 0=Mon ... 6=Sun
    day_rows = [c for c in rate_configs if c.day_of_week == booking_dow]
    rate_configs = day_rows if day_rows else [c for c in rate_configs if c.day_of_week is None]
    if not rate_configs:
        return

    def to_min(dt):
        return dt.hour * 60 + dt.minute

    trip_s = to_min(mileage.actual_start)
    trip_e = to_min(mileage.actual_end)
    if trip_e <= trip_s:
        return  # invalid same-day end

    new_slots = []
    for cfg in rate_configs:
        h, m   = cfg.start_time.split(':')
        band_s = int(h) * 60 + int(m)
        h, m   = cfg.end_time.split(':')
        band_e = 1440 if cfg.end_time == '24:00' else int(h) * 60 + int(m)

        ov_s = max(trip_s, band_s)
        ov_e = min(trip_e, band_e)
        ov   = max(0, ov_e - ov_s)
        if ov == 0:
            continue

        hrs    = round(ov / 60, 2)
        rate   = float(cfg.rate)
        new_slots.append(DriverOTSlot(
            rate_config_id=cfg.id,
            slot_label=cfg.label,
            start_time=f"{ov_s // 60:02d}:{ov_s % 60:02d}",
            end_time  =f"{ov_e // 60:02d}:{ov_e % 60:02d}",
            hours=hrs, rate=rate,
            amount=round(hrs * rate, 2),
        ))

    if not new_slots:
        return

    yr   = mileage.actual_end.year
    last = DriverOT.query.filter(DriverOT.ot_number.like(f'OT-{yr}-%')) \
                         .order_by(DriverOT.id.desc()).first()
    seq  = (int(last.ot_number.split('-')[-1]) + 1) if last else 1

    ot = DriverOT(
        booking_id   =booking.id,
        driver_id    =booking.driver_id,
        ot_number    =f'OT-{yr}-{seq:04d}',
        date         =mileage.actual_end.date(),
        total_hours  =round(sum(float(s.hours)  for s in new_slots), 2),
        total_amount =round(sum(float(s.amount) for s in new_slots), 2),
        status       ='pending',
        created_at   =get_bkk_time(),
        created_by_id=current_user.id,
    )
    ot.slots = new_slots
    db.session.add(ot)
    db.session.flush()  # ไม่ commit เอง — ให้ caller ที่เรียก commit() ครอบ transaction ไว้


@admincost_bp.route('/vehicle/mileage/override-fuel', methods=['POST'])
@login_required
def override_fuel():
    if not is_vehicle_admin():
        flash('คุณไม่มีสิทธิ์', 'danger')
        return redirect(url_for('vehicle.index'))
    booking_id = int(request.form.get('booking_id'))
    fuel_cost  = float(request.form.get('fuel_cost', 0))
    mileage = VehicleMileage.query.filter_by(booking_id=booking_id).first()
    if not mileage:
        mileage = VehicleMileage(booking_id=booking_id, noted_by=current_user.id)
        db.session.add(mileage)
    mileage.fuel_cost = fuel_cost

    # ถ้า mileage เคยถูกหักงบไปแล้ว → refund เก่า แล้ว deduct ใหม่ด้วยจำนวนใหม่
    if mileage.id and mileage.last_budget_log_id:
        booking = mileage.booking
        target_date = mileage.actual_end.date() if mileage.actual_end else date.today()
        if booking and booking.trip_department and booking.expense_type in ['central', 'department']:
            budget = VehicleBudget.query.filter_by(
                department_id=booking.trip_department_id,
                year=target_date.year, month=target_date.month,
                budget_type_id=booking.expense_type_id
            ).first()
            if budget:
                budget_svc.rededuct_for_mileage(
                    mileage, budget, fuel_cost,
                    snap={'distance': (mileage.odometer_end - mileage.odometer_start)
                          if (mileage.odometer_end and mileage.odometer_start) else None,
                          'fuel_rate': float(booking.assigned_vehicle.fuel_rate) if booking.assigned_vehicle else None,
                          'fuel_price': None},
                    note=f'override_fuel by {current_user.username} → {fuel_cost}',
                )

    db.session.commit()
    flash(f'Override ค่าน้ำมัน #{booking_id} เป็น {fuel_cost:,.2f} บาท เรียบร้อย', 'success')
    return redirect(request.referrer or url_for('admincost.cost_summary'))


TH_MONTHS = ['','ม.ค.','ก.พ.','มี.ค.','เม.ย.','พ.ค.','มิ.ย.','ก.ค.','ส.ค.','ก.ย.','ต.ค.','พ.ย.','ธ.ค.']

@admincost_bp.route('/admin/cost', methods=['GET'])
@login_required
def cost_summary():
    if not is_vehicle_admin():
        flash('คุณไม่มีสิทธิ์เข้าหน้านี้', 'danger')
        return redirect(url_for('vehicle.index'))

    now        = get_bkk_time()
    from_month = int(request.args.get('from_month', now.month))
    from_year  = int(request.args.get('from_year',  now.year))
    to_month   = int(request.args.get('to_month',   now.month))
    to_year    = int(request.args.get('to_year',    now.year))
    sel_driver = request.args.get('driver_id', type=int)
    sel_status = request.args.get('status', '')

    from_date = date(from_year, from_month, 1)
    to_date   = date(to_year + 1, 1, 1) if to_month == 12 else date(to_year, to_month + 1, 1)

    # KPI — query รวมทุก status
    base_q = DriverOT.query.filter(DriverOT.date >= from_date, DriverOT.date < to_date)
    if sel_driver:
        base_q = base_q.filter(DriverOT.driver_id == sel_driver)
    all_ots = base_q.all()

    kpi_records  = len(all_ots)
    kpi_hours    = round(sum(float(o.total_hours)  for o in all_ots), 2)
    kpi_total    = round(sum(float(o.total_amount) for o in all_ots), 2)
    kpi_pending  = round(sum(float(o.total_amount) for o in all_ots if o.status == 'pending'),  2)
    kpi_approved = round(sum(float(o.total_amount) for o in all_ots if o.status == 'approved'), 2)
    kpi_paid     = round(sum(float(o.total_amount) for o in all_ots if o.status == 'paid'),     2)
    count_pending  = sum(1 for o in all_ots if o.status == 'pending')
    count_approved = sum(1 for o in all_ots if o.status == 'approved')
    count_paid     = sum(1 for o in all_ots if o.status == 'paid')

    # Filtered list
    list_q = DriverOT.query.filter(DriverOT.date >= from_date, DriverOT.date < to_date)
    if sel_driver:
        list_q = list_q.filter(DriverOT.driver_id == sel_driver)
    if sel_status:
        list_q = list_q.filter(DriverOT.status == sel_status)
    ots = list_q.order_by(DriverOT.date.desc()).all()

    drivers      = Driver.query.order_by(Driver.name).all()
    rate_configs = OTRateConfig.query.filter_by(is_active=True).order_by(OTRateConfig.sort_order).all()

    range_label = TH_MONTHS[from_month] + ' ' + str(from_year + 543)
    if from_month != to_month or from_year != to_year:
        range_label += f" – {TH_MONTHS[to_month]} {to_year + 543}"

    return render_template('vehicle/admin/vehicle_cost.html',
        ots=ots, drivers=drivers, rate_configs=rate_configs,
        from_month=from_month, from_year=from_year,
        to_month=to_month, to_year=to_year,
        sel_driver=sel_driver, sel_status=sel_status,
        kpi_records=kpi_records, kpi_hours=kpi_hours,
        kpi_total=kpi_total, kpi_pending=kpi_pending,
        kpi_approved=kpi_approved, kpi_paid=kpi_paid,
        count_pending=count_pending, count_approved=count_approved, count_paid=count_paid,
        range_label=range_label, now=now,
    )


@admincost_bp.route('/admin/ot/<int:ot_id>/approve', methods=['POST'])
@login_required
def ot_approve(ot_id):
    if not is_vehicle_admin():
        flash('คุณไม่มีสิทธิ์', 'danger')
        return redirect(url_for('vehicle.index'))
    ot = DriverOT.query.get_or_404(ot_id)
    if ot.status == 'pending':
        ot.status         = 'approved'
        ot.approved_by_id = current_user.id
        ot.approved_at    = get_bkk_time()
        db.session.commit()
        flash(f'อนุมัติ {ot.ot_number} เรียบร้อย', 'success')
    else:
        flash('สถานะไม่ถูกต้องสำหรับการอนุมัติ', 'warning')
    return redirect(request.referrer or url_for('admincost.cost_summary'))


@admincost_bp.route('/admin/ot/<int:ot_id>/mark_paid', methods=['POST'])
@login_required
def ot_mark_paid(ot_id):
    if not is_vehicle_admin():
        flash('คุณไม่มีสิทธิ์', 'danger')
        return redirect(url_for('vehicle.index'))
    ot = DriverOT.query.get_or_404(ot_id)
    if ot.status == 'approved':
        ot.status     = 'paid'
        ot.paid_by_id = current_user.id
        ot.paid_at    = get_bkk_time()
        db.session.commit()
        flash(f'บันทึกการจ่าย {ot.ot_number} เรียบร้อย', 'success')
    else:
        flash('ต้องอนุมัติก่อนจึงจะบันทึกการจ่ายได้', 'warning')
    return redirect(request.referrer or url_for('admincost.cost_summary'))


@admincost_bp.route('/admin/ot/<int:ot_id>/edit', methods=['POST'])
@login_required
def ot_edit(ot_id):
    if not is_vehicle_admin():
        flash('คุณไม่มีสิทธิ์', 'danger')
        return redirect(url_for('vehicle.index'))
    ot = DriverOT.query.get_or_404(ot_id)

    ot.driver_id = int(request.form.get('driver_id', ot.driver_id))
    ot.date      = datetime.strptime(request.form.get('date'), '%Y-%m-%d').date()
    ot.note      = request.form.get('note', '').strip() or None

    slot_labels = request.form.getlist('slot_label[]')
    slot_starts = request.form.getlist('slot_start[]')
    slot_ends   = request.form.getlist('slot_end[]')
    slot_rates  = request.form.getlist('slot_rate[]')
    slot_cfgids = request.form.getlist('slot_cfg_id[]')

    new_slots = []
    for i, label in enumerate(slot_labels):
        try:
            start  = slot_starts[i]; end = slot_ends[i]
            rate   = float(slot_rates[i])
            cfg_id = int(slot_cfgids[i]) if i < len(slot_cfgids) and slot_cfgids[i] else None
            sh, sm = map(int, start.split(':'))
            eh, em = map(int, end.split(':'))
            mins   = max(0, (eh * 60 + em) - (sh * 60 + sm))
            hrs    = round(mins / 60, 2)
            new_slots.append(DriverOTSlot(
                rate_config_id=cfg_id, slot_label=label,
                start_time=start, end_time=end,
                hours=hrs, rate=rate, amount=round(hrs * rate, 2),
            ))
        except (ValueError, IndexError):
            continue

    ot.slots        = new_slots
    ot.total_hours  = round(sum(float(s.hours)  for s in new_slots), 2)
    ot.total_amount = round(sum(float(s.amount) for s in new_slots), 2)
    db.session.commit()
    flash(f'แก้ไข {ot.ot_number} เรียบร้อย', 'success')
    return redirect(request.referrer or url_for('admincost.cost_summary'))


@admincost_bp.route('/admin/ot/<int:ot_id>/delete', methods=['POST'])
@login_required
def ot_delete(ot_id):
    if not is_vehicle_admin():
        flash('คุณไม่มีสิทธิ์', 'danger')
        return redirect(url_for('vehicle.index'))
    ot = DriverOT.query.get_or_404(ot_id)
    ot_num = ot.ot_number
    db.session.delete(ot)
    db.session.commit()
    flash(f'ลบ {ot_num} เรียบร้อย', 'success')
    return redirect(request.referrer or url_for('admincost.cost_summary'))


@admincost_bp.route('/admin/ot/rate_config/update', methods=['POST'])
@login_required
def ot_rate_config_update():
    if not is_vehicle_admin():
        flash('คุณไม่มีสิทธิ์', 'danger')
        return redirect(url_for('vehicle.index'))

    # Soft-delete existing rows the user removed in the modal
    for did in request.form.getlist('cfg_delete[]'):
        if did:
            cfg = OTRateConfig.query.get(int(did))
            if cfg:
                cfg.is_active = False

    # Update existing (cfg_id present) or create new (cfg_id == '')
    max_order = db.session.query(db.func.coalesce(db.func.max(OTRateConfig.sort_order), 0)).scalar()
    for cfg_id, label, start, end, rate, day in zip(
        request.form.getlist('cfg_id[]'),
        request.form.getlist('cfg_label[]'),
        request.form.getlist('cfg_start[]'),
        request.form.getlist('cfg_end[]'),
        request.form.getlist('cfg_rate[]'),
        request.form.getlist('cfg_day[]'),
    ):
        if not label or not start or not end or rate == '':
            continue
        day_val = int(day) if day not in ('', None) else None
        if cfg_id:
            cfg = OTRateConfig.query.get(int(cfg_id))
            if cfg:
                cfg.label = label; cfg.start_time = start
                cfg.end_time = end; cfg.rate = float(rate)
                cfg.day_of_week = day_val
        else:
            max_order += 10
            db.session.add(OTRateConfig(
                label=label, start_time=start, end_time=end,
                rate=float(rate), is_active=True, sort_order=max_order,
                day_of_week=day_val,
            ))
    db.session.commit()
    flash('อัปเดตอัตรา OT เรียบร้อย', 'success')
    return redirect(request.referrer or url_for('admincost.cost_summary'))

# ─────────────────────────────────────────────
# Driver View — หน้าคนขับ (mobile-friendly)
# ─────────────────────────────────────────────
@driver_bp.route('/driver')
@login_required
def driver_home():
    # หา Driver record ที่ผูกกับ user นี้
    driver = Driver.query.filter_by(user_id=current_user.id).first()
    if not driver:
        flash('บัญชีของคุณยังไม่ได้ผูกกับพนักงานขับรถ กรุณาติดต่อ Admin', 'warning')
        return redirect(url_for('vehicle.index'))

    # ดึงทริปที่ approved และคนขับคือตัวเอง
    bookings = VehicleBooking.query.filter(
        VehicleBooking.status == 'approved',
        VehicleBooking.driver_id == driver.id
    ).order_by(VehicleBooking.start_datetime.desc()).all()

    today_start    = datetime.now().replace(hour=0,  minute=0,  second=0,  microsecond=0)
    today_end      = datetime.now().replace(hour=23, minute=59, second=59, microsecond=0)
    tomorrow_start = today_start + timedelta(days=1)
    tomorrow_end   = today_end   + timedelta(days=1)

    # สำหรับ modal "งานนอกระบบ"
    vehicles = Vehicle.query.filter_by(status='active').order_by(Vehicle.id).all()
    users    = User.query.order_by(User.full_name).all()

    return render_template('vehicle/driver_home.html',
                           driver=driver,
                           bookings=bookings,
                           today_start=today_start,
                           today_end=today_end,
                           tomorrow_start=tomorrow_start,
                           tomorrow_end=tomorrow_end,
                           vehicles=vehicles,
                           users=users)


# ─────────────────────────────────────────────
# งานนอกระบบ — driver สร้าง booking เอง (ad-hoc)
# auto-status=approved, driver_id=self, start=now
# end ตอนหลัง driver กรอกไมล์ขากลับ → จบงาน
# expense_type=NULL → admin มาเลือกที่หลัง
# is_ad_hoc=True → ซ่อนจากหน้าปฏิทิน /vehicle
# ─────────────────────────────────────────────
@driver_bp.route('/driver/ad-hoc-trip', methods=['POST'])
@login_required
def driver_ad_hoc_trip():
    driver = Driver.query.filter_by(user_id=current_user.id).first()
    if not driver:
        flash('ไม่พบข้อมูลคนขับ', 'danger')
        return redirect(url_for('driver.driver_home'))

    contact_user_id_raw = request.form.get('contact_user_id', '').strip()
    contact_name_raw    = request.form.get('contact_name', '').strip()
    vehicle_id_raw      = request.form.get('vehicle_id', '').strip()
    destination         = request.form.get('destination', '').strip()

    if not vehicle_id_raw or not destination:
        flash('กรุณาเลือกรถและกรอกปลายทาง', 'warning')
        return redirect(url_for('driver.driver_home'))

    if not contact_user_id_raw and not contact_name_raw:
        flash('กรุณาเลือกผู้จองหรือพิมพ์ชื่อผู้ติดต่อ', 'warning')
        return redirect(url_for('driver.driver_home'))

    vehicle = Vehicle.query.get(int(vehicle_id_raw))
    if not vehicle:
        flash('ไม่พบรถที่เลือก', 'danger')
        return redirect(url_for('driver.driver_home'))

    # contact_user_id ถ้ามี → ใช้ user คนนั้น; ถ้าไม่มี → driver เป็นเจ้าของ + เก็บชื่อ free-text
    if contact_user_id_raw:
        user_id      = int(contact_user_id_raw)
        contact_name = contact_name_raw or None
    else:
        user_id      = current_user.id
        contact_name = contact_name_raw

    now = get_bkk_time()
    end_placeholder = now.replace(hour=23, minute=59, second=0, microsecond=0)

    booking = VehicleBooking(
        user_id             = user_id,
        contact_name        = contact_name,
        start_datetime      = now,
        end_datetime        = end_placeholder,
        destination         = destination,
        purpose             = 'งานนอกระบบ',
        passenger_count     = 1,
        need_driver         = True,
        driver_id           = driver.id,
        assigned_vehicle_id = vehicle.id,
        status              = 'approved',
        is_ad_hoc           = True,
        snap_vehicle_plate  = vehicle.license_plate,
        snap_driver_name    = driver.name,
    )
    db.session.add(booking)
    db.session.flush()
    _n_booking_created(booking)   # แจ้ง admin ให้รู้ว่ามี ad-hoc มาใหม่ → admin มาเลือก expense_type ที่หลัง
    db.session.commit()

    flash(f'สร้างงานนอกระบบเรียบร้อย (BK-{booking.id:04d}) ไปบันทึกเลขไมล์ออกได้เลย', 'success')
    return redirect(url_for('driver.driver_home'))


@driver_bp.route('/driver/mileage', methods=['POST'])
@login_required
def driver_mileage():
    driver = Driver.query.filter_by(user_id=current_user.id).first()
    if not driver:
        flash('ไม่พบข้อมูลคนขับ', 'danger')
        return redirect(url_for('driver.driver_home'))

    booking_id = int(request.form.get('booking_id'))
    booking    = VehicleBooking.query.get_or_404(booking_id)

    # ตรวจสอบว่าทริปนี้เป็นของคนขับคนนี้จริง
    if booking.driver_id != driver.id:
        flash('คุณไม่มีสิทธิ์บันทึกทริปนี้', 'danger')
        return redirect(url_for('driver.driver_home'))

    entry_type = request.form.get('entry_type')
    mileage    = VehicleMileage.query.filter_by(booking_id=booking_id).first()
    if not mileage:
        mileage = VehicleMileage(booking_id=booking_id, noted_by=current_user.id)
        db.session.add(mileage)

    upload_folder = os.path.join('static', 'uploads', 'mileage')
    os.makedirs(upload_folder, exist_ok=True)

    if entry_type == 'start':
        mileage.odometer_start = int(request.form.get('odometer_start', 0))
        mileage.actual_start   = datetime.strptime(request.form.get('actual_start'), '%Y-%m-%dT%H:%M')
        img = request.files.get('odometer_start_img')
        if img and img.filename:
            fname = f"{int(time.time())}_start_{secure_filename(img.filename)}"
            img.save(os.path.join(upload_folder, fname))
            mileage.odometer_start_img = fname
        db.session.flush()
        _n_mileage_start(booking, mileage)   # Event #8
        flash('บันทึกเลขไมล์ก่อนออกเรียบร้อย', 'success')

    elif entry_type == 'end':
        submitted_end_mileage = int(request.form.get('odometer_end', 0))
        # 🌟 เช็คว่าเลขไมล์ตอนจบ ต้องมากกว่าเลขไมล์ตอนเริ่ม
        # (ดักเผื่อกรณี mileage.odometer_start มีค่าอยู่แล้ว)
        if mileage.odometer_start is not None and submitted_end_mileage <= mileage.odometer_start:
            flash(f'❌ บันทึกไม่สำเร็จ! เลขไมล์ตอนจบ ({submitted_end_mileage}) ต้องมากกว่าเลขไมล์ตอนเริ่ม ({mileage.odometer_start})', 'danger')
            return redirect(url_for('driver.driver_home')) # เด้งกลับไปให้กรอกใหม่
        
        # ถ้าเลขไมล์ถูกต้อง ค่อยเอาไปใส่ใน object
        mileage.odometer_end = submitted_end_mileage
        mileage.actual_end   = datetime.strptime(request.form.get('actual_end'), '%Y-%m-%dT%H:%M')
        img = request.files.get('odometer_end_img')
        if img and img.filename:
            fname = f"{int(time.time())}_end_{secure_filename(img.filename)}"
            img.save(os.path.join(upload_folder, fname))
            mileage.odometer_end_img = fname
        mileage.refuel = True if request.form.get('refuel') else False
        if mileage.refuel:
            amt = request.form.get('refuel_amount', '').strip()
            if amt:
                mileage.refuel_amount = float(amt)
            ri = request.files.get('refuel_img')
            if ri and ri.filename:
                fname = f"{int(time.time())}_refuel_{secure_filename(ri.filename)}"
                ri.save(os.path.join(upload_folder, fname))
                mileage.refuel_img = fname
        flash('ปิดงานเรียบร้อย', 'success')

    # Event #9 — แจ้งเมื่อปิดงาน
    if entry_type == 'end':
        _n_mileage_end(booking, mileage)

    db.session.commit()

    # สร้าง OT record อัตโนมัติเมื่อปิดงาน
    if entry_type == 'end':
        auto_generate_ot(booking, mileage)

    # ── หักงบประมาณอัตโนมัติเมื่อปิดงาน + Events #10, #11, #12 ──
    if entry_type == 'end':
        m2       = VehicleMileage.query.filter_by(booking_id=booking_id).first()
        distance = (m2.odometer_end - m2.odometer_start) if (m2 and m2.odometer_end and m2.odometer_start) else None
        target_date = m2.actual_end.date() if (m2 and m2.actual_end) else date.today()
        fuel_price = FuelPrice.get_for_date(target_date) or float(SystemConfig.get('fuel_price', '40') or 40)

        if m2 and m2.fuel_cost and float(m2.fuel_cost) > 0:
            trip_cost = float(m2.fuel_cost)
        elif distance and booking.assigned_vehicle and booking.assigned_vehicle.fuel_rate:
            trip_cost = round((distance / float(booking.assigned_vehicle.fuel_rate)) * fuel_price, 2)
        else:
            trip_cost = 0

        # หัก central/department — ผ่าน BudgetService (ledger + idempotent)
        if booking.trip_department and booking.expense_type in ['central', 'department'] and trip_cost > 0:
            # Bug fix: booking.expense_type_id เป็น NULL — ใช้ name lookup แทน
            # - department: ใช้ trip_department_id (กองตัวเอง)
            # - central:    ใช้ central_category → match vehicle_department row
            bt = BudgetType.query.filter_by(name=booking.expense_type).first()
            if booking.expense_type == 'central':
                central_dept = VehicleDepartment.query.filter_by(name=booking.central_category).first() \
                               if booking.central_category else None
                target_dept_id = central_dept.id if central_dept else None
            else:
                target_dept_id = booking.trip_department_id
            budget = (VehicleBudget.query.filter_by(
                department_id=target_dept_id,
                year=target_date.year, month=target_date.month,
                budget_type_id=bt.id
            ).first()) if (bt and target_dept_id) else None
            if budget:
                budget_svc.deduct_for_mileage(
                    m2, budget, trip_cost,
                    snap={'distance': distance,
                          'fuel_rate': float(booking.assigned_vehicle.fuel_rate) if booking.assigned_vehicle else None,
                          'fuel_price': fuel_price},
                    note=f'driver_mileage booking #{booking.id}',
                )
            else:
                _key_label = booking.central_category if booking.expense_type == 'central' else booking.trip_department
                current_app.logger.warning(
                    '[budget-deduct skip] booking #%s (driver): ไม่พบ VehicleBudget '
                    '(target_dept_id=%s, year=%s, month=%s, expense_type=%s→bt_id=%s, key_label=%s, trip_cost=%s)',
                    booking.id, target_dept_id, target_date.year, target_date.month,
                    booking.expense_type, (bt.id if bt else None), _key_label, trip_cost,
                )
                flash(
                    f'⚠️ ปิดทริปแล้ว แต่ไม่ได้หักงบ '
                    f'(ไม่พบงบ {booking.expense_type} ของ "{_key_label or "—"}" '
                    f'เดือน {target_date.month}/{target_date.year})',
                    'warning'
                )
            _n_budget(booking, trip_cost, booking.expense_type)   # Event #10 / #11
            db.session.commit()
        elif booking.expense_type in ['central', 'department']:
            current_app.logger.warning(
                '[budget-deduct skip] booking #%s (driver): ข้ามการหักงบ '
                '(trip_department=%s, expense_type=%s, trip_cost=%s)',
                booking.id, booking.trip_department, booking.expense_type, trip_cost,
            )
            if trip_cost == 0:
                flash(
                    f'⚠️ ปิดทริปแล้ว แต่ไม่ได้หักงบ '
                    f'(trip_cost = 0 — ตรวจ fuel_cost หรือ vehicle.fuel_rate)',
                    'warning'
                )

        # ต้องจ่ายส่วนตัว
        elif booking.expense_type == 'personal' and trip_cost > 0:
            _n_payment_required(booking, m2, trip_cost)   # Event #12
            db.session.commit()

    return redirect(url_for('driver.driver_home'))


def _fmt_date_th(d):
    """แปลง date เป็นรูปแบบไทย เช่น 1 เม.ย. 68"""
    TH_MON = ['','ม.ค.','ก.พ.','มี.ค.','เม.ย.','พ.ค.','มิ.ย.','ก.ค.','ส.ค.','ก.ย.','ต.ค.','พ.ย.','ธ.ค.']
    return f"{d.day} {TH_MON[d.month]} {str(d.year+543)[2:]}"

# ══════════════════════════════════════════════════════
# Feature 3: Budget Routes
# ══════════════════════════════════════════════════════
@adminfleet_bp.route('/admin/budget', methods=['GET', 'POST'])
@login_required
def budget_manage():
    if not is_vehicle_admin():
        flash('คุณไม่มีสิทธิ์', 'danger')
        return redirect(url_for('vehicle.index'))
 
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'set_budget':
            dept        = request.form.get('department', '').strip()
            year        = int(request.form.get('year'))
            month       = int(request.form.get('month'))
            amount      = float(request.form.get('budget_amount', 0))
            budget_type = request.form.get('budget_type', 'department')
            approver_id = request.form.get('approver_id') or None
            if approver_id:
                approver_id = int(approver_id)

            bt_obj = BudgetType.query.filter_by(name=budget_type).first()
            if not bt_obj:
                flash('ไม่พบประเภทงบ กรุณาตรวจสอบข้อมูล', 'danger')
                return redirect(url_for('adminfleet.budget_manage'))

            # หา VehicleDepartment — central: auto-create ถ้าไม่มี
            dept_obj = VehicleDepartment.query.filter_by(name=dept, budget_type_id=bt_obj.id).first()
            if not dept_obj:
                if budget_type == 'central':
                    dept_obj = VehicleDepartment(name=dept, budget_type_id=bt_obj.id)
                    db.session.add(dept_obj)
                    db.session.flush()
                else:
                    flash('ไม่พบกอง/แผนก กรุณาตรวจสอบข้อมูล', 'danger')
                    return redirect(url_for('adminfleet.budget_manage'))

            budget = VehicleBudget.query.filter_by(
                department_id=dept_obj.id, year=year, month=month, budget_type_id=bt_obj.id
            ).first()

            # parse date range
            start_date_str = request.form.get('start_date', '').strip()
            end_date_str   = request.form.get('end_date', '').strip()
            from datetime import date as date_cls
            start_date = date_cls.fromisoformat(start_date_str) if start_date_str else None
            end_date   = date_cls.fromisoformat(end_date_str)   if end_date_str   else None

            if budget:
                # log การเปลี่ยน budget_amount (ผ่าน BudgetService)
                budget_svc.set_budget_amount(
                    budget, amount,
                    note=f'admin {current_user.username}: update budget {budget_type} {dept} {year}-{month:02d} → {amount}',
                )
                budget.start_date = start_date
                budget.end_date   = end_date
                if budget_type == 'department':
                    budget.approver_id = approver_id
            else:
                budget = VehicleBudget(
                    department_id=dept_obj.id, year=year, month=month,
                    budget_amount=amount, budget_type_id=bt_obj.id,
                    approver_id=approver_id if budget_type == 'department' else None,
                    start_date=start_date, end_date=end_date
                )
                db.session.add(budget)
                db.session.flush()
                # log การสร้าง budget ใหม่
                budget_svc.set_budget_amount(
                    budget, amount,
                    note=f'admin {current_user.username}: create budget {budget_type} {dept} {year}-{month:02d} = {amount}',
                )
            db.session.commit()

            type_label = "ส่วนกลาง" if budget_type == 'central' else "งานกอง"
            flash(f'ตั้งงบ{type_label} "{dept}" เดือน {month}/{year} = {amount:,.0f} บาท เรียบร้อย', 'success')

        elif action == 'top_up':
            try:
                bid    = int(request.form.get('budget_id'))
                delta  = float(request.form.get('delta', 0))
                ntext  = (request.form.get('note') or '').strip()
                if delta <= 0:
                    raise ValueError('top-up ต้องเป็นจำนวนบวก')
                budget = VehicleBudget.query.get_or_404(bid)
                if not budget.is_active:
                    raise ValueError(f'งบ "{budget.department.name}" ถูกปิดใช้งานอยู่ — เปิดใช้งานก่อน')
                new_total = float(budget.budget_amount or 0) + delta
                budget_svc.set_budget_amount(
                    budget, new_total,
                    note=f'top-up +{delta:,.0f} by {current_user.username}'
                         + (f' | {ntext}' if ntext else ''))
                db.session.commit()
                flash(f'เพิ่มงบ "{budget.department.name}" +{delta:,.0f} ฿ เรียบร้อย', 'success')
            except ValueError as e:
                db.session.rollback()
                flash(f'เพิ่มงบไม่สำเร็จ: {e}', 'danger')
            except Exception as e:
                db.session.rollback()
                flash(f'เกิดข้อผิดพลาด: {e}', 'danger')

        elif action == 'manual_adjust':
            try:
                bid   = int(request.form.get('budget_id'))
                delta = float(request.form.get('delta', 0))
                ntext = (request.form.get('note') or '').strip()
                if not ntext:
                    raise ValueError('ต้องระบุเหตุผล (note) สำหรับ manual adjust')
                budget = VehicleBudget.query.get_or_404(bid)
                if not budget.is_active:
                    raise ValueError(f'งบ "{budget.department.name}" ถูกปิดใช้งานอยู่ — เปิดใช้งานก่อน')
                budget_svc.manual_adjust(
                    budget, delta,
                    note=f'manual_adjust by {current_user.username}: {ntext}')
                db.session.commit()
                sign = '+' if delta >= 0 else ''
                flash(f'ปรับยอด "{budget.department.name}" {sign}{delta:,.2f} ฿', 'success')
            except ValueError as e:
                db.session.rollback()
                flash(f'ปรับยอดไม่สำเร็จ: {e}', 'danger')
            except Exception as e:
                db.session.rollback()
                flash(f'เกิดข้อผิดพลาด: {e}', 'danger')

        elif action == 'toggle_active':
            try:
                bid    = int(request.form.get('budget_id'))
                target = request.form.get('to_active') == '1'
                budget = VehicleBudget.query.get_or_404(bid)
                log = budget_svc.set_active(
                    budget, target,
                    note=f'{"เปิด" if target else "ปิด"}ใช้งานโดย {current_user.username}',
                )
                db.session.commit()
                if log is None:
                    flash(f'งบ "{budget.department.name}" อยู่ในสถานะที่ต้องการอยู่แล้ว', 'info')
                elif target:
                    flash(f'เปิดใช้งานงบ "{budget.department.name}" เรียบร้อย', 'success')
                else:
                    flash(f'ปิดใช้งานงบ "{budget.department.name}" — booking ใหม่จะถูกบล็อก', 'warning')
            except Exception as e:
                db.session.rollback()
                flash(f'เปลี่ยนสถานะไม่สำเร็จ: {e}', 'danger')

        elif action == 'refund_booking':
            try:
                bk_id   = int(request.form.get('booking_id'))
                booking = VehicleBooking.query.get_or_404(bk_id)
                refunds = budget_svc.refund_for_booking(
                    booking,
                    note=f'cancel + refund booking #{bk_id} by {current_user.username}')
                if booking.status not in ('rejected', 'cancelled'):
                    booking.status = 'cancelled'
                db.session.commit()
                if refunds:
                    flash(f'ยกเลิก booking #{bk_id} + คืนงบ {len(refunds)} รายการ', 'success')
                else:
                    flash(f'ยกเลิก booking #{bk_id} (ยังไม่เคยหักงบ ไม่ต้องคืน)', 'info')
            except Exception as e:
                db.session.rollback()
                flash(f'ยกเลิก booking ไม่สำเร็จ: {e}', 'danger')

        return redirect(url_for('adminfleet.budget_manage',
                                year=request.form.get('year') or '',
                                month=request.form.get('month') or ''))

    now       = datetime.now()
    sel_year  = int(request.args.get('year', now.year))
    sel_month = int(request.args.get('month', now.month))

    raw_budgets = VehicleBudget.query.filter_by(year=sel_year, month=sel_month)\
                                     .join(VehicleBudget.department)\
                                     .order_by(VehicleDepartment.name).all()

    # ── Pending: bookings approved + expense_type in (central, department) ที่ยังไม่หักงบ
    #    (ใช้ outerjoin VehicleMileage; pending = mileage IS NULL หรือ budget_deducted_at IS NULL)
    pending_q = (VehicleBooking.query
                 .outerjoin(VehicleMileage,
                            VehicleMileage.booking_id == VehicleBooking.id)
                 .filter(VehicleBooking.status == 'approved',
                         VehicleBooking.expense_type.in_(['central', 'department']),
                         or_(VehicleMileage.id.is_(None),
                             VehicleMileage.budget_deducted_at.is_(None)))
                 .order_by(VehicleBooking.start_datetime.desc()))
    pending_bookings = pending_q.all()

    # นับต่อ (department_id, expense_type_id) → match กับ budget row
    pending_count_map = {}
    for pb in pending_bookings:
        if pb.trip_department_id and pb.expense_type_id:
            key = (pb.trip_department_id, pb.expense_type_id)
            pending_count_map[key] = pending_count_map.get(key, 0) + 1

    budgets = []
    for b in raw_budgets:
        pct = round(min(float(b.used_amount) / float(b.budget_amount) * 100, 100), 1) if b.budget_amount > 0 else 0
        pkey = (b.department_id, b.budget_type_id)
        budgets.append({
            'id':            b.id,
            'department':    b.department.name,
            'budget_amount': b.budget_amount,
            'used_amount':   b.used_amount,
            'remaining':     round(float(b.budget_amount) - float(b.used_amount), 2),
            'pct':           pct,
            'budget_type':   b.budget_type.name,
            'approver_id':   b.approver_id,
            'approver_name': (b.approver.full_name or b.approver.username) if b.approver else None,
            'start_date':    b.start_date.isoformat() if b.start_date else '',
            'end_date':      b.end_date.isoformat()   if b.end_date   else '',
            'start_date_th': _fmt_date_th(b.start_date) if b.start_date else '',
            'end_date_th':   _fmt_date_th(b.end_date)   if b.end_date   else '',
            'pending_count': pending_count_map.get(pkey, 0),
            'is_active':     b.is_active,
        })

    central_budgets = [b for b in budgets if b['budget_type'] == 'central']
    dept_budgets    = [b for b in budgets if b['budget_type'] == 'department']

    # KPI summary stats — รวมเฉพาะ active เท่านั้น (inactive ยังแสดงในการ์ดแต่ไม่นับ KPI)
    _active_central = [b for b in central_budgets if b['is_active']]
    _active_dept    = [b for b in dept_budgets    if b['is_active']]
    total_central_budget  = sum(float(b['budget_amount']) for b in _active_central)
    total_dept_budget     = sum(float(b['budget_amount']) for b in _active_dept)
    total_central_used    = sum(float(b['used_amount'])   for b in _active_central)
    total_dept_used       = sum(float(b['used_amount'])   for b in _active_dept)
    total_central_pending = sum(b['pending_count']        for b in _active_central)
    total_dept_pending    = sum(b['pending_count']        for b in _active_dept)

    # งบส่วนตัวที่ได้รับจริง (personal_status=1 ในเดือนที่เลือก)
    personal_mileages = VehicleMileage.query.join(VehicleBooking).filter(
        VehicleBooking.expense_type == 'personal',
        VehicleMileage.personal_status == 1,
        extract('year',  VehicleMileage.personal_paid_at) == sel_year,
        extract('month', VehicleMileage.personal_paid_at) == sel_month,
    ).all()
    fuel_price = float(SystemConfig.get('fuel_price', '40'))
    total_personal_received = 0.0
    for m in personal_mileages:
        if m.fuel_cost:
            total_personal_received += float(m.fuel_cost)
        elif m.odometer_end and m.odometer_start and m.booking.assigned_vehicle:
            dist = m.odometer_end - m.odometer_start
            rate = float(m.booking.assigned_vehicle.fuel_rate or 10)
            total_personal_received += round((dist / rate) * fuel_price, 2)

    # ── ส่วนตัวค้างจ่าย: trip ปิดทริปแล้ว แต่ admin ยังไม่ได้กดรับเงิน
    #    Scope: เดือนที่เลือก (จับคู่กับ KPI อื่น). Trigger จาก actual_end (ทริปปิด)
    personal_unpaid_mileages = VehicleMileage.query.join(VehicleBooking).filter(
        VehicleBooking.expense_type == 'personal',
        VehicleMileage.odometer_end.isnot(None),
        ((VehicleMileage.personal_status == 0) | (VehicleMileage.personal_status.is_(None))),
        extract('year',  VehicleMileage.actual_end) == sel_year,
        extract('month', VehicleMileage.actual_end) == sel_month,
    ).all()
    total_personal_unpaid_amount = 0.0
    for m in personal_unpaid_mileages:
        if m.fuel_cost:
            total_personal_unpaid_amount += float(m.fuel_cost)
        elif m.odometer_end and m.odometer_start and m.booking.assigned_vehicle:
            dist = m.odometer_end - m.odometer_start
            rate = float(m.booking.assigned_vehicle.fuel_rate or 10)
            total_personal_unpaid_amount += round((dist / rate) * fuel_price, 2)

    # ── นับ budget rows ที่ใช้เกินเพดาน (used > cap, active เท่านั้น) สำหรับ critical signal
    over_budget_rows = [b for b in (_active_central + _active_dept)
                        if float(b['used_amount']) > float(b['budget_amount']) > 0]

    kpi = {
        'central_budget':       total_central_budget,
        'dept_budget':          total_dept_budget,
        'total_budget':         total_central_budget + total_dept_budget,
        'central_used':         total_central_used,
        'dept_used':            total_dept_used,
        'total_used':           total_central_used + total_dept_used,
        'central_remaining':    total_central_budget - total_central_used,
        'dept_remaining':       total_dept_budget - total_dept_used,
        'total_remaining':      (total_central_budget + total_dept_budget)
                                - (total_central_used + total_dept_used),
        'central_pending_count': total_central_pending,
        'dept_pending_count':    total_dept_pending,
        'total_pending_count':   total_central_pending + total_dept_pending,
        'personal_received':     total_personal_received,
        # Phase 2 redesign (2026-05-22): new signals สำหรับ summary card footer
        'personal_unpaid_count':  len(personal_unpaid_mileages),
        'personal_unpaid_amount': total_personal_unpaid_amount,
        'over_budget_count':      len(over_budget_rows),
        'pct_of_cap':             (((total_central_used + total_dept_used) /
                                    (total_central_budget + total_dept_budget)) * 100)
                                   if (total_central_budget + total_dept_budget) > 0 else 0,
    }

    # ── Phase 2E (2026-05-22): personal mileage rows สำหรับ section ส่วนตัว
    #    Scope: เดือนที่เลือก (จับคู่ filter), รวมทั้ง paid + unpaid.
    #    Trigger window: paid → personal_paid_at; unpaid → actual_end (วันปิดทริป)
    personal_rows = []
    _personal_all = VehicleMileage.query.join(VehicleBooking).filter(
        VehicleBooking.expense_type == 'personal',
        VehicleMileage.odometer_end.isnot(None),
        or_(
            and_(VehicleMileage.personal_status == 1,
                 extract('year',  VehicleMileage.personal_paid_at) == sel_year,
                 extract('month', VehicleMileage.personal_paid_at) == sel_month),
            and_(or_(VehicleMileage.personal_status == 0,
                     VehicleMileage.personal_status.is_(None)),
                 extract('year',  VehicleMileage.actual_end) == sel_year,
                 extract('month', VehicleMileage.actual_end) == sel_month),
        ),
    ).order_by(VehicleMileage.actual_end.desc()).all()

    for pm in _personal_all:
        if pm.fuel_cost:
            pcost = float(pm.fuel_cost)
        elif pm.odometer_end and pm.odometer_start and pm.booking and pm.booking.assigned_vehicle:
            dist  = pm.odometer_end - pm.odometer_start
            rate  = float(pm.booking.assigned_vehicle.fuel_rate or 10)
            pcost = round((dist / rate) * fuel_price, 2)
        else:
            pcost = 0.0

        bk = pm.booking
        personal_rows.append({
            'mileage_id':   pm.id,
            'booking_id':   bk.id if bk else None,
            'date':         pm.actual_end,
            'user':         (bk.user.full_name or bk.user.username) if (bk and bk.user) else '—',
            'destination':  (bk.destination if bk else '') or '—',
            'fuel_cost':    pcost,
            'is_paid':      (pm.personal_status == 1),
            'paid_at':      pm.personal_paid_at,
        })

    # pending_list สำหรับ refund modal — ตัด field ลงให้พอดี
    pending_list = []
    for pb in pending_bookings:
        m = VehicleMileage.query.filter_by(booking_id=pb.id).first()
        pending_list.append({
            'id':           pb.id,
            'department':   pb.trip_department or '—',
            'expense_type': pb.expense_type or '—',
            'destination':  pb.destination or '—',
            'start':        pb.start_datetime,
            'user':         (pb.user.full_name or pb.user.username) if pb.user else '—',
            'has_mileage':  bool(m),
            'has_deduct':   bool(m and m.budget_deducted_at),
        })

    # แยก datalist ตาม type
    central_dept_names = [cat['label'] for cat in EXPENSE_CATEGORIES['central']]
    dept_dept_names    = [d.name for d in VehicleDepartment.query
                          .filter(VehicleDepartment.is_disable == 0)
                          .join(VehicleDepartment.budget_type)
                          .filter(BudgetType.name == 'department')
                          .order_by(VehicleDepartment.name).all()]

    eligible_approvers = User.query.order_by(User.full_name).all()

    TH_MONTHS = ['','ม.ค.','ก.พ.','มี.ค.','เม.ย.','พ.ค.','มิ.ย.','ก.ค.','ส.ค.','ก.ย.','ต.ค.','พ.ย.','ธ.ค.']

    # ── Phase 7 (2026-05-22) — Pivot งบส่วนกลาง/แผนก × เดือน (ปีงบ Mar→Feb)
    #    fiscal_year_start_ad = ปีที่ "เริ่มเดือน 3"; ถ้า sel_month >= 3 → start = sel_year, else start = sel_year - 1
    fiscal_year_start_ad = sel_year if sel_month >= 3 else sel_year - 1
    pivot = _build_budget_pivot(fiscal_year_start_ad)

    return render_template('vehicle/admin/budget_manage.html',
                           central_budgets=central_budgets,
                           dept_budgets=dept_budgets,
                           central_dept_names=central_dept_names,
                           dept_dept_names=dept_dept_names,
                           eligible_approvers=eligible_approvers,
                           kpi=kpi,
                           pending_list=pending_list,
                           personal_rows=personal_rows,
                           sel_year=sel_year, sel_month=sel_month,
                           month_label=f"{TH_MONTHS[sel_month]} {sel_year+543}",
                           TH_MONTHS=TH_MONTHS,
                           pivot=pivot,
                           fiscal_year_start_ad=fiscal_year_start_ad,
                           now=now)


def _build_budget_pivot(fiscal_year_start_ad):
    """Build fiscal-year (Mar→Feb) pivot for budget_manage page.

    Phase 7 (2026-05-22). Fiscal year = months [3..12] of `fiscal_year_start_ad`
    + months [1..2] of `fiscal_year_start_ad + 1`. Filter `is_active=True` only
    (inactive budgets excluded from pivot per design intent).

    Phase 2 (2026-05-22, redesign continuation): เพิ่ม `personal` row —
    sum fuel_cost ของ VehicleMileage ที่ expense_type='personal' + personal_status=1
    (admin ยืนยันรับเงินแล้ว) ภายใน fiscal year. Aggregate ตาม personal_paid_at.

    Returns dict:
      {
        'central':        { dept_id: { month_num: used_amount } },
        'central_labels': { dept_id: dept_name },
        'central_max':    float,           # max used cell (for heat scale)
        'dept':           { dept_id: { month_num: used_amount } },
        'dept_labels':    { dept_id: dept_name },
        'dept_max':       float,
        'personal':       { month_num: total_received },   # 1 row across fiscal year
        'personal_max':   float,
        'fiscal_months':  [(month, year_ad), ...]   # ordered Mar→Feb (12 tuples)
      }
    """
    fiscal_months = [(m, fiscal_year_start_ad) for m in range(3, 13)] \
                  + [(m, fiscal_year_start_ad + 1) for m in (1, 2)]

    # Build OR(year=Y AND month=M) conditions across 12 (m,y) pairs
    year_month_conds = [and_(VehicleBudget.year == y, VehicleBudget.month == m)
                        for (m, y) in fiscal_months]

    rows = (VehicleBudget.query
            .filter(VehicleBudget.is_active.is_(True))
            .filter(or_(*year_month_conds))
            .join(VehicleBudget.budget_type)
            .join(VehicleBudget.department)
            .all())

    central, dept = {}, {}
    labels_c, labels_d = {}, {}
    for b in rows:
        is_central = (b.budget_type.name == 'central')
        bucket = central if is_central else dept
        labels = labels_c if is_central else labels_d
        did = b.department_id
        if did not in bucket:
            bucket[did] = {}
            labels[did] = b.department.name
        bucket[did][b.month] = float(b.used_amount or 0)

    max_c = max((v for row in central.values() for v in row.values() if v > 0), default=0)
    max_d = max((v for row in dept.values()    for v in row.values() if v > 0), default=0)

    # ── Personal row: aggregate ทุก mileage ที่ admin รับเงินแล้ว (personal_status=1)
    #    ภายใน fiscal year (group by month of personal_paid_at)
    fy_start = datetime(fiscal_year_start_ad,     3, 1)
    fy_end   = datetime(fiscal_year_start_ad + 1, 3, 1)
    personal_mileages = (VehicleMileage.query
                         .join(VehicleBooking)
                         .filter(VehicleBooking.expense_type == 'personal',
                                 VehicleMileage.personal_status == 1,
                                 VehicleMileage.personal_paid_at >= fy_start,
                                 VehicleMileage.personal_paid_at <  fy_end)
                         .all())
    fuel_price = float(SystemConfig.get('fuel_price', '40'))
    personal = {}
    for mi in personal_mileages:
        if not mi.personal_paid_at:
            continue
        if mi.fuel_cost:
            cost = float(mi.fuel_cost)
        elif mi.odometer_end and mi.odometer_start and mi.booking.assigned_vehicle:
            dist = mi.odometer_end - mi.odometer_start
            rate = float(mi.booking.assigned_vehicle.fuel_rate or 10)
            cost = round((dist / rate) * fuel_price, 2)
        else:
            cost = 0.0
        mkey = mi.personal_paid_at.month
        personal[mkey] = personal.get(mkey, 0.0) + cost

    max_p = max((v for v in personal.values() if v > 0), default=0)

    return {
        'central':        central,
        'central_labels': labels_c,
        'central_max':    max_c,
        'dept':           dept,
        'dept_labels':    labels_d,
        'dept_max':       max_d,
        'personal':       personal,
        'personal_max':   max_p,
        'fiscal_months':  fiscal_months,
    }


# ══════════════════════════════════════════════════════
# Feature 3.1: Personal Reimbursement
# ══════════════════════════════════════════════════════
@adminfleet_bp.route('/admin/budget/personal', methods=['GET'])
@login_required
def budget_personal():
    if not is_vehicle_admin():
        flash('คุณไม่มีสิทธิ์', 'danger')
        return redirect(url_for('vehicle.index'))

    now       = datetime.now()
    sel_year  = int(request.args.get('year',  now.year))
    sel_month = int(request.args.get('month', now.month))
    status_filter = request.args.get('status', 'all')  # all | pending | paid

    # ดึง mileage ที่ trip เป็น personal และปิดงานแล้ว (odometer_end มีค่า)
    q = VehicleMileage.query.join(VehicleBooking).filter(
        VehicleBooking.expense_type == 'personal',
        VehicleMileage.odometer_end.isnot(None),
        extract('year',  VehicleMileage.actual_end) == sel_year,
        extract('month', VehicleMileage.actual_end) == sel_month,
    )
    if status_filter == 'pending':
        q = q.filter(VehicleMileage.personal_status == 0)
    elif status_filter == 'paid':
        q = q.filter(VehicleMileage.personal_status == 1)

    mileages = q.order_by(VehicleMileage.actual_end.desc()).all()

    fuel_price = float(SystemConfig.get('fuel_price', '40'))
    rows = []
    total_pending = 0.0
    total_paid    = 0.0
    for m in mileages:
        b = m.booking
        distance = (m.odometer_end - m.odometer_start) if (m.odometer_end and m.odometer_start) else 0
        if m.fuel_cost and float(m.fuel_cost) > 0:
            cost = float(m.fuel_cost)
        elif distance and b.assigned_vehicle and b.assigned_vehicle.fuel_rate:
            cost = round((distance / float(b.assigned_vehicle.fuel_rate)) * fuel_price, 2)
        else:
            cost = 0.0

        if m.personal_status == 0:
            total_pending += cost
        else:
            total_paid += cost

        rows.append({
            'mileage_id':   m.id,
            'booking_id':   b.id,
            'user_name':    b.user.full_name or b.user.username,
            'department':   b.snap_department_name or b.trip_department or '—',
            'destination':  b.destination,
            'actual_end':   m.actual_end,
            'distance':     distance,
            'cost':         cost,
            'status':       m.personal_status,  # 0=pending, 1=paid
            'paid_at':      m.personal_paid_at,
            'paid_by':      (m.personal_paid_by.full_name or m.personal_paid_by.username) if m.personal_paid_by else None,
        })

    TH_MONTHS = ['','ม.ค.','ก.พ.','มี.ค.','เม.ย.','พ.ค.','มิ.ย.','ก.ค.','ส.ค.','ก.ย.','ต.ค.','พ.ย.','ธ.ค.']
    return render_template('vehicle/admin/budget_personal.html',
                           rows=rows,
                           total_pending=total_pending,
                           total_paid=total_paid,
                           sel_year=sel_year, sel_month=sel_month,
                           month_label=f"{TH_MONTHS[sel_month]} {sel_year+543}",
                           status_filter=status_filter,
                           now=now)


@adminfleet_bp.route('/admin/budget/personal/mark_paid', methods=['POST'])
@login_required
def budget_personal_mark_paid():
    if not is_vehicle_admin():
        return jsonify({'ok': False, 'msg': 'ไม่มีสิทธิ์'}), 403

    mileage_id = request.form.get('mileage_id', type=int)
    m = VehicleMileage.query.get_or_404(mileage_id)

    m.personal_status     = 1
    m.personal_paid_at    = datetime.now()
    m.personal_paid_by_id = current_user.id

    # ปิด sticky payment notifications ที่ค้างของ booking นี้ (ทั้งของ user และ admin)
    Notification.query.filter(
        Notification.booking_id == m.booking_id,
        Notification.category.in_(['payment', 'payment_admin']),
        Notification.is_read == False
    ).update({'is_read': True, 'is_sticky': False}, synchronize_session=False)

    # แจ้ง user ว่ายืนยันแล้ว
    _n_payment_confirmed(m.booking, m)
    db.session.commit()

    return jsonify({'ok': True})


@adminfleet_bp.route('/admin/budget/personal/mark_unpaid', methods=['POST'])
@login_required
def budget_personal_mark_unpaid():
    if not is_vehicle_admin():
        return jsonify({'ok': False, 'msg': 'ไม่มีสิทธิ์'}), 403

    mileage_id = request.form.get('mileage_id', type=int)
    m = VehicleMileage.query.get_or_404(mileage_id)

    m.personal_status     = 0
    m.personal_paid_at    = None
    m.personal_paid_by_id = None
    db.session.commit()

    return jsonify({'ok': True})


# ══════════════════════════════════════════════════════
# Feature 4: Vehicle History (API — ใช้ใน manage-fleet)
# ══════════════════════════════════════════════════════
@adminfleet_bp.route('/api/vehicle/<int:vid>/history')
@login_required
def vehicle_history(vid):
    vehicle  = Vehicle.query.get_or_404(vid)
    bookings = VehicleBooking.query.filter(
        VehicleBooking.assigned_vehicle_id == vid,
        VehicleBooking.status == 'approved'
    ).order_by(VehicleBooking.start_datetime.desc()).limit(20).all()

    rows = []
    total_km = 0
    for b in bookings:
        m = b.mileage[0] if b.mileage else None
        dist = (m.odometer_end - m.odometer_start) if (m and m.odometer_end and m.odometer_start) else None
        if dist:
            total_km += dist
        rows.append({
            'id':          b.id,
            'date':        b.start_datetime.strftime('%d/%m/%Y'),
            'destination': b.destination,
            'driver':      b.driver.name if b.driver else '-',
            'distance':    dist,
            'odometer_end': m.odometer_end if m else None,
        })
    return jsonify({'vehicle': f"{vehicle.brand} {vehicle.model}", 'total_km': total_km, 'rows': rows})


# ══════════════════════════════════════════════════════
# Feature 5: Excel Export
# ══════════════════════════════════════════════════════
@admincost_bp.route('/admin/cost/export')
@login_required
def cost_export():
    if not is_vehicle_admin():
        flash('คุณไม่มีสิทธิ์', 'danger')
        return redirect(url_for('vehicle.index'))

    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        flash('ไม่พบ openpyxl — รัน: pip install openpyxl', 'danger')
        return redirect(url_for('admincost.cost_summary'))

    from flask import send_file
    now        = get_bkk_time()
    from_month = int(request.args.get('from_month', now.month))
    from_year  = int(request.args.get('from_year',  now.year))
    to_month   = int(request.args.get('to_month',   now.month))
    to_year    = int(request.args.get('to_year',    now.year))
    sel_driver = request.args.get('driver_id', type=int)
    sel_status = request.args.get('status', '')

    from_date = date(from_year, from_month, 1)
    to_date   = date(to_year + 1, 1, 1) if to_month == 12 else date(to_year, to_month + 1, 1)

    q = DriverOT.query.filter(DriverOT.date >= from_date, DriverOT.date < to_date)
    if sel_driver:
        q = q.filter(DriverOT.driver_id == sel_driver)
    if sel_status:
        q = q.filter(DriverOT.status == sel_status)
    ots = q.order_by(DriverOT.date).all()

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = f"OT {TH_MONTHS[from_month]}{from_year+543}"

    headers  = ['เลขที่','วันที่','คนขับ','Booking','สถานที่','ช่วงเวลา OT','ชม.','ยอด(฿)','สถานะ','หมายเหตุ']
    hdr_fill = PatternFill('solid', fgColor='4F46E5')
    hdr_font = Font(bold=True, color='FFFFFF', name='Sarabun')
    thin     = Side(style='thin', color='E4E4E7')
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)
    ST_LABEL = {'pending':'รออนุมัติ','approved':'อนุมัติแล้ว','paid':'จ่ายแล้ว'}

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = border

    total_hrs = total_amt = 0
    for ri, ot in enumerate(ots, 2):
        slot_str = ' | '.join(
            f"{s.slot_label} {s.start_time}–{s.end_time} ฿{s.rate}" for s in ot.slots
        )
        row_data = [
            ot.ot_number,
            ot.date.strftime('%d/%m/%Y'),
            ot.driver.name if ot.driver else '-',
            ot.booking.id if ot.booking else '-',
            ot.booking.destination if ot.booking else '-',
            slot_str,
            float(ot.total_hours),
            float(ot.total_amount),
            ST_LABEL.get(ot.status, ot.status),
            ot.note or '',
        ]
        total_hrs += float(ot.total_hours)
        total_amt += float(ot.total_amount)
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border    = border
            cell.alignment = Alignment(horizontal='center' if ci in [1,2,7,8,9] else 'left')
            if ri % 2 == 0:
                cell.fill = PatternFill('solid', fgColor='F7F7F8')

    tr = len(ots) + 2
    ws.cell(row=tr, column=6, value='รวม').font = Font(bold=True)
    ws.cell(row=tr, column=7, value=round(total_hrs, 2)).font = Font(bold=True)
    ws.cell(row=tr, column=8, value=round(total_amt, 2)).font = Font(bold=True)

    for ci, w in enumerate([14,12,18,10,24,36,8,12,12,20], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 22

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"driver_ot_{from_year}_{from_month:02d}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ══════════════════════════════════════════════════════
# Feature 6: Vehicle Service/Tax Date Update
# ══════════════════════════════════════════════════════
@adminfleet_bp.route('/admin/manage-fleet/service', methods=['POST'])
@login_required
def update_vehicle_service():
    if not is_vehicle_admin():
        flash('คุณไม่มีสิทธิ์', 'danger')
        return redirect(url_for('adminfleet.manage_fleet'))

    vid     = int(request.form.get('vehicle_id'))
    vehicle = Vehicle.query.get_or_404(vid)

    svc_date = request.form.get('next_service_date', '').strip()
    svc_km   = request.form.get('next_service_km', '').strip()
    tax_date = request.form.get('tax_due_date', '').strip()

    vehicle.next_service_date = date.fromisoformat(svc_date) if svc_date else None
    vehicle.next_service_km   = int(svc_km) if svc_km else None
    vehicle.tax_due_date      = date.fromisoformat(tax_date) if tax_date else None
    db.session.commit()
    flash(f'อัปเดตวันนัดซ่อม/ต่อภาษี {vehicle.brand} {vehicle.model} เรียบร้อย', 'success')
    return redirect(url_for('adminfleet.manage_fleet'))

# ══════════════════════════════════════════════════════
# Feature 7: Merge Validation API (สำหรับ Drag & Drop Merge)
# ══════════════════════════════════════════════════════
@vehicle_bp.route('/api/check-merge', methods=['POST'])
@login_required
def api_check_merge():
    """
    ตรวจสอบ validity ก่อน merge จริง
    รับ: { booking_ids: [1,2,...], vehicle_id: 3 (optional) }
    คืน: { valid, errors: [], warnings: [], merged_range, total_pax, destinations }
    """
    if not is_vehicle_admin():
        return jsonify({'error': 'ไม่มีสิทธิ์'}), 403

    data        = request.get_json(force=True)
    booking_ids = [int(x) for x in data.get('booking_ids', [])]
    vehicle_id  = data.get('vehicle_id')

    if len(booking_ids) < 2:
        return jsonify({'valid': False, 'errors': ['ต้องเลือกอย่างน้อย 2 รายการ'], 'warnings': []}), 200

    bookings = [VehicleBooking.query.get(bid) for bid in booking_ids]
    bookings = [b for b in bookings if b]

    errors   = []
    warnings = []

    # ── คำนวณ merged time range ──────────────────────────────
    starts   = [b.start_datetime for b in bookings]
    ends     = [b.end_datetime   for b in bookings]
    merged_start = min(starts)
    merged_end   = max(ends)

    # ── ตรวจสอบ time overlap (warning ถ้าไม่ overlap) ────────
    # หา overlap จริง: ถ้าทุก pair overlap กัน → ok, ไม่งั้น warning
    has_overlap = False
    for i in range(len(bookings)):
        for j in range(i + 1, len(bookings)):
            a, b2 = bookings[i], bookings[j]
            if a.start_datetime < b2.end_datetime and a.end_datetime > b2.start_datetime:
                has_overlap = True
    if not has_overlap:
        warnings.append('ช่วงเวลาของรายการที่เลือกไม่ทับซ้อนกัน การรวมทริปอาจไม่สมเหตุสมผล')

    # ── ตรวจสอบปลายทาง (warning ถ้าต่างกัน) ─────────────────
    destinations = list(dict.fromkeys(b.destination for b in bookings))  # unique + ordered
    if len(destinations) > 1:
        warnings.append(f'ปลายทางต่างกัน: {", ".join(destinations)}')

    # ── รวม passengers ───────────────────────────────────────
    total_pax = sum(b.passenger_count for b in bookings)

    # ── ตรวจสอบรถ (ถ้าเลือกมาแล้ว) ───────────────────────────
    if vehicle_id:
        vehicle = Vehicle.query.get(int(vehicle_id))
        if vehicle:
            # capacity check
            if total_pax > vehicle.capacity:
                errors.append(
                    f'จำนวนผู้โดยสารรวม ({total_pax} คน) เกินความจุรถ {vehicle.brand} {vehicle.model} ({vehicle.capacity} ที่นั่ง)'
                )
            # overlap check — รถคันนี้ถูกใช้ในช่วงเวลาที่ merge แล้วไหม?
            conflict = VehicleBooking.query.filter(
                VehicleBooking.assigned_vehicle_id == vehicle.id,
                VehicleBooking.status.in_(['approved', 'waiting_approver']),
                VehicleBooking.id.notin_(booking_ids),
                VehicleBooking.start_datetime < merged_end,
                VehicleBooking.end_datetime   > merged_start,
            ).first()
            if conflict:
                errors.append(
                    f'รถ {vehicle.brand} {vehicle.model} ({vehicle.license_plate}) '
                    f'มีการจองทับซ้อนในช่วงเวลานี้ (#{conflict.id})'
                )

    # ── ตรวจสอบ driver (ถ้าเลือกมาแล้ว) ──────────────────────
    driver_id = data.get('driver_id')
    if driver_id:
        conflict_d = VehicleBooking.query.filter(
            VehicleBooking.driver_id == int(driver_id),
            VehicleBooking.status.in_(['approved', 'waiting_approver']),
            VehicleBooking.id.notin_(booking_ids),
            VehicleBooking.start_datetime < merged_end,
            VehicleBooking.end_datetime   > merged_start,
        ).first()
        if conflict_d:
            driver = Driver.query.get(int(driver_id))
            dname  = driver.name if driver else f'#{driver_id}'
            errors.append(
                f'คนขับ {dname} มีทริปอื่นทับซ้อนในช่วงเวลานี้ (#{conflict_d.id})'
            )

    # ── Format response ───────────────────────────────────────
    TH_MONTHS = ['','ม.ค.','ก.พ.','มี.ค.','เม.ย.','พ.ค.','มิ.ย.',
                 'ก.ค.','ส.ค.','ก.ย.','ต.ค.','พ.ย.','ธ.ค.']

    def fmt_dt(dt):
        return f"{dt.day} {TH_MONTHS[dt.month]} {dt.year + 543}  {dt.strftime('%H:%M')}"

    booking_summaries = []
    for b in bookings:
        booking_summaries.append({
            'id':         b.id,
            'booker':     b.user.full_name or b.user.username,
            'dept':       b.user.department or '–',
            'dest':       b.destination,
            'pax':        b.passenger_count,
            'start':      fmt_dt(b.start_datetime),
            'end':        fmt_dt(b.end_datetime),
            'need_driver': b.need_driver,
        })

    return jsonify({
        'valid':        len(errors) == 0,
        'errors':       errors,
        'warnings':     warnings,
        'merged_start': fmt_dt(merged_start),
        'merged_end':   fmt_dt(merged_end),
        'total_pax':    total_pax,
        'destinations': destinations,
        'bookings':     booking_summaries,
    })


# ── Booking data สำหรับ Admin page (JSON) ─────────────────
@vehicle_bp.route('/api/admin/bookings')
@login_required
def api_admin_bookings():
    """คืน booking list ล่าสุดสำหรับ admin page (ใช้ refresh หลัง merge)"""
    if not is_vehicle_admin():
        return jsonify({'error': 'ไม่มีสิทธิ์'}), 403

    bookings = VehicleBooking.query.order_by(VehicleBooking.created_at.desc()).all()
    result   = []
    for b in bookings:
        result.append({
            'id':         b.id,
            'booker':     b.user.full_name or b.user.username,
            'dept':       b.user.department or '–',
            'dest':       b.destination,
            'purpose':    b.purpose,
            'pax':        b.passenger_count,
            'start':      b.start_datetime.strftime('%d/%m/%y %H:%M'),
            'end':        b.end_datetime.strftime('%H:%M'),
            'start_iso':  b.start_datetime.isoformat(),
            'end_iso':    b.end_datetime.isoformat(),
            'status':     b.status,
            'trip_group': b.trip_group or '',
            'vehicle':    (f"{b.assigned_vehicle.brand} {b.assigned_vehicle.model} ({b.assigned_vehicle.license_plate})" if b.assigned_vehicle else ''),
            'driver':     (b.driver.name if b.driver else ''),
            'need_driver': b.need_driver,
            'detail_url': f'/vehicle/detail/{b.id}',
            'assign_url': f'/vehicle/admin/assign/{b.id}',
            'ungroup_url': f'/vehicle/admin/assign/{b.id}',
        })
    return jsonify(result)