"""
Fuel management blueprint — /admin/fuel
=======================================
Vehicle admin records fuel bills paid to drivers, batches them into
reimbursement claims, and tracks the cash reserve (เงินสำรอง).

Status (computed, not stored):
    รอเบิก   = bill.reimbursement_id IS NULL
    อนุมัติ  = reimbursement_id NOT NULL AND received_at IS NULL
    ได้เงิน  = received_at NOT NULL

Phase 1: DB + backend routes (UI = phase 2, Excel/PDF = phase 4).
"""
from datetime import datetime, date
from decimal import Decimal
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, abort
from flask_login import login_required, current_user
from sqlalchemy import extract, func
from models import (
    db, SystemConfig,
    Vehicle, Driver,
    FuelBill, FuelReimbursement, FuelPrice,
    FuelReserveConfig, FuelReserveLog,
)

fuel_bp = Blueprint('fuel', __name__)


# ─────────────────────────────────────────────
# Permission
# ─────────────────────────────────────────────
def is_vehicle_admin():
    return current_user.role_vehicle == 'admin' or current_user.is_superadmin


def _guard():
    if not current_user.is_authenticated or not is_vehicle_admin():
        abort(403)


# ─────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────
PAYMENT_METHODS = ('transfer', 'card', 'self')   # เงินสด / ตัดบัตร / จ่ายเอง
PAYMENT_LABEL_TH = {'transfer': 'เงินสด', 'card': 'ตัดบัตร', 'self': 'จ่ายเอง'}
MONTH_LABEL_TH = ['ม.ค.','ก.พ.','มี.ค.','เม.ย.','พ.ค.','มิ.ย.',
                  'ก.ค.','ส.ค.','ก.ย.','ต.ค.','พ.ย.','ธ.ค.']


def _bill_status(bill):
    """รอเบิก | อนุมัติ | ได้เงิน — derived from FK + received_at."""
    if bill.reimbursement_id is None:
        return 'รอเบิก'
    if bill.reimbursement and bill.reimbursement.received_at is None:
        return 'อนุมัติ'
    return 'ได้เงิน'


def _depletes_reserve(method):
    """Only `transfer` (เงินสด — เบิกจากกองกลาง) depletes reserve.
    `card` = company card (no reserve impact). `self` = ผู้โดยสารจ่ายเอง (เก็บประวัติเฉย ๆ)."""
    return method == 'transfer'


def _parse_date(s, default=None):
    if not s:
        return default
    try:
        return datetime.strptime(s, '%Y-%m-%d').date()
    except ValueError:
        return default


def _parse_int(s, default=None):
    try:
        return int(s) if s not in (None, '') else default
    except (ValueError, TypeError):
        return default


def _parse_decimal(s, default=Decimal('0')):
    try:
        return Decimal(str(s)) if s not in (None, '') else default
    except Exception:
        return default


def _read_filters(args):
    """Return (year, month, vehicle_id, driver_id) parsed from query string."""
    today = date.today()
    return (
        _parse_int(args.get('year'),  today.year),
        _parse_int(args.get('month'), 0),
        _parse_int(args.get('vehicle_id')),
        _parse_int(args.get('driver_id')),
    )


def _filtered_bills_query(year, month, vehicle_id, driver_id):
    """Build the FuelBill query used by both admin_fuel() and exports."""
    q = FuelBill.query.filter(extract('year', FuelBill.bill_date) == year)
    if month:      q = q.filter(extract('month', FuelBill.bill_date) == month)
    if vehicle_id: q = q.filter(FuelBill.vehicle_id == vehicle_id)
    if driver_id:  q = q.filter(FuelBill.driver_id == driver_id)
    return q.order_by(FuelBill.bill_date.desc(), FuelBill.id.desc())


# ─────────────────────────────────────────────
# Main page — GET /admin/fuel
# ─────────────────────────────────────────────
@fuel_bp.route('/admin/fuel', methods=['GET'])
@login_required
def admin_fuel():
    _guard()

    # Filters
    f_year, f_month, f_veh, f_drv = _read_filters(request.args)

    # Bill query (filtered)
    bills = _filtered_bills_query(f_year, f_month, f_veh, f_drv).all()

    # Reimbursements (visible separately; shows all in current year regardless of bill filter)
    reimbursements = (FuelReimbursement.query
                      .filter(
                          (extract('year', FuelReimbursement.submitted_at) == f_year) |
                          (extract('year', FuelReimbursement.created_at) == f_year)
                      )
                      .order_by(FuelReimbursement.id.desc())
                      .all())

    # ── KPI: cash reserve side ─────────────────────────────────────
    reserve_amount = FuelReserveConfig.get_amount()

    # All bills not yet "ได้เงิน" (regardless of filter — reserve is global)
    pending_q = FuelBill.query.outerjoin(FuelReimbursement,
                                         FuelBill.reimbursement_id == FuelReimbursement.id)
    pending_q = pending_q.filter(
        (FuelBill.reimbursement_id.is_(None)) |
        (FuelReimbursement.received_at.is_(None))
    )
    pending_bills = pending_q.all()

    pending_total       = sum((float(b.amount) for b in pending_bills), 0.0)
    reserve_used        = sum((float(b.amount) for b in pending_bills if _depletes_reserve(b.payment_method)), 0.0)
    reserve_balance     = float(reserve_amount) - reserve_used

    # "จ่ายเอง" tracker — ผู้โดยสารจ่ายเอง ไม่หัก reserve · pending scope
    self_paid_bills  = [b for b in pending_bills if b.payment_method == 'self']
    self_paid_total  = sum((float(b.amount) for b in self_paid_bills), 0.0)
    self_paid_count  = len(self_paid_bills)

    # ── KPI: annual budget side (sum across all bills in f_year, no other filter) ──
    year_bills = FuelBill.query.filter(extract('year', FuelBill.bill_date) == f_year).all()
    year_used  = sum((float(b.amount) for b in year_bills), 0.0)
    annual_budget = float(SystemConfig.get('fuel_annual_budget', 0) or 0)
    year_remaining = annual_budget - year_used

    # ── KPI: by payment method (current filter scope) ──
    by_method = {'transfer': 0.0, 'card': 0.0, 'self': 0.0}
    for b in bills:
        by_method[b.payment_method] = by_method.get(b.payment_method, 0.0) + float(b.amount)

    # Pending breakdown for status chips — only bills that deplete reserve (transfer)
    pending_by_status = {'รอเบิก': 0.0, 'อนุมัติ': 0.0}
    for b in pending_bills:
        if not _depletes_reserve(b.payment_method):
            continue
        pending_by_status[_bill_status(b)] = pending_by_status.get(_bill_status(b), 0.0) + float(b.amount)

    # ── Pivot: vehicle × month for f_year ──────────────────────────
    pivot_rows = (db.session.query(
                      FuelBill.vehicle_id,
                      extract('month', FuelBill.bill_date).label('m'),
                      func.sum(FuelBill.amount).label('total'))
                  .filter(extract('year', FuelBill.bill_date) == f_year)
                  .group_by(FuelBill.vehicle_id, 'm')
                  .all())
    pivot = {}  # {vehicle_id: {month: total}}
    for vid, m, total in pivot_rows:
        pivot.setdefault(vid, {})[int(m)] = float(total or 0)

    # Status + balance-after-this-bill (computed running balance, descending date)
    # Iterate bills in chronological order so balance_after reflects state when bill was created.
    chrono = sorted(bills, key=lambda b: (b.bill_date, b.id))
    running = float(reserve_amount)
    bill_meta = {}  # {bill_id: {status, balance_after}}
    for b in chrono:
        if _depletes_reserve(b.payment_method):
            running -= float(b.amount)
        # if bill is already 'ได้เงิน', reserve was refunded — but we want snapshot at bill date
        bill_meta[b.id] = {
            'status': _bill_status(b),
            'balance_after': running,
        }

    # Dropdowns
    vehicles = Vehicle.query.order_by(Vehicle.license_plate).all()
    drivers  = Driver.query.filter_by(is_active=True).order_by(Driver.name).all()

    pivot_labels = {v.id: v.license_plate for v in vehicles if v.id in pivot}
    all_pivot_vals = [val for row in pivot.values() for val in row.values()]
    pivot_max = max(all_pivot_vals) if all_pivot_vals else 1
    year_bills_count = len(year_bills)

    # Reserve history + price history (latest 20 each)
    reserve_logs = FuelReserveLog.query.order_by(FuelReserveLog.created_at.desc()).limit(20).all()
    fuel_prices  = FuelPrice.query.order_by(FuelPrice.effective_date.desc()).limit(20).all()

    # Distinct years that actually have bills (desc) — always include f_year
    year_rows = (db.session.query(extract('year', FuelBill.bill_date))
                 .filter(FuelBill.bill_date.isnot(None))
                 .distinct().all())
    available_years = sorted({int(y[0]) for y in year_rows if y[0]} | {f_year}, reverse=True)

    return render_template(
        'vehicle/admin/admin_fuel.html',
        # data
        bills=bills,
        reimbursements=reimbursements,
        bill_meta=bill_meta,
        vehicles=vehicles,
        drivers=drivers,
        reserve_logs=reserve_logs,
        fuel_prices=fuel_prices,
        pivot=pivot,
        pivot_labels=pivot_labels,
        pivot_max=pivot_max,
        year_bills_count=year_bills_count,
        # KPIs
        reserve_amount=reserve_amount,
        reserve_used=reserve_used,
        reserve_balance=reserve_balance,
        pending_total=pending_total,
        pending_by_status=pending_by_status,
        self_paid_total=self_paid_total,
        self_paid_count=self_paid_count,
        annual_budget=annual_budget,
        year_used=year_used,
        year_remaining=year_remaining,
        by_method=by_method,
        # filters
        f_year=f_year,
        f_month=f_month,
        f_veh=f_veh,
        f_drv=f_drv,
        available_years=available_years,
    )


# ─────────────────────────────────────────────
# Bills CRUD
# ─────────────────────────────────────────────
@fuel_bp.route('/admin/fuel/bill', methods=['POST'])
@login_required
def create_bill():
    _guard()
    bill_date = _parse_date(request.form.get('bill_date'))
    vehicle_id = _parse_int(request.form.get('vehicle_id'))
    driver_id  = _parse_int(request.form.get('driver_id'))
    amount     = _parse_decimal(request.form.get('amount'))
    method     = (request.form.get('payment_method') or '').strip()
    mileage    = _parse_int(request.form.get('mileage'))
    note       = (request.form.get('note') or '').strip() or None

    if not (bill_date and vehicle_id and driver_id and amount > 0 and method in PAYMENT_METHODS):
        flash('กรุณากรอกข้อมูลให้ครบ (วันที่, รถ, ผู้เติม, จำนวนเงิน, ช่องทาง)', 'danger')
        return redirect(request.referrer or url_for('fuel.admin_fuel'))

    bill = FuelBill(
        bill_date=bill_date,
        vehicle_id=vehicle_id,
        driver_id=driver_id,
        amount=amount,
        payment_method=method,
        mileage=mileage,
        note=note,
        created_by=current_user.id,
    )
    db.session.add(bill)
    db.session.commit()
    flash('บันทึกบิลเรียบร้อย', 'success')
    return redirect(request.referrer or url_for('fuel.admin_fuel'))


@fuel_bp.route('/admin/fuel/bill/<int:bill_id>/edit', methods=['POST'])
@login_required
def edit_bill(bill_id):
    _guard()
    bill = FuelBill.query.get_or_404(bill_id)

    bill.bill_date = _parse_date(request.form.get('bill_date'), bill.bill_date)
    bill.vehicle_id = _parse_int(request.form.get('vehicle_id'), bill.vehicle_id)
    bill.driver_id  = _parse_int(request.form.get('driver_id'),  bill.driver_id)
    bill.amount     = _parse_decimal(request.form.get('amount'), bill.amount)
    method = (request.form.get('payment_method') or '').strip()
    if method in PAYMENT_METHODS:
        bill.payment_method = method
    bill.mileage = _parse_int(request.form.get('mileage'), bill.mileage)
    bill.note    = (request.form.get('note') or '').strip() or None

    db.session.commit()
    flash('แก้ไขบิลเรียบร้อย', 'success')
    return redirect(request.referrer or url_for('fuel.admin_fuel'))


@fuel_bp.route('/admin/fuel/bill/<int:bill_id>/delete', methods=['POST'])
@login_required
def delete_bill(bill_id):
    _guard()
    bill = FuelBill.query.get_or_404(bill_id)
    db.session.delete(bill)
    db.session.commit()
    flash('ลบบิลเรียบร้อย', 'success')
    return redirect(request.referrer or url_for('fuel.admin_fuel'))


# ─────────────────────────────────────────────
# Reimbursement (รวมบิล)
# ─────────────────────────────────────────────
@fuel_bp.route('/admin/fuel/reimbursement', methods=['POST'])
@login_required
def create_reimbursement():
    _guard()
    rb_no = (request.form.get('reimbursement_no') or '').strip()
    source = (request.form.get('source') or '').strip() or None
    submitted_at = _parse_date(request.form.get('submitted_at'), date.today())
    note = (request.form.get('note') or '').strip() or None
    bill_ids = request.form.getlist('bill_ids')
    bill_ids = [int(x) for x in bill_ids if x.isdigit()]

    if not rb_no or not bill_ids:
        flash('กรุณาเลือกบิลและกรอกเลขใบเบิก', 'danger')
        return redirect(request.referrer or url_for('fuel.admin_fuel'))

    rb = FuelReimbursement(
        reimbursement_no=rb_no,
        source=source,
        submitted_at=submitted_at,
        note=note,
        created_by=current_user.id,
    )
    db.session.add(rb)
    db.session.flush()   # get rb.id

    # Attach bills (only those still รอเบิก)
    bills = FuelBill.query.filter(FuelBill.id.in_(bill_ids),
                                  FuelBill.reimbursement_id.is_(None)).all()
    for b in bills:
        b.reimbursement_id = rb.id

    db.session.commit()
    flash(f'รวม {len(bills)} บิลเป็นใบเบิก {rb_no} เรียบร้อย', 'success')
    return redirect(request.referrer or url_for('fuel.admin_fuel'))


@fuel_bp.route('/admin/fuel/reimbursement/<int:rb_id>/edit', methods=['POST'])
@login_required
def edit_reimbursement(rb_id):
    _guard()
    rb = FuelReimbursement.query.get_or_404(rb_id)
    rb.reimbursement_no = (request.form.get('reimbursement_no') or rb.reimbursement_no).strip()
    rb.source = (request.form.get('source') or '').strip() or None
    rb.submitted_at = _parse_date(request.form.get('submitted_at'), rb.submitted_at)
    rb.received_at  = _parse_date(request.form.get('received_at'),  rb.received_at)
    rb.note = (request.form.get('note') or '').strip() or None
    db.session.commit()
    flash('แก้ไขใบเบิกเรียบร้อย', 'success')
    return redirect(request.referrer or url_for('fuel.admin_fuel'))


@fuel_bp.route('/admin/fuel/reimbursement/<int:rb_id>/receive', methods=['POST'])
@login_required
def receive_reimbursement(rb_id):
    _guard()
    rb = FuelReimbursement.query.get_or_404(rb_id)
    rb.received_at = _parse_date(request.form.get('received_at'), date.today())
    db.session.commit()
    flash(f'บันทึกได้เงินคืน ใบเบิก {rb.reimbursement_no}', 'success')
    return redirect(request.referrer or url_for('fuel.admin_fuel'))


@fuel_bp.route('/admin/fuel/reimbursement/<int:rb_id>/delete', methods=['POST'])
@login_required
def delete_reimbursement(rb_id):
    _guard()
    rb = FuelReimbursement.query.get_or_404(rb_id)
    # Detach all bills (set reimbursement_id back to NULL)
    FuelBill.query.filter_by(reimbursement_id=rb.id).update({'reimbursement_id': None})
    db.session.delete(rb)
    db.session.commit()
    flash('ลบใบเบิกเรียบร้อย (บิลกลับสู่สถานะรอเบิก)', 'success')
    return redirect(request.referrer or url_for('fuel.admin_fuel'))


# ─────────────────────────────────────────────
# Reserve config + log
# ─────────────────────────────────────────────
@fuel_bp.route('/admin/fuel/reserve', methods=['POST'])
@login_required
def adjust_reserve():
    _guard()
    change = _parse_decimal(request.form.get('change_amount'))
    note   = (request.form.get('note') or '').strip()

    if change == 0 or not note:
        flash('กรุณากรอกจำนวนเงิน (≠ 0) และเหตุผล', 'danger')
        return redirect(request.referrer or url_for('fuel.admin_fuel'))

    cfg = FuelReserveConfig.query.get(1)
    if not cfg:
        cfg = FuelReserveConfig(id=1, amount=Decimal('0'))
        db.session.add(cfg)
        db.session.flush()

    cfg.amount = (cfg.amount or Decimal('0')) + change
    cfg.updated_by = current_user.id

    log = FuelReserveLog(
        change_amount=change,
        new_balance=cfg.amount,
        note=note,
        created_by=current_user.id,
    )
    db.session.add(log)
    db.session.commit()
    flash('ปรับเงินสำรองเรียบร้อย', 'success')
    return redirect(request.referrer or url_for('fuel.admin_fuel'))


# ─────────────────────────────────────────────
# Fuel price (effective-dated)
# ─────────────────────────────────────────────
@fuel_bp.route('/admin/fuel/price', methods=['POST'])
@login_required
def add_price():
    _guard()
    eff_date = _parse_date(request.form.get('effective_date'))
    price    = _parse_decimal(request.form.get('price_per_liter'))
    note     = (request.form.get('note') or '').strip() or None

    if not eff_date or price <= 0:
        flash('กรุณากรอกวันที่และราคา/ลิตร', 'danger')
        return redirect(request.referrer or url_for('fuel.admin_fuel'))

    existing = FuelPrice.query.filter_by(effective_date=eff_date).first()
    if existing:
        existing.price_per_liter = price
        existing.note = note
        existing.created_by = current_user.id
        flash('อัปเดตราคา/ลิตร สำหรับวันที่นี้แล้ว', 'success')
    else:
        db.session.add(FuelPrice(
            effective_date=eff_date,
            price_per_liter=price,
            note=note,
            created_by=current_user.id,
        ))
        flash('เพิ่มราคา/ลิตร เรียบร้อย', 'success')
    db.session.commit()
    return redirect(request.referrer or url_for('fuel.admin_fuel'))


@fuel_bp.route('/admin/fuel/price/<int:price_id>/delete', methods=['POST'])
@login_required
def delete_price(price_id):
    _guard()
    p = FuelPrice.query.get_or_404(price_id)
    db.session.delete(p)
    db.session.commit()
    flash('ลบราคา/ลิตรเรียบร้อย', 'success')
    return redirect(request.referrer or url_for('fuel.admin_fuel'))


# ─────────────────────────────────────────────
# Annual budget config (uses SystemConfig key='fuel_annual_budget')
# ─────────────────────────────────────────────
@fuel_bp.route('/admin/fuel/annual-budget', methods=['POST'])
@login_required
def set_annual_budget():
    _guard()
    amount = _parse_decimal(request.form.get('amount'))
    if amount < 0:
        flash('จำนวนเงินไม่ถูกต้อง', 'danger')
        return redirect(request.referrer or url_for('fuel.admin_fuel'))
    SystemConfig.set('fuel_annual_budget', float(amount))
    flash('ตั้งงบน้ำมันรายปีเรียบร้อย', 'success')
    return redirect(request.referrer or url_for('fuel.admin_fuel'))


# ─────────────────────────────────────────────
# JSON helper (mileage page badge — phase 3 will consume)
# ─────────────────────────────────────────────
@fuel_bp.route('/api/fuel/bill-by-mileage', methods=['GET'])
@login_required
def api_bill_by_mileage():
    """Lookup FuelBill by (vehicle_id, mileage). Returns list (could be 0..N)."""
    _guard()
    vid = _parse_int(request.args.get('vehicle_id'))
    m   = _parse_int(request.args.get('mileage'))
    if not (vid and m):
        return jsonify([])
    bills = FuelBill.query.filter_by(vehicle_id=vid, mileage=m).all()
    return jsonify([{
        'id': b.id,
        'date': b.bill_date.isoformat() if b.bill_date else None,
        'amount': float(b.amount),
        'payment_method': b.payment_method,
    } for b in bills])


# ─────────────────────────────────────────────
# Excel export — GET /admin/fuel/export/excel
# ─────────────────────────────────────────────
@fuel_bp.route('/admin/fuel/export/excel', methods=['GET'])
@login_required
def export_excel():
    """3-sheet workbook: บิล / ใบเบิก / Pivot — honors year/month/vehicle/driver filters."""
    _guard()

    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        flash('ไม่พบ openpyxl — รัน: pip install openpyxl', 'danger')
        return redirect(url_for('fuel.admin_fuel'))
    from flask import send_file

    today = date.today()
    f_year, f_month, f_veh, f_drv = _read_filters(request.args)

    bills = _filtered_bills_query(f_year, f_month, f_veh, f_drv).all()

    reimbursements = (FuelReimbursement.query
                      .filter(
                          (extract('year', FuelReimbursement.submitted_at) == f_year) |
                          (extract('year', FuelReimbursement.created_at) == f_year)
                      )
                      .order_by(FuelReimbursement.id.desc())
                      .all())

    pivot_rows = (db.session.query(
                      FuelBill.vehicle_id,
                      extract('month', FuelBill.bill_date).label('m'),
                      func.sum(FuelBill.amount).label('total'))
                  .filter(extract('year', FuelBill.bill_date) == f_year)
                  .group_by(FuelBill.vehicle_id, 'm')
                  .all())
    pivot = {}  # {vehicle_id: {month: total}}
    for vid, m, total in pivot_rows:
        pivot.setdefault(vid, {})[int(m)] = float(total or 0)
    vehicles = {v.id: v for v in Vehicle.query.order_by(Vehicle.license_plate).all()}

    # Workbook style helpers
    hdr_fill = PatternFill('solid', fgColor='4F46E5')
    hdr_font = Font(bold=True, color='FFFFFF', name='Sarabun')
    bold     = Font(bold=True, name='Sarabun')
    thin     = Side(style='thin', color='E4E4E7')
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)
    money_fmt = '#,##0.00'

    def write_header(ws, headers, widths):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = hdr_font; c.fill = hdr_fill
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = border
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[1].height = 22

    wb = openpyxl.Workbook()

    # ── Sheet 1: บิล ───────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'บิล'
    h1 = ['วันที่','รถ','ผู้เติม','จำนวน(฿)','ช่องทาง','ไมล์','สถานะ','เลขใบเบิก']
    write_header(ws1, h1, [12, 22, 22, 14, 12, 12, 12, 18])
    ri = 2
    bills_total = 0.0
    for b in bills:
        veh = b.vehicle
        veh_str = f"{veh.license_plate} {veh.brand} {veh.model}".strip() if veh else '-'
        drv_str = b.driver.name if b.driver else '-'
        rb_no = b.reimbursement.reimbursement_no if b.reimbursement else ''
        amt = float(b.amount or 0)
        bills_total += amt
        row = [
            b.bill_date.strftime('%d/%m/%Y') if b.bill_date else '',
            veh_str, drv_str, amt,
            PAYMENT_LABEL_TH.get(b.payment_method, b.payment_method or '-'),
            b.mileage if b.mileage is not None else '',
            _bill_status(b),
            rb_no,
        ]
        for ci, val in enumerate(row, 1):
            c = ws1.cell(row=ri, column=ci, value=val)
            c.border = border
            c.alignment = Alignment(horizontal='center' if ci in (1, 4, 5, 6, 7) else 'left')
            if ci == 4: c.number_format = money_fmt
            if ri % 2 == 0: c.fill = PatternFill('solid', fgColor='FAFAFA')
        ri += 1
    # totals
    if bills:
        c = ws1.cell(row=ri, column=3, value='รวม'); c.font = bold; c.alignment = Alignment(horizontal='right')
        c = ws1.cell(row=ri, column=4, value=round(bills_total, 2)); c.font = bold; c.number_format = money_fmt

    # ── Sheet 2: ใบเบิก ─────────────────────────────────────────────
    ws2 = wb.create_sheet('ใบเบิก')
    h2 = ['เลขใบเบิก','แหล่งเบิก','วันส่ง','วันได้เงิน','จำนวนบิล','รวมเงิน(฿)']
    write_header(ws2, h2, [18, 22, 12, 12, 12, 16])
    ri = 2
    rb_total = 0.0
    for rb in reimbursements:
        rb_bills = rb.bills if hasattr(rb, 'bills') else []
        rb_sum = sum((float(b.amount or 0) for b in rb_bills), 0.0)
        rb_total += rb_sum
        row = [
            rb.reimbursement_no or '',
            rb.source or '-',
            rb.submitted_at.strftime('%d/%m/%Y') if rb.submitted_at else '-',
            rb.received_at.strftime('%d/%m/%Y')  if rb.received_at  else 'รอเงิน',
            len(rb_bills),
            rb_sum,
        ]
        for ci, val in enumerate(row, 1):
            c = ws2.cell(row=ri, column=ci, value=val)
            c.border = border
            c.alignment = Alignment(horizontal='center' if ci in (1, 3, 4, 5, 6) else 'left')
            if ci == 6: c.number_format = money_fmt
            if ri % 2 == 0: c.fill = PatternFill('solid', fgColor='FAFAFA')
        ri += 1
    if reimbursements:
        c = ws2.cell(row=ri, column=5, value='รวม'); c.font = bold; c.alignment = Alignment(horizontal='right')
        c = ws2.cell(row=ri, column=6, value=round(rb_total, 2)); c.font = bold; c.number_format = money_fmt

    # ── Sheet 3: Pivot ─────────────────────────────────────────────
    ws3 = wb.create_sheet('Pivot')
    h3 = ['รถ'] + MONTH_LABEL_TH + ['รวมทั้งปี']
    write_header(ws3, h3, [22] + [12]*12 + [16])
    ri = 2
    month_totals = [0.0]*12
    grand_total = 0.0
    # sort vehicles by license plate for stable order
    sorted_vids = sorted(pivot.keys(), key=lambda vid: (vehicles[vid].license_plate if vid in vehicles else ''))
    for vid in sorted_vids:
        veh = vehicles.get(vid)
        veh_str = f"{veh.license_plate} {veh.brand} {veh.model}".strip() if veh else f"#{vid}"
        row_total = 0.0
        c = ws3.cell(row=ri, column=1, value=veh_str)
        c.border = border; c.alignment = Alignment(horizontal='left')
        if ri % 2 == 0: c.fill = PatternFill('solid', fgColor='FAFAFA')
        for mi in range(1, 13):
            v = pivot[vid].get(mi, 0.0)
            row_total += v
            month_totals[mi-1] += v
            c = ws3.cell(row=ri, column=1+mi, value=(v if v else None))
            c.border = border; c.alignment = Alignment(horizontal='right')
            c.number_format = money_fmt
            if ri % 2 == 0: c.fill = PatternFill('solid', fgColor='FAFAFA')
        grand_total += row_total
        c = ws3.cell(row=ri, column=14, value=row_total)
        c.border = border; c.alignment = Alignment(horizontal='right')
        c.number_format = money_fmt; c.font = bold
        if ri % 2 == 0: c.fill = PatternFill('solid', fgColor='FAFAFA')
        ri += 1
    # footer: รวมต่อเดือน
    if pivot:
        c = ws3.cell(row=ri, column=1, value='รวมต่อเดือน')
        c.font = bold; c.border = border; c.alignment = Alignment(horizontal='right')
        for mi in range(1, 13):
            c = ws3.cell(row=ri, column=1+mi, value=month_totals[mi-1] or None)
            c.font = bold; c.border = border; c.alignment = Alignment(horizontal='right')
            c.number_format = money_fmt
        c = ws3.cell(row=ri, column=14, value=grand_total)
        c.font = bold; c.border = border; c.alignment = Alignment(horizontal='right')
        c.number_format = money_fmt

    # Filename: fuel_<year>_<month>_<vehicle>_<date>.xlsx
    month_part = f"{f_month:02d}" if f_month else 'all'
    if f_veh and f_veh in vehicles:
        veh_part = ''.join(ch for ch in vehicles[f_veh].license_plate if ch.isalnum()) or 'veh'
    else:
        veh_part = 'all'
    fname = f"fuel_{f_year}_{month_part}_{veh_part}_{today.strftime('%Y%m%d')}.xlsx"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
