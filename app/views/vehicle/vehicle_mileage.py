from flask import render_template, request, redirect, url_for, flash
from flask_login import login_required, current_user
from models import (db, get_bkk_time, Vehicle, VehicleBooking, Driver, VehicleMileage,
                    VehicleBudget, FuelBill, DriverOT)
from sqlalchemy import extract
from datetime import datetime, timedelta
from views.core.notification_service import (
    notify_mileage_started      as _n_mileage_start,
    notify_mileage_ended        as _n_mileage_end,
)
import os, time
from werkzeug.utils import secure_filename
from views.vehicle.vehicle_common import (
    vehicle_bp, is_vehicle_admin, _build_budget_subs,
    get_fuel_price, calc_fuel_cost,
)
import services.vehicle.mileage_service as mileage_svc


def _compute_mileage_cost(b, m):
    """Return (distance, fuel_cost, status_key, fuel_price) for a booking+mileage pair.
    fuel_price เพิ่ม 2026-07-22 (Case 17 merge) — ราคาน้ำมัน/ลิตร ที่ใช้คำนวณจริง
    (None ถ้ายังไม่ complete)"""
    if not m:
        return (None, None, 'none', None)
    if m.odometer_start and m.odometer_end:
        d  = m.odometer_end - m.odometer_start
        td = m.actual_end.date() if m.actual_end else b.start_datetime.date()
        fp = get_fuel_price(td)
        c  = calc_fuel_cost(b.assigned_vehicle, d, fp, m.fuel_cost)
        return (d, c, 'complete', fp)
    if m.odometer_start:
        return (None, None, 'partial', None)
    return (None, None, 'none', None)


def _get_mileage_budget_info(b):
    """Return (budget_type, budget_label, budget_sub) for a booking."""
    et = (b.expense_type or '').strip()
    if et == 'central':
        return ('central', 'งบส่วนกลาง', (b.central_category or '').strip() or None)
    if et == 'department':
        sub = (b.trip_department or (b.user.department if b.user else '') or '').strip()
        return ('department', 'งบส่วนกอง', sub or None)
    if et == 'personal':
        return ('personal', 'งบส่วนตัว', None)
    return ('', '—', None)


def _handle_mileage_start(booking, mileage, upload_folder):
    mileage.odometer_start = int(request.form.get('odometer_start', 0))
    mileage.actual_start   = datetime.strptime(request.form.get('actual_start'), '%Y-%m-%dT%H:%M')
    img = request.files.get('odometer_start_img')
    if img and img.filename:
        fname = f"{int(time.time())}_start_{secure_filename(img.filename)}"
        img.save(os.path.join(upload_folder, fname))
        mileage.odometer_start_img = fname
    db.session.flush()
    stale_msgs = mileage_svc.auto_close_stale_trips(
        booking.assigned_vehicle_id, mileage.odometer_start, mileage.actual_start,
        booking.id, actor_id=current_user.id)
    for msg, cat in stale_msgs:
        flash(msg, cat)
    _n_mileage_start(booking, mileage)
    flash(f'บันทึกเลขไมล์ก่อนออก #{booking.id} เรียบร้อย', 'success')


def _handle_mileage_end(booking, mileage, upload_folder):
    """Set end-mileage fields. Returns False (with flash) if odometer validation fails."""
    submitted_end = int(request.form.get('odometer_end', 0))
    if mileage.odometer_start is not None and submitted_end <= mileage.odometer_start:
        flash(
            f'❌ บันทึกไม่สำเร็จ! เลขไมล์ตอนจบ ({submitted_end}) '
            f'ต้องมากกว่าเลขไมล์ตอนเริ่ม ({mileage.odometer_start})',
            'danger'
        )
        return False
    # REQ-3 (Phase 3.5, 2026-07-19): เพดานระยะทาง — confirm ผ่านได้ ไม่ hard block
    # (ตกลงกับเจ้าของโปรเจกต์) — JS ถาม confirm() ก่อนแล้วเซ็ต confirm_distance=1 ให้ปกติ
    # เกิดที่นี่เฉพาะกรณี JS ถูกข้าม/ปิดไป (safety net)
    if mileage.odometer_start is not None:
        distance = submitted_end - mileage.odometer_start
        cap = mileage_svc.get_distance_cap_km()
        if distance > cap and request.form.get('confirm_distance') != '1':
            flash(
                f'⚠️ ระยะทาง {distance:,} กม. เกินเพดานปกติ ({cap:,.0f} กม.) — '
                f'กรุณาตรวจสอบเลขไมล์แล้วยืนยันอีกครั้งถ้าถูกต้อง',
                'warning'
            )
            return False
    mileage.odometer_end = submitted_end
    mileage.actual_end   = datetime.strptime(request.form.get('actual_end'), '%Y-%m-%dT%H:%M')
    img = request.files.get('odometer_end_img')
    if img and img.filename:
        fname = f"{int(time.time())}_end_{secure_filename(img.filename)}"
        img.save(os.path.join(upload_folder, fname))
        mileage.odometer_end_img = fname
    mileage.refuel = True if request.form.get('refuel') else False
    if mileage.refuel:
        refuel_amt = request.form.get('refuel_amount', '').strip()
        if refuel_amt:
            mileage.refuel_amount = float(refuel_amt)
        refuel_img = request.files.get('refuel_img')
        if refuel_img and refuel_img.filename:
            fname = f"{int(time.time())}_refuel_{secure_filename(refuel_img.filename)}"
            refuel_img.save(os.path.join(upload_folder, fname))
            mileage.refuel_img = fname
    fuel = request.form.get('fuel_cost', '').strip()
    if fuel:
        mileage.fuel_cost = float(fuel)
    flash(f'บันทึกเลขไมล์หลังกลับ #{booking.id} เรียบร้อย', 'success')
    return True



def _query_mileage_bookings(cutoff, f_date_start, f_date_end, f_vehicle,
                             f_driver, f_budget_type, f_budget_sub):
    q = VehicleBooking.query.filter(
        VehicleBooking.status == 'approved',
        VehicleBooking.start_datetime < cutoff,
    )
    if f_date_start:
        try:
            q = q.filter(VehicleBooking.start_datetime >= datetime.strptime(f_date_start, '%Y-%m-%d'))
        except ValueError:
            pass
    if f_date_end:
        try:
            end_dt = datetime.strptime(f_date_end, '%Y-%m-%d') + timedelta(days=1)
            q = q.filter(VehicleBooking.start_datetime < end_dt)
        except ValueError:
            pass
    if f_vehicle:
        q = q.filter(VehicleBooking.assigned_vehicle_id == f_vehicle)
    if f_driver:
        q = q.filter(VehicleBooking.driver_id == f_driver)
    if f_budget_type in ('central', 'department', 'personal'):
        q = q.filter(VehicleBooking.expense_type == f_budget_type)
    if f_budget_sub and f_budget_type == 'central':
        q = q.filter(VehicleBooking.central_category == f_budget_sub)
    elif f_budget_sub and f_budget_type == 'department':
        q = q.filter(VehicleBooking.trip_department == f_budget_sub)
    return q.order_by(VehicleBooking.start_datetime.desc()).all()


def _build_mileage_rows(bookings, fuel_by_vehicle, ot_records_by_booking, f_status,
                         f_pending_personal=False):
    rows = []
    for b in bookings:
        m = b.mileage[0] if b.mileage else None
        distance, fuel_cost, status_key, fuel_price = _compute_mileage_cost(b, m)
        if f_status == 'incomplete':
            if status_key not in ('none', 'partial'):
                continue
        elif f_status and f_status != status_key:
            continue
        if f_pending_personal:
            is_pending_personal = (
                (b.expense_type or '') == 'personal'
                and m is not None and m.odometer_end is not None
                and m.personal_status == 0
            )
            if not is_pending_personal:
                continue
        budget_type, budget_label, budget_sub = _get_mileage_budget_info(b)
        bills = fuel_by_vehicle.get(b.assigned_vehicle_id, []) if b.assigned_vehicle_id else []
        if m and m.odometer_start and m.odometer_end:
            refuel_odo = next((km for km in bills if m.odometer_start <= km <= m.odometer_end), None)
            has_refuel = refuel_odo is not None
        else:
            refuel_odo = None
            has_refuel = False

        # OT breakdown ต่อ slot (Case 17 merge, 2026-07-22) — 1 booking = 1 DriverOT
        # (auto_generate_ot idempotent ต่อ booking_id), slots มาจาก DriverOTSlot ตัวจริง
        ot_record       = ot_records_by_booking.get(b.id)
        ot_hours        = float(ot_record.total_hours)  if ot_record else 0
        ot_total_amount = float(ot_record.total_amount) if ot_record else 0
        ot_slots = [
            {'label': s.slot_label, 'rate': float(s.rate), 'hours': float(s.hours),
             'amount': float(s.amount), 'start_time': s.start_time, 'end_time': s.end_time}
            for s in ot_record.slots
        ] if ot_record else []

        rows.append({
            'b': b, 'm': m,
            'distance':        distance,
            'fuel_cost':       fuel_cost,
            'fuel_price':      fuel_price,
            'status_key':      status_key,
            'budget_type':     budget_type,
            'budget_label':    budget_label,
            'budget_sub':      budget_sub,
            'has_refuel':      has_refuel,
            'refuel_odo':      refuel_odo,
            'ot_hours':        ot_hours,
            'ot_total_amount': ot_total_amount,
            'ot_slots':        ot_slots,
        })

    display_rows = []
    seen_groups  = set()
    for r in rows:
        tg = r['b'].trip_group
        if tg is None:
            display_rows.append({'kind': 'single', 'row': r, 'count': 1, 'members': [r]})
            continue
        if tg in seen_groups:
            continue
        seen_groups.add(tg)
        members = [x for x in rows if x['b'].trip_group == tg]
        display_rows.append({'kind': 'group', 'row': members[0], 'count': len(members), 'members': members})

    return rows, display_rows


def _calc_mileage_kpi(now, cutoff):
    year_budgets    = VehicleBudget.query.filter_by(year=now.year).all()
    total_budget    = sum(float(bu.budget_amount) for bu in year_budgets)
    total_used      = sum(float(bu.used_amount)   for bu in year_budgets)
    total_remaining = total_budget - total_used

    month_trips = VehicleBooking.query.filter(
        VehicleBooking.status == 'approved',
        extract('year',  VehicleBooking.start_datetime) == now.year,
        extract('month', VehicleBooking.start_datetime) == now.month,
    ).all()
    month_total_cost = sum(
        c for b in month_trips
        for _, c, st, _ in [_compute_mileage_cost(b, b.mileage[0] if b.mileage else None)]
        if st == 'complete' and c
    )

    pending_personal_count = VehicleMileage.query.join(VehicleBooking).filter(
        VehicleBooking.expense_type == 'personal',
        VehicleMileage.personal_status == 0,
        VehicleMileage.odometer_end.isnot(None),
    ).count()

    all_past = VehicleBooking.query.filter(
        VehicleBooking.status == 'approved',
        VehicleBooking.start_datetime < cutoff,
    ).all()
    missing_count = sum(
        1 for b in all_past
        if _compute_mileage_cost(b, b.mileage[0] if b.mileage else None)[2] != 'complete'
    )

    return {
        'month_total_cost':       month_total_cost,
        'total_budget':           total_budget,
        'total_used':             total_used,
        'total_remaining':        total_remaining,
        'pending_personal_count': pending_personal_count,
        'missing_count':          missing_count,
    }


def _parse_mileage_filters(today):
    """Parse GET query param (mileage dashboard filter) (extract จาก mileage_log ตอน Phase 5,
    logic เดิม 100% รวม default-date เมื่อไม่ระบุช่วงและไม่กด show_all)
    คืน tuple 9 ค่าตามลำดับที่ route ใช้"""
    show_all      = request.args.get('show_all', '') == '1'
    f_date_start  = request.args.get('date_start', '').strip()
    f_date_end    = request.args.get('date_end', '').strip()
    f_vehicle     = request.args.get('vehicle_id', type=int)
    f_driver      = request.args.get('driver_id', type=int)
    f_status      = request.args.get('status_filter', '').strip()
    f_budget_type = request.args.get('budget_type', '').strip()
    f_budget_sub  = request.args.get('budget_sub', '').strip()
    f_pending_personal = request.args.get('pending_personal', '') == '1'

    if not show_all and not f_date_start and not f_date_end:
        f_date_start = today.replace(day=1).strftime('%Y-%m-%d')
        f_date_end   = today.strftime('%Y-%m-%d')

    return (show_all, f_date_start, f_date_end, f_vehicle, f_driver, f_status,
            f_budget_type, f_budget_sub, f_pending_personal)


def _handle_mileage_post():
    """POST ของ mileage_log — บันทึกไมล์ start/end/both (extract จาก mileage_log ตอน
    Phase 5, logic เดิม 100%). entry_type='both' เพิ่ม 2026-07-22 (Case 17 merge) —
    admin กรอกเลขไมล์ออก+กลับพร้อมกันได้ในคำขอเดียว (driver ฝั่ง /driver/mileage คนละ
    route/JS ไม่ถูกแตะ ยังคง 2 ขั้นตอนแยกเหมือนเดิม). คืน response (redirect) ให้ route
    return ตรง"""
    booking_id = int(request.form.get('booking_id'))
    booking    = VehicleBooking.query.get_or_404(booking_id)
    entry_type = request.form.get('entry_type')

    mileage = VehicleMileage.query.filter_by(booking_id=booking_id).first()
    if not mileage:
        mileage = VehicleMileage(booking_id=booking_id, noted_by=current_user.id)
        db.session.add(mileage)

    upload_folder = os.path.join('static', 'uploads', 'mileage')
    os.makedirs(upload_folder, exist_ok=True)

    if entry_type == 'both':
        # เช็ก end > start จากฟอร์มดิบก่อนแตะ mileage object เลย (all-or-nothing) — กัน
        # flash "บันทึกไมล์ออกสำเร็จ" ค้างจาก _handle_mileage_start() ทั้งที่ end ยังไม่ผ่าน
        # (_handle_mileage_end() ยังตรวจซ้ำเป็น source of truth หลัง odometer_start ถูกเซ็ตจริง)
        submitted_start = int(request.form.get('odometer_start', 0))
        submitted_end   = int(request.form.get('odometer_end', 0))
        if submitted_end <= submitted_start:
            flash(
                f'❌ บันทึกไม่สำเร็จ! เลขไมล์ตอนจบ ({submitted_end}) '
                f'ต้องมากกว่าเลขไมล์ตอนเริ่ม ({submitted_start})',
                'danger'
            )
            return redirect(url_for('vehicle.mileage_log'))

    if entry_type in ('start', 'both'):
        _handle_mileage_start(booking, mileage, upload_folder)
    if entry_type in ('end', 'both'):
        if not _handle_mileage_end(booking, mileage, upload_folder):
            db.session.rollback()  # all-or-nothing: เคลียร์ odometer_start ที่ flush ไปแล้วด้วย
            return redirect(url_for('vehicle.mileage_log'))
        _n_mileage_end(booking, mileage)

    db.session.commit()

    if entry_type in ('end', 'both'):
        # notify_ot_created อยู่ใน auto_generate_ot() แล้ว (Phase 4, 2026-07-19)
        mileage_svc.auto_generate_ot(booking, mileage, actor_id=current_user.id)
        m2 = VehicleMileage.query.filter_by(booking_id=booking_id).first()
        result = mileage_svc.close_trip(booking, m2, source='mileage_log')
        for msg, cat in result['flash_messages']:
            flash(msg, cat)

    return redirect(url_for('vehicle.mileage_log'))


@vehicle_bp.route('/vehicle/mileage', methods=['GET', 'POST'])
@login_required
def mileage_log():
    if not is_vehicle_admin():
        flash('คุณไม่มีสิทธิ์เข้าหน้านี้', 'danger')
        return redirect(url_for('vehicle.index'))

    if request.method == 'POST':
        return _handle_mileage_post()

    # ── GET: Admin mileage dashboard ─────────────────────────────
    today      = get_bkk_time().date()
    now        = get_bkk_time()
    fuel_price = mileage_svc.get_fuel_price(today)
    (show_all, f_date_start, f_date_end, f_vehicle, f_driver, f_status,
     f_budget_type, f_budget_sub, f_pending_personal) = _parse_mileage_filters(today)

    cutoff   = datetime.combine(today + timedelta(days=1), datetime.min.time())
    bookings = _query_mileage_bookings(cutoff, f_date_start, f_date_end, f_vehicle,
                                       f_driver, f_budget_type, f_budget_sub)

    fuel_by_vehicle = {}
    for vid, mil in (db.session.query(FuelBill.vehicle_id, FuelBill.mileage)
                                .filter(FuelBill.mileage.isnot(None)).all()):
        fuel_by_vehicle.setdefault(vid, []).append(mil)

    # DriverOT เต็ม object (ไม่ใช่แค่ summed hours) — ให้ template/JS ดึง slot breakdown
    # (label/rate/amount ต่อ time-band) ได้ (Case 17 merge, 2026-07-22)
    ot_records_by_booking = {}
    for ot in (DriverOT.query
               .filter(DriverOT.booking_id.isnot(None), DriverOT.is_deleted.is_(False))
               .all()):
        ot_records_by_booking[ot.booking_id] = ot  # 1 booking = 1 DriverOT (auto_generate_ot idempotent)

    rows, display_rows        = _build_mileage_rows(bookings, fuel_by_vehicle, ot_records_by_booking,
                                                     f_status, f_pending_personal)
    kpi                       = _calc_mileage_kpi(now, cutoff)
    vehicles_all              = Vehicle.query.order_by(Vehicle.license_plate).all()
    drivers_all               = Driver.query.filter_by(is_active=True).order_by(Driver.name).all()
    budget_subs               = _build_budget_subs()

    return render_template('vehicle/admin/vehicle_mileage.html',
        rows=rows,
        display_rows=display_rows,
        fuel_price=fuel_price,
        distance_cap=mileage_svc.get_distance_cap_km(),
        today=today,
        **kpi,
        vehicles_all=vehicles_all,
        drivers_all=drivers_all,
        budget_subs=budget_subs,
        f={'date_start': f_date_start, 'date_end': f_date_end,
           'vehicle_id': f_vehicle or '', 'driver_id': f_driver or '',
           'status_filter': f_status,
           'budget_type': f_budget_type,
           'budget_sub': f_budget_sub,
           'pending_personal': '1' if f_pending_personal else '',
           'show_all': show_all},
    )


# ─────────────────────────────────────────────
# Export Excel — mileage (admin)
# ─────────────────────────────────────────────

def _filter_and_calc_mileage_rows(bookings, f_status):
    """คำนวณ distance/fuel_cost/status ต่อ booking + กรองตาม status (post-DB filter —
    fuel_cost มาจากสูตร ไม่ใช่ column ตรง กรองใน SQL ไม่ได้) (extract จาก mileage_export
    ตอน Phase 5 — logic เดิม 100%)
    คืน (rows, total_distance, total_fuel) — rows = [(booking, mileage, distance, fuel_cost,
    status_key), ...]"""
    rows = []
    total_distance = 0.0
    total_fuel     = 0.0
    for b in bookings:
        m = b.mileage[0] if b.mileage else None
        distance = fuel_cost = None
        status_key = 'none'
        if m and m.odometer_start and m.odometer_end:
            distance = m.odometer_end - m.odometer_start
            td = m.actual_end.date() if m.actual_end else b.start_datetime.date()
            fuel_cost  = calc_fuel_cost(b.assigned_vehicle, distance, get_fuel_price(td), m.fuel_cost) or None
            status_key = 'complete'
        elif m and m.odometer_start:
            status_key = 'partial'

        if f_status and f_status != status_key: continue

        if distance: total_distance += distance
        if fuel_cost: total_fuel += fuel_cost
        rows.append((b, m, distance, fuel_cost, status_key))
    return rows, total_distance, total_fuel


def _build_mileage_workbook(rows, total_distance, total_fuel, today):
    """สร้าง openpyxl Workbook จาก rows ที่กรอง+คำนวณแล้ว (extract จาก mileage_export ตอน
    Phase 5 — logic/styling เดิม 100%) คืน wb ให้ route save+send

    import openpyxl ในนี้ตั้งใจ (ไม่ย้ายขึ้น top-of-file) — เหตุผลเดียวกับ apscheduler ใน
    notification_cron.py (Phase 5): openpyxl เป็น optional dependency เฉพาะ export feature
    นี้ ถ้าย้ายขึ้น top-level ของ vehicle_mileage.py ทั้งไฟล์ (mileage dashboard/POST
    start-end) จะ import ไม่ได้ทันทีถ้า deployment ไหนไม่ได้ติดตั้ง openpyxl ทั้งที่ฟีเจอร์
    หลักไม่เกี่ยวกับ Excel เลย — mileage_export() (route ที่เรียกฟังก์ชันนี้) เช็ก
    ImportError + flash แจ้งเตือนไว้ก่อนเรียกแล้ว จึงมั่นใจได้ว่าถึงจุดนี้ import สำเร็จแน่"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Mileage {today.strftime('%Y-%m')}"

    headers = ['Booking','ผู้จอง','รถ','ทะเบียน','คนขับ','ปลายทาง','วันเดินทาง',
               'ไมล์ออก','ไมล์กลับ','ระยะทาง(กม.)','ค่าน้ำมัน(฿)','สถานะ','ประเภท']
    hdr_fill = PatternFill('solid', fgColor='4F46E5')
    hdr_font = Font(bold=True, color='FFFFFF', name='Sarabun')
    thin = Side(style='thin', color='E4E4E7')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border

    ST_LABEL  = {'complete':'ครบ','partial':'รอกลับ','none':'รอกรอก'}
    EXP_LABEL = {'central':'ส่วนกลาง','department':'หน่วยงาน','personal':'ส่วนตัว'}
    ri = 2
    for b, m, distance, fuel_cost, status_key in rows:
        row_data = [
            f"BK-{b.id}",
            b.user.full_name or b.user.username,
            f"{b.assigned_vehicle.brand} {b.assigned_vehicle.model}" if b.assigned_vehicle else '-',
            b.assigned_vehicle.license_plate if b.assigned_vehicle else '-',
            b.driver.name if b.driver else '-',
            b.destination or '-',
            b.start_datetime.strftime('%d/%m/%Y'),
            m.odometer_start if m and m.odometer_start else None,
            m.odometer_end   if m and m.odometer_end   else None,
            distance,
            round(fuel_cost, 2) if fuel_cost else None,
            ST_LABEL.get(status_key, '-'),
            EXP_LABEL.get(b.expense_type or '', 'ไม่ระบุ'),
        ]
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = border
            c.alignment = Alignment(horizontal='center' if ci in [1,7,8,9,10,11,12] else 'left')
            if ri % 2 == 0:
                c.fill = PatternFill('solid', fgColor='FAFAFA')
        ri += 1

    # Totals row
    tr = ri
    ws.cell(row=tr, column=9,  value='รวม').font = Font(bold=True)
    ws.cell(row=tr, column=10, value=round(total_distance, 2)).font = Font(bold=True)
    ws.cell(row=tr, column=11, value=round(total_fuel, 2)).font = Font(bold=True)

    col_widths = [10,22,18,14,16,28,12,12,12,14,14,12,14]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 22
    return wb


@vehicle_bp.route('/vehicle/mileage/export')
@login_required
def mileage_export():
    if not is_vehicle_admin():
        flash('คุณไม่มีสิทธิ์', 'danger')
        return redirect(url_for('vehicle.index'))

    import io
    try:
        import openpyxl
    except ImportError:
        flash('ไม่พบ openpyxl — รัน: pip install openpyxl', 'danger')
        return redirect(url_for('vehicle.mileage_log'))

    from flask import send_file
    today = get_bkk_time().date()

    f_date_start = request.args.get('date_start', '').strip()
    f_date_end   = request.args.get('date_end', '').strip()
    f_vehicle    = request.args.get('vehicle_id', type=int)
    f_driver     = request.args.get('driver_id', type=int)
    f_status     = request.args.get('status_filter', '').strip()

    cutoff = datetime.combine(today + timedelta(days=1), datetime.min.time())
    q = VehicleBooking.query.filter(
        VehicleBooking.status == 'approved',
        VehicleBooking.start_datetime < cutoff,
    )
    if f_date_start:
        try: q = q.filter(VehicleBooking.start_datetime >= datetime.strptime(f_date_start, '%Y-%m-%d'))
        except ValueError: pass
    if f_date_end:
        try: q = q.filter(VehicleBooking.start_datetime < datetime.strptime(f_date_end, '%Y-%m-%d') + timedelta(days=1))
        except ValueError: pass
    if f_vehicle: q = q.filter(VehicleBooking.assigned_vehicle_id == f_vehicle)
    if f_driver:  q = q.filter(VehicleBooking.driver_id == f_driver)
    bookings = q.order_by(VehicleBooking.start_datetime.desc()).all()

    rows, total_distance, total_fuel = _filter_and_calc_mileage_rows(bookings, f_status)
    wb = _build_mileage_workbook(rows, total_distance, total_fuel, today)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"mileage_{today.strftime('%Y%m%d')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ─────────────────────────────────────────────
# สรุปค่าใช้จ่าย (admin + superadmin)
# ─────────────────────────────────────────────
