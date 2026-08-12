"""
Fuel management blueprint — /admin/fuel
=======================================
Vehicle admin records fuel bills paid to drivers, batches them into
reimbursement claims, and tracks the cash reserve (เงินสำรอง).

Status (computed, not stored):
    รอเบิก   = bill.reimbursement_id IS NULL
    อนุมัติ  = reimbursement_id NOT NULL AND received_at IS NULL
    ได้เงิน  = received_at NOT NULL

Phase 1: DB + backend routes (UI = phase 2, Excel/PDF = phase 4).
"""
from datetime import datetime, date
from decimal import Decimal
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, abort, current_app
from flask_login import login_required, current_user
from sqlalchemy import extract, func, or_, and_
from models import (
    db, get_bkk_time, SystemConfig, User,
    Vehicle, Driver,
    FuelBill, FuelReimbursement, FuelPrice,
    FuelReserveLog, ExpenseHolder, ReimbursementSource,
    ReimbursementSettlement,
)
from domain.vehicle.fuel import PAYMENT_LABEL_TH, CATEGORY_LABEL_TH
import services.vehicle.fuel_service as fuel_svc

fuel_bp = Blueprint('fuel', __name__)


# ─────────────────────────────────────────────
# Permission
# ─────────────────────────────────────────────
def is_vehicle_admin():
    return current_user.role_vehicle == 'admin' or current_user.is_superadmin


def _guard():
    if not current_user.is_authenticated or not is_vehicle_admin():
        abort(403)


# ─────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────
MONTH_LABEL_TH = ['ม.ค.','ก.พ.','มี.ค.','เม.ย.','พ.ค.','มิ.ย.',
                  'ก.ค.','ส.ค.','ก.ย.','ต.ค.','พ.ย.','ธ.ค.']


def _bill_status(bill):
    """รอเบิก | อนุมัติ | ได้เงิน — derived from FK + received_at."""
    if bill.reimbursement_id is None:
        return 'รอเบิก'
    if bill.reimbursement and bill.reimbursement.received_at is None:
        return 'อนุมัติ'
    return 'ได้เงิน'


def _pending_status(bill):
    """ใช้ไปแล้ว | อยู่ในใบร่าง | ทำเรื่องเบิกแล้ว — tab ค้างเบิก (§4.3, ปรับหลัง P4 มี submit จริง)
    เช็กเลือกได้เฉพาะ "ใช้ไปแล้ว" เท่านั้น — 2 สถานะหลังแปลว่าเข้าใบเบิกไปแล้ว
    (attach_bills_to_reimbursement กรองแค่ reimbursement_id IS NULL อยู่แล้ว ไม่ถูกดึงซ้ำ)
    """
    if bill.reimbursement_id is None:
        return 'ใช้ไปแล้ว'
    if bill.reimbursement.status == 'draft':
        return 'อยู่ในใบร่าง'
    return 'ทำเรื่องเบิกแล้ว'


def _finished_status(bill):
    """ได้เงินคืน | ตัดบัตร | จ่ายเอง — tab จบแล้ว (D4, §5.4)"""
    if bill.payment_method == 'card':
        return 'ตัดบัตร'
    if bill.payment_method == 'self':
        return 'จ่ายเอง'
    return 'ได้เงินคืน'


def _category_filter_key(bill):
    """chip กรอง (D5) — ทะเบียนรถ ถ้าเป็นบิลน้ำมันที่ผูกรถ · ที่เหลือ (ไม่ใช่น้ำมัน/ไม่มีรถ) → 'other'"""
    return str(bill.vehicle_id) if (bill.category == 'fuel' and bill.vehicle_id) else 'other'


def _parse_date(s, default=None):
    if not s:
        return default
    try:
        return datetime.strptime(s, '%Y-%m-%d').date()
    except ValueError:
        return default


def _parse_int(s, default=None):
    try:
        return int(s) if s not in (None, '') else default
    except (ValueError, TypeError):
        return default


def _parse_decimal(s, default=Decimal('0')):
    try:
        return Decimal(str(s)) if s not in (None, '') else default
    except Exception:
        return default


def _read_filters(args):
    """Return (year, month, vehicle_id, driver_id) parsed from query string."""
    today = get_bkk_time().date()
    return (
        _parse_int(args.get('year'),  today.year),
        _parse_int(args.get('month'), 0),
        _parse_int(args.get('vehicle_id')),
        _parse_int(args.get('driver_id')),
    )


def _filtered_bills_query(year, month, vehicle_id, driver_id):
    """Build the FuelBill query used by both admin_fuel() and exports."""
    q = FuelBill.query.filter(extract('year', FuelBill.bill_date) == year)
    if month:      q = q.filter(extract('month', FuelBill.bill_date) == month)
    if vehicle_id: q = q.filter(FuelBill.vehicle_id == vehicle_id)
    if driver_id:  q = q.filter(FuelBill.driver_id == driver_id)
    return q.order_by(FuelBill.bill_date.desc(), FuelBill.id.desc())


# ─────────────────────────────────────────────
# Main page — GET /admin/fuel
# ─────────────────────────────────────────────
@fuel_bp.route('/admin/fuel', methods=['GET'])
@login_required
def admin_fuel():
    """หน้า "เงินสำรองและค่าใช้จ่าย" — 5 tab (review 2026-08-10 #6, spec §5.1)

    KPI strip เดิม + ตาราง "บิลเบิก" ledger เดิม (ผูก FuelReserveConfig singleton ที่เลิกใช้แล้ว)
    ถูกลบออกจากหน้านี้ทั้งคู่ — แทนที่ด้วย KPI เงินสำรองรายคน (my_kpi) + tab ค้างเบิก/จบแล้ว
    ที่ query ตรงจาก FuelBill ตามสถานะจริง ไม่ต้องพึ่ง running-balance คำนวณสดทุก request แล้ว
    """
    _guard()

    today = get_bkk_time().date()
    f_year = _parse_int(request.args.get('year'), today.year)

    # ปีที่มีบิลจริง (สำหรับ selector ของ pivot) — เสมอรวม f_year แม้ปีนั้นยังไม่มีบิล
    year_rows = (db.session.query(extract('year', FuelBill.bill_date))
                 .filter(FuelBill.bill_date.isnot(None))
                 .distinct().all())
    available_years = sorted({int(y[0]) for y in year_rows if y[0]} | {f_year}, reverse=True)

    # ── tab ภาพรวมทั้งปี — pivot รถ × เดือน (นับทุก payment_method, มิติน้ำมัน) ──
    pivot_rows = (db.session.query(
                      FuelBill.vehicle_id,
                      extract('month', FuelBill.bill_date).label('m'),
                      func.sum(FuelBill.amount).label('total'))
                  .filter(extract('year', FuelBill.bill_date) == f_year)
                  .group_by(FuelBill.vehicle_id, 'm')
                  .all())
    pivot = {}  # {vehicle_id: {month: total}}
    for vid, m, total in pivot_rows:
        pivot.setdefault(vid, {})[int(m)] = float(total or 0)

    vehicles = Vehicle.query.order_by(Vehicle.license_plate).all()
    drivers  = Driver.query.filter_by(is_active=True).order_by(Driver.name).all()
    vehicle_labels = {v.id: v.license_plate for v in vehicles}
    pivot_labels = {v.id: v.license_plate for v in vehicles if v.id in pivot}
    all_pivot_vals = [val for row in pivot.values() for val in row.values()]
    pivot_max = max(all_pivot_vals) if all_pivot_vals else 1

    # ── งบทั้งปี (แสดงใน modal ตั้งค่า → segment "งบทั้งปี") ──
    year_used_amt = (db.session.query(func.coalesce(func.sum(FuelBill.amount), 0))
                     .filter(extract('year', FuelBill.bill_date) == f_year).scalar())
    year_used = float(year_used_amt or 0)
    annual_budget = float(SystemConfig.get('fuel_annual_budget', 0) or 0)
    year_remaining = annual_budget - year_used
    fuel_prices = FuelPrice.query.order_by(FuelPrice.effective_date.desc()).limit(20).all()
    sources_all = ReimbursementSource.query.order_by(ReimbursementSource.is_default.desc(),
                                                     ReimbursementSource.name).all()

    # ── เงินสำรองรายคน (P2) — KPI ของคนล็อกอินคนเดียวเท่านั้น (D1) ──
    # ไม่ใช่ผู้สำรองเงิน → my_holder เป็น None → holder_kpi คืนศูนย์ทั้งหมด (D2)
    my_holder = fuel_svc.get_holder(current_user.id)
    my_kpi = fuel_svc.holder_kpi(my_holder)
    my_quota_lines = fuel_svc.quota_lines(today.year, today.month)
    source_labels = {s.id: s.name for s in sources_all if s.is_active}

    # tab เจ้าหน้าที่ — KPI ทุกคน + ประวัติปรับล่าสุดต่อคน (query อ่านล้วน ไม่ผ่าน service)
    # log serialize เป็น dict เอง (ไม่ใช่ |tojson ตรงบน ORM object) เพื่อให้ jinja tojson ใช้ในโมดัลได้
    holder_rows = fuel_svc.all_holder_kpis()
    for row in holder_rows:
        logs = (FuelReserveLog.query
                .filter_by(holder_id=row['holder'].id)
                .order_by(FuelReserveLog.created_at.desc())
                .limit(15).all())
        row['logs'] = [{
            'date':    log.created_at.strftime('%d/%m/%Y %H:%M') if log.created_at else '',
            'type':    log.log_type,
            'change':  float(log.change_amount),
            'balance': float(log.new_balance),
            'note':    log.note,
        } for log in logs]
    holder_candidate_users = (User.query
                              .filter(~User.id.in_(db.session.query(ExpenseHolder.user_id)))
                              .order_by(User.full_name)
                              .all())

    # ── tab ค้างเบิก (P3) — เฉพาะบิล reserve ที่ยังไม่ "ได้เงินคืนแล้ว" (reimbursement.status != received) ──
    pending_reserve_bills = (FuelBill.query
                             .outerjoin(FuelReimbursement, FuelBill.reimbursement_id == FuelReimbursement.id)
                             .filter(FuelBill.payment_method == 'reserve')
                             .filter(or_(FuelBill.reimbursement_id.is_(None),
                                        FuelReimbursement.status != 'received'))
                             .order_by(FuelBill.bill_date.desc(), FuelBill.id.desc())
                             .all())
    pending_vehicle_ids = sorted(
        {b.vehicle_id for b in pending_reserve_bills if b.vehicle_id},
        key=lambda vid: vehicle_labels.get(vid, ''))
    pending_meta = {
        b.id: {'status': _pending_status(b), 'filter_key': _category_filter_key(b)}
        for b in pending_reserve_bills
    }

    # ── tab จบแล้ว (P5, §5.4/D4) — reserve ที่ได้เงินคืนแล้ว + card + self ──
    finished_bills = (FuelBill.query
                      .outerjoin(FuelReimbursement, FuelBill.reimbursement_id == FuelReimbursement.id)
                      .filter(or_(
                          FuelBill.payment_method.in_(('card', 'self')),
                          and_(FuelBill.payment_method == 'reserve', FuelReimbursement.status == 'received'),
                      ))
                      .order_by(FuelBill.bill_date.desc(), FuelBill.id.desc())
                      .all())
    finished_vehicle_ids = sorted(
        {b.vehicle_id for b in finished_bills if b.vehicle_id},
        key=lambda vid: vehicle_labels.get(vid, ''))
    finished_meta = {
        b.id: {'status': _finished_status(b), 'filter_key': _category_filter_key(b)}
        for b in finished_bills
    }
    draft_reimbursements = (FuelReimbursement.query
                            .filter_by(status='draft')
                            .order_by(FuelReimbursement.id.desc())
                            .all())
    reimbursement_sources = (ReimbursementSource.query
                             .filter_by(is_active=True)
                             .order_by(ReimbursementSource.is_default.desc(), ReimbursementSource.name)
                             .all())
    all_holders = ExpenseHolder.query.filter_by(is_active=True).order_by(ExpenseHolder.id).all()

    # ── tab ใบเบิกเงิน (P4) — ทุกสถานะ (ไม่กรองปี) + ตารางคืนเงินรายคนต่อใบ ──
    all_reimbursements = FuelReimbursement.query.order_by(FuelReimbursement.id.desc()).all()
    rb_meta = {}
    for rb in all_reimbursements:
        rb_bills = rb.bills
        settlements = ReimbursementSettlement.query.filter_by(reimbursement_id=rb.id).all()
        rb_meta[rb.id] = {
            'bill_count': len(rb_bills),
            'total': sum((float(b.amount) for b in rb_bills), 0.0),
            'settlements': settlements,
            'stale': (rb.status == 'draft' and rb.created_at is not None
                     and (today - rb.created_at.date()).days > 14),
        }

    return render_template(
        'vehicle/admin/admin_fuel.html',
        # data
        vehicles=vehicles,
        drivers=drivers,
        fuel_prices=fuel_prices,
        sources_all=sources_all,
        pivot=pivot,
        pivot_labels=pivot_labels,
        pivot_max=pivot_max,
        # งบทั้งปี (modal ตั้งค่า)
        annual_budget=annual_budget,
        year_used=year_used,
        year_remaining=year_remaining,
        # เงินสำรองรายคน
        my_kpi=my_kpi,
        my_quota_lines=my_quota_lines,
        vehicle_labels=vehicle_labels,
        source_labels=source_labels,
        holder_rows=holder_rows,
        holder_candidate_users=holder_candidate_users,
        pending_reserve_bills=pending_reserve_bills,
        pending_vehicle_ids=pending_vehicle_ids,
        pending_meta=pending_meta,
        finished_bills=finished_bills,
        finished_vehicle_ids=finished_vehicle_ids,
        finished_meta=finished_meta,
        category_labels=CATEGORY_LABEL_TH,
        PAYMENT_LABEL_TH=PAYMENT_LABEL_TH,
        draft_reimbursements=draft_reimbursements,
        reimbursement_sources=reimbursement_sources,
        all_holders=all_holders,
        all_reimbursements=all_reimbursements,
        rb_meta=rb_meta,
        today=today,
        # filters
        f_year=f_year,
        available_years=available_years,
    )


# ─────────────────────────────────────────────
# Bills CRUD — validation (§4.5) อยู่ services/vehicle/fuel_service.py ทั้งหมด (ADR 0001)
# ─────────────────────────────────────────────
def _read_bill_form():
    """parse ค่าจากฟอร์มบิล (ใช้ร่วม create/edit) — คืน dict ตรงกับ kwargs ของ fuel_svc.create_bill/update_bill"""
    return dict(
        bill_date=_parse_date(request.form.get('bill_date')),
        vehicle_id=_parse_int(request.form.get('vehicle_id')),
        driver_id=_parse_int(request.form.get('driver_id')),
        amount=_parse_decimal(request.form.get('amount')),
        method=(request.form.get('payment_method') or '').strip(),
        category=(request.form.get('category') or 'fuel').strip(),
        liters=_parse_decimal(request.form.get('liters'), None),
        mileage=_parse_int(request.form.get('mileage')),
        note=(request.form.get('note') or '').strip() or None,
        paid_by_holder_id=_parse_int(request.form.get('paid_by_holder_id')),
    )


def _flash_bill_warnings(warnings):
    for w in warnings:
        flash(w, 'warning')


@fuel_bp.route('/admin/fuel/bill', methods=['POST'])
@login_required
def create_bill():
    _guard()
    form = _read_bill_form()
    if not (form['bill_date'] and form['driver_id'] and form['amount'] and form['amount'] > 0):
        flash('กรุณากรอกข้อมูลให้ครบ (วันที่, ผู้เติม, จำนวนเงิน, ช่องทาง)', 'danger')
        return redirect(request.referrer or url_for('fuel.admin_fuel'))
    try:
        bill, warnings = fuel_svc.create_bill(actor_id=current_user.id, **form)
        flash('บันทึกบิลเรียบร้อย', 'success')
        _flash_bill_warnings(warnings)
    except ValueError as e:
        db.session.rollback()
        flash(str(e), 'danger')
    except Exception:
        db.session.rollback()
        current_app.logger.exception('create_bill failed')
        flash('เกิดข้อผิดพลาด กรุณาลองใหม่', 'danger')
    return redirect(request.referrer or url_for('fuel.admin_fuel'))


@fuel_bp.route('/admin/fuel/bill/<int:bill_id>/edit', methods=['POST'])
@login_required
def edit_bill(bill_id):
    _guard()
    bill = FuelBill.query.get_or_404(bill_id)
    form = _read_bill_form()
    try:
        bill, warnings = fuel_svc.update_bill(bill, actor_id=current_user.id, **form)
        flash('แก้ไขบิลเรียบร้อย', 'success')
        _flash_bill_warnings(warnings)
    except ValueError as e:
        db.session.rollback()
        flash(str(e), 'danger')
    except Exception:
        db.session.rollback()
        current_app.logger.exception('edit_bill failed')
        flash('เกิดข้อผิดพลาด กรุณาลองใหม่', 'danger')
    return redirect(request.referrer or url_for('fuel.admin_fuel'))


@fuel_bp.route('/admin/fuel/bill/<int:bill_id>/delete', methods=['POST'])
@login_required
def delete_bill(bill_id):
    _guard()
    bill = FuelBill.query.get_or_404(bill_id)
    try:
        fuel_svc.delete_bill(bill, current_user.id)
        flash('ลบบิลเรียบร้อย', 'success')
    except ValueError as e:
        db.session.rollback()
        flash(str(e), 'danger')
    except Exception:
        db.session.rollback()
        current_app.logger.exception('delete_bill failed')
        flash('เกิดข้อผิดพลาด กรุณาลองใหม่', 'danger')
    return redirect(request.referrer or url_for('fuel.admin_fuel'))


@fuel_bp.route('/admin/fuel/bill/<int:bill_id>/detach', methods=['POST'])
@login_required
def detach_bill(bill_id):
    _guard()
    bill = FuelBill.query.get_or_404(bill_id)
    try:
        fuel_svc.detach_bill(bill, current_user.id)
        flash('ถอดบิลออกจากใบเบิกเรียบร้อย', 'success')
    except ValueError as e:
        db.session.rollback()
        flash(str(e), 'danger')
    except Exception:
        db.session.rollback()
        current_app.logger.exception('detach_bill failed')
        flash('เกิดข้อผิดพลาด กรุณาลองใหม่', 'danger')
    return redirect(request.referrer or url_for('fuel.admin_fuel'))


# ─────────────────────────────────────────────
# Reimbursement — ใบเบิกเงิน (tab ใบเบิกเงิน, P4)
# วงจรสถานะ: draft (แก้/ลบ/ใส่-ถอดบิลได้) → submitted (ล็อก, snapshot settlement) → received (ล็อก)
# logic ทั้งหมดอยู่ fuel_service.py (ADR 0001)
# ─────────────────────────────────────────────
@fuel_bp.route('/admin/fuel/reimbursement', methods=['POST'])
@login_required
def create_reimbursement():
    """สร้างใบเบิกเปล่า (draft, ไม่ต้องมีบิล) — ปุ่ม "สร้างใบเบิก" ใน tab ใบเบิกเงิน"""
    _guard()
    rb_no = (request.form.get('reimbursement_no') or '').strip()
    source_id = _parse_int(request.form.get('source_id'))
    note = (request.form.get('note') or '').strip() or None
    try:
        rb = fuel_svc.create_draft_reimbursement(rb_no, source_id, note, current_user.id)
        flash(f'สร้างใบเบิก {rb.reimbursement_no} เรียบร้อย', 'success')
    except ValueError as e:
        db.session.rollback()
        flash(str(e), 'danger')
    except Exception:
        db.session.rollback()
        current_app.logger.exception('create_reimbursement failed')
        flash('เกิดข้อผิดพลาด กรุณาลองใหม่', 'danger')
    return redirect(request.referrer or url_for('fuel.admin_fuel'))


@fuel_bp.route('/admin/fuel/reimbursement/<int:rb_id>/edit', methods=['POST'])
@login_required
def edit_reimbursement(rb_id):
    """แก้เลขที่/แหล่งเบิก/หมายเหตุ — เฉพาะใบ draft"""
    _guard()
    rb = FuelReimbursement.query.get_or_404(rb_id)
    rb_no = (request.form.get('reimbursement_no') or '').strip()
    source_id = _parse_int(request.form.get('source_id'))
    note = (request.form.get('note') or '').strip()
    try:
        fuel_svc.update_reimbursement_meta(rb, reimbursement_no=rb_no, source_id=source_id,
                                           note=note, actor_id=current_user.id)
        flash('แก้ไขใบเบิกเรียบร้อย', 'success')
    except ValueError as e:
        db.session.rollback()
        flash(str(e), 'danger')
    except Exception:
        db.session.rollback()
        current_app.logger.exception('edit_reimbursement failed')
        flash('เกิดข้อผิดพลาด กรุณาลองใหม่', 'danger')
    return redirect(request.referrer or url_for('fuel.admin_fuel'))


@fuel_bp.route('/admin/fuel/reimbursement/<int:rb_id>/submit', methods=['POST'])
@login_required
def submit_reimbursement(rb_id):
    """draft → submitted: snapshot settlement รายคน + ล็อกบิลข้างใน"""
    _guard()
    rb = FuelReimbursement.query.get_or_404(rb_id)
    amount_requested = _parse_decimal(request.form.get('amount_requested'), None)
    try:
        fuel_svc.submit_reimbursement(rb, amount_requested, current_user.id)
        flash(f'ส่งเรื่องใบเบิก {rb.reimbursement_no} เรียบร้อย', 'success')
    except ValueError as e:
        db.session.rollback()
        flash(str(e), 'danger')
    except Exception:
        db.session.rollback()
        current_app.logger.exception('submit_reimbursement failed')
        flash('เกิดข้อผิดพลาด กรุณาลองใหม่', 'danger')
    return redirect(request.referrer or url_for('fuel.admin_fuel'))


@fuel_bp.route('/admin/fuel/reimbursement/<int:rb_id>/receive', methods=['POST'])
@login_required
def receive_reimbursement(rb_id):
    """submitted → received: บันทึกยอดที่ได้เงินจริง"""
    _guard()
    rb = FuelReimbursement.query.get_or_404(rb_id)
    amount_received = _parse_decimal(request.form.get('amount_received'))
    received_at = _parse_date(request.form.get('received_at'), get_bkk_time().date())
    try:
        fuel_svc.receive_reimbursement(rb, amount_received, received_at, current_user.id)
        flash(f'บันทึกได้เงินคืน ใบเบิก {rb.reimbursement_no}', 'success')
    except ValueError as e:
        db.session.rollback()
        flash(str(e), 'danger')
    except Exception:
        db.session.rollback()
        current_app.logger.exception('receive_reimbursement failed')
        flash('เกิดข้อผิดพลาด กรุณาลองใหม่', 'danger')
    return redirect(request.referrer or url_for('fuel.admin_fuel'))


@fuel_bp.route('/admin/fuel/reimbursement/<int:rb_id>/delete', methods=['POST'])
@login_required
def delete_reimbursement(rb_id):
    """ลบใบเบิก — เฉพาะ draft (บิลข้างในกลับสู่ "ใช้ไปแล้ว")"""
    _guard()
    rb = FuelReimbursement.query.get_or_404(rb_id)
    try:
        fuel_svc.delete_draft_reimbursement(rb, current_user.id)
        flash('ลบใบเบิกเรียบร้อย (บิลกลับสู่สถานะใช้ไปแล้ว)', 'success')
    except ValueError as e:
        db.session.rollback()
        flash(str(e), 'danger')
    except Exception:
        db.session.rollback()
        current_app.logger.exception('delete_reimbursement failed')
        flash('เกิดข้อผิดพลาด กรุณาลองใหม่', 'danger')
    return redirect(request.referrer or url_for('fuel.admin_fuel'))


@fuel_bp.route('/admin/fuel/settlement/<int:settlement_id>/settle', methods=['POST'])
@login_required
def settle_holder(settlement_id):
    """คืนเงินให้ผู้สำรอง 1 คน — คืนคนละวันได้ (§5.5 ตารางคืนเงินรายคน)"""
    _guard()
    settlement = ReimbursementSettlement.query.get_or_404(settlement_id)
    settled_at = _parse_date(request.form.get('settled_at'), get_bkk_time().date())
    try:
        fuel_svc.settle_holder(settlement, settled_at, current_user.id)
        flash('บันทึกคืนเงินเรียบร้อย', 'success')
    except ValueError as e:
        db.session.rollback()
        flash(str(e), 'danger')
    except Exception:
        db.session.rollback()
        current_app.logger.exception('settle_holder failed')
        flash('เกิดข้อผิดพลาด กรุณาลองใหม่', 'danger')
    return redirect(request.referrer or url_for('fuel.admin_fuel'))


# ─────────────────────────────────────────────
# tab ค้างเบิก — "ใส่ใบเบิก" (P3): เลือกใบ draft ที่มีอยู่ หรือสร้างใหม่ในตัว
# logic (attach/quota check) อยู่ fuel_service.py ทั้งหมด (ADR 0001)
# ─────────────────────────────────────────────
@fuel_bp.route('/admin/fuel/attach-bills', methods=['POST'])
@login_required
def attach_bills():
    _guard()
    rb_id = _parse_int(request.form.get('reimbursement_id'))
    bill_ids = [int(x) for x in request.form.getlist('bill_ids') if x.isdigit()]
    if not (rb_id and bill_ids):
        flash('กรุณาเลือกบิลและใบเบิก', 'danger')
        return redirect(request.referrer or url_for('fuel.admin_fuel'))
    rb = FuelReimbursement.query.get_or_404(rb_id)
    try:
        warnings = fuel_svc.attach_bills_to_reimbursement(bill_ids, rb, current_user.id)
        flash(f'ใส่ {len(bill_ids)} บิลเข้าใบเบิก {rb.reimbursement_no} เรียบร้อย', 'success')
        _flash_bill_warnings(warnings)
    except ValueError as e:
        db.session.rollback()
        flash(str(e), 'danger')
    except Exception:
        db.session.rollback()
        current_app.logger.exception('attach_bills failed')
        flash('เกิดข้อผิดพลาด กรุณาลองใหม่', 'danger')
    return redirect(request.referrer or url_for('fuel.admin_fuel'))


@fuel_bp.route('/admin/fuel/reimbursement/draft', methods=['POST'])
@login_required
def create_reimbursement_draft():
    """"+ สร้างใบเบิกใหม่" ในตัว modal ใส่ใบเบิก — ต่างจาก create_reimbursement เดิม
    (legacy, มี submitted_at ทันที) ตัวนี้สร้าง status='draft' + source_id (FK) ล้วนๆ"""
    _guard()
    rb_no = (request.form.get('reimbursement_no') or '').strip()
    source_id = _parse_int(request.form.get('source_id'))
    note = (request.form.get('note') or '').strip() or None
    bill_ids = [int(x) for x in request.form.getlist('bill_ids') if x.isdigit()]
    if not (rb_no and bill_ids):
        flash('กรุณาเลือกบิลและกรอกเลขใบเบิก', 'danger')
        return redirect(request.referrer or url_for('fuel.admin_fuel'))
    try:
        rb = fuel_svc.create_draft_with_bills(bill_ids, rb_no, source_id, note, current_user.id)
        flash(f'สร้างใบเบิก {rb.reimbursement_no} + ใส่ {len(bill_ids)} บิลเรียบร้อย', 'success')
    except ValueError as e:
        db.session.rollback()
        flash(str(e), 'danger')
    except Exception:
        db.session.rollback()
        current_app.logger.exception('create_reimbursement_draft failed')
        flash('เกิดข้อผิดพลาด กรุณาลองใหม่', 'danger')
    return redirect(request.referrer or url_for('fuel.admin_fuel'))


# ─────────────────────────────────────────────
# API โควตา — ให้ JS ตอนกรอกบิล (§4.2 จุดที่ 2)
# ─────────────────────────────────────────────
@fuel_bp.route('/api/fuel/quota', methods=['GET'])
@login_required
def api_quota():
    _guard()
    vid = _parse_int(request.args.get('vehicle_id'))
    d   = _parse_date(request.args.get('bill_date'), get_bkk_time().date())
    if not vid:
        return jsonify({'ok': True, 'card': None})
    st = fuel_svc.quota_status(vid, 'card', d.year, d.month)
    if st is None:
        return jsonify({'ok': True, 'card': None})
    return jsonify({'ok': True, 'card': {
        'limit': float(st['limit']), 'used': float(st['used']), 'remaining': float(st['remaining']),
    }})


# ─────────────────────────────────────────────
# Fuel price (effective-dated)
# ─────────────────────────────────────────────
@fuel_bp.route('/admin/fuel/price', methods=['POST'])
@login_required
def add_price():
    _guard()
    eff_date = _parse_date(request.form.get('effective_date'))
    price    = _parse_decimal(request.form.get('price_per_liter'))
    note     = (request.form.get('note') or '').strip() or None

    if not eff_date or price <= 0:
        flash('กรุณากรอกวันที่และราคา/ลิตร', 'danger')
        return redirect(request.referrer or url_for('fuel.admin_fuel'))

    existing = FuelPrice.query.filter_by(effective_date=eff_date).first()
    if existing:
        existing.price_per_liter = price
        existing.note = note
        existing.created_by = current_user.id
        flash('อัปเดตราคา/ลิตร สำหรับวันที่นี้แล้ว', 'success')
    else:
        db.session.add(FuelPrice(
            effective_date=eff_date,
            price_per_liter=price,
            note=note,
            created_by=current_user.id,
        ))
        flash('เพิ่มราคา/ลิตร เรียบร้อย', 'success')
    db.session.commit()
    return redirect(request.referrer or url_for('fuel.admin_fuel'))


@fuel_bp.route('/admin/fuel/price/<int:price_id>/delete', methods=['POST'])
@login_required
def delete_price(price_id):
    _guard()
    p = FuelPrice.query.get_or_404(price_id)
    db.session.delete(p)
    db.session.commit()
    flash('ลบราคา/ลิตรเรียบร้อย', 'success')
    return redirect(request.referrer or url_for('fuel.admin_fuel'))


# ─────────────────────────────────────────────
# Annual budget config (uses SystemConfig key='fuel_annual_budget')
# ─────────────────────────────────────────────
@fuel_bp.route('/admin/fuel/annual-budget', methods=['POST'])
@login_required
def set_annual_budget():
    _guard()
    amount = _parse_decimal(request.form.get('amount'))
    if amount < 0:
        flash('จำนวนเงินไม่ถูกต้อง', 'danger')
        return redirect(request.referrer or url_for('fuel.admin_fuel'))
    SystemConfig.set('fuel_annual_budget', float(amount))
    flash('ตั้งงบน้ำมันรายปีเรียบร้อย', 'success')
    return redirect(request.referrer or url_for('fuel.admin_fuel'))


# ─────────────────────────────────────────────
# แหล่งเบิก (ReimbursementSource) — CRUD (review 2026-08-10 #8)
# logic อยู่ fuel_service.py ทั้งหมด (ADR 0001) — route แค่ parse form → เรียก service → flash
# ─────────────────────────────────────────────
@fuel_bp.route('/admin/fuel/source', methods=['POST'])
@login_required
def source_create():
    _guard()
    try:
        src = fuel_svc.create_source(request.form.get('name'),
                                     bool(request.form.get('is_default')), current_user.id)
        flash(f'เพิ่มแหล่งเบิก {src.name} เรียบร้อย', 'success')
    except ValueError as e:
        db.session.rollback()
        flash(str(e), 'danger')
    except Exception:
        db.session.rollback()
        current_app.logger.exception('source_create failed')
        flash('เกิดข้อผิดพลาด กรุณาลองใหม่', 'danger')
    return redirect(request.referrer or url_for('fuel.admin_fuel'))


@fuel_bp.route('/admin/fuel/source/<int:source_id>/toggle-active', methods=['POST'])
@login_required
def source_toggle_active(source_id):
    _guard()
    source = ReimbursementSource.query.get_or_404(source_id)
    try:
        fuel_svc.toggle_source_active(source, current_user.id)
        flash('เปิด/ปิดใช้งานแหล่งเบิกเรียบร้อย', 'success')
    except ValueError as e:
        db.session.rollback()
        flash(str(e), 'danger')
    except Exception:
        db.session.rollback()
        current_app.logger.exception('source_toggle_active failed')
        flash('เกิดข้อผิดพลาด กรุณาลองใหม่', 'danger')
    return redirect(request.referrer or url_for('fuel.admin_fuel'))


@fuel_bp.route('/admin/fuel/source/<int:source_id>/set-default', methods=['POST'])
@login_required
def source_set_default(source_id):
    _guard()
    source = ReimbursementSource.query.get_or_404(source_id)
    try:
        fuel_svc.set_default_source(source, current_user.id)
        flash(f'ตั้ง {source.name} เป็นแหล่งเบิกค่าเริ่มต้นเรียบร้อย', 'success')
    except ValueError as e:
        db.session.rollback()
        flash(str(e), 'danger')
    except Exception:
        db.session.rollback()
        current_app.logger.exception('source_set_default failed')
        flash('เกิดข้อผิดพลาด กรุณาลองใหม่', 'danger')
    return redirect(request.referrer or url_for('fuel.admin_fuel'))


@fuel_bp.route('/admin/fuel/source/<int:source_id>/delete', methods=['POST'])
@login_required
def source_delete(source_id):
    _guard()
    source = ReimbursementSource.query.get_or_404(source_id)
    try:
        fuel_svc.delete_source(source, current_user.id)
        flash('ลบแหล่งเบิกเรียบร้อย', 'success')
    except ValueError as e:
        db.session.rollback()
        flash(str(e), 'danger')
    except Exception:
        db.session.rollback()
        current_app.logger.exception('source_delete failed')
        flash('เกิดข้อผิดพลาด กรุณาลองใหม่', 'danger')
    return redirect(request.referrer or url_for('fuel.admin_fuel'))


# ─────────────────────────────────────────────
# เจ้าหน้าที่ผู้สำรองเงิน (ExpenseHolder) — Phase 2
# logic เงินอยู่ services/vehicle/fuel_service.py ทั้งหมด (ADR 0001) —
# route แค่ parse form → เรียก service → flash ผลลัพธ์
# ─────────────────────────────────────────────
@fuel_bp.route('/admin/fuel/holder', methods=['POST'])
@login_required
def holder_create():
    _guard()
    try:
        uid  = _parse_int(request.form.get('user_id'))
        amt  = _parse_decimal(request.form.get('float_amount'))
        note = (request.form.get('note') or '').strip()
        if not uid:
            raise ValueError('กรุณาเลือกผู้ใช้')
        holder = fuel_svc.create_holder(uid, amt, note, current_user.id)
        name = holder.user.full_name or holder.user.username
        flash(f'เพิ่มเจ้าหน้าที่ {name} เป็นผู้สำรองเงินเรียบร้อย', 'success')
    except ValueError as e:
        db.session.rollback()
        flash(str(e), 'danger')
    except Exception:
        db.session.rollback()
        current_app.logger.exception('holder_create failed')
        flash('เกิดข้อผิดพลาด กรุณาลองใหม่', 'danger')
    return redirect(request.referrer or url_for('fuel.admin_fuel'))


@fuel_bp.route('/admin/fuel/holder/<int:holder_id>/set-float', methods=['POST'])
@login_required
def holder_update(holder_id):
    _guard()
    holder = ExpenseHolder.query.get_or_404(holder_id)
    try:
        amt  = _parse_decimal(request.form.get('float_amount'))
        note = (request.form.get('note') or '').strip()
        fuel_svc.set_float(holder, amt, note, current_user.id)
        flash('ตั้งวงเงินสำรองเรียบร้อย', 'success')
    except ValueError as e:
        db.session.rollback()
        flash(str(e), 'danger')
    except Exception:
        db.session.rollback()
        current_app.logger.exception('holder_update failed')
        flash('เกิดข้อผิดพลาด กรุณาลองใหม่', 'danger')
    return redirect(request.referrer or url_for('fuel.admin_fuel'))


@fuel_bp.route('/admin/fuel/holder/<int:holder_id>/topup', methods=['POST'])
@login_required
def holder_topup(holder_id):
    _guard()
    holder = ExpenseHolder.query.get_or_404(holder_id)
    try:
        amt  = _parse_decimal(request.form.get('amount'))
        note = (request.form.get('note') or '').strip()
        fuel_svc.top_up(holder, amt, note, current_user.id)
        flash('เติมเงินสำรองเรียบร้อย', 'success')
    except ValueError as e:
        db.session.rollback()
        flash(str(e), 'danger')
    except Exception:
        db.session.rollback()
        current_app.logger.exception('holder_topup failed')
        flash('เกิดข้อผิดพลาด กรุณาลองใหม่', 'danger')
    return redirect(request.referrer or url_for('fuel.admin_fuel'))


@fuel_bp.route('/admin/fuel/holder/<int:holder_id>/adjust', methods=['POST'])
@login_required
def holder_adjust(holder_id):
    _guard()
    holder = ExpenseHolder.query.get_or_404(holder_id)
    try:
        change = _parse_decimal(request.form.get('change_amount'))
        note   = (request.form.get('note') or '').strip()
        fuel_svc.adjust_float(holder, change, note, current_user.id)
        flash('ปรับวงเงินสำรองเรียบร้อย', 'success')
    except ValueError as e:
        db.session.rollback()
        flash(str(e), 'danger')
    except Exception:
        db.session.rollback()
        current_app.logger.exception('holder_adjust failed')
        flash('เกิดข้อผิดพลาด กรุณาลองใหม่', 'danger')
    return redirect(request.referrer or url_for('fuel.admin_fuel'))


@fuel_bp.route('/admin/fuel/holder/<int:holder_id>/count', methods=['POST'])
@login_required
def holder_count(holder_id):
    _guard()
    holder = ExpenseHolder.query.get_or_404(holder_id)
    try:
        counted = _parse_decimal(request.form.get('counted_amount'))
        note    = (request.form.get('note') or '').strip()
        variance = fuel_svc.count_cash(holder, counted, note, current_user.id)
        if variance == 0:
            flash('นับเงินตรงกับคงเหลือในระบบพอดี', 'success')
        else:
            sign = '+' if variance > 0 else ''
            flash(f'นับเงินจริงต่างจากระบบ {sign}{variance:,.2f} บาท — บันทึกไว้แล้ว '
                  f'กด "ปรับยอด" ถ้าต้องการแก้ให้ตรง', 'warning')
    except ValueError as e:
        db.session.rollback()
        flash(str(e), 'danger')
    except Exception:
        db.session.rollback()
        current_app.logger.exception('holder_count failed')
        flash('เกิดข้อผิดพลาด กรุณาลองใหม่', 'danger')
    return redirect(request.referrer or url_for('fuel.admin_fuel'))


# ─────────────────────────────────────────────
# JSON helper (mileage page badge — phase 3 will consume)
# ─────────────────────────────────────────────
@fuel_bp.route('/api/fuel/bill-by-mileage', methods=['GET'])
@login_required
def api_bill_by_mileage():
    """Lookup FuelBill by (vehicle_id, mileage). Returns list (could be 0..N)."""
    _guard()
    vid = _parse_int(request.args.get('vehicle_id'))
    m   = _parse_int(request.args.get('mileage'))
    if not (vid and m):
        return jsonify([])
    bills = FuelBill.query.filter_by(vehicle_id=vid, mileage=m).all()
    return jsonify([{
        'id': b.id,
        'date': b.bill_date.isoformat() if b.bill_date else None,
        'amount': float(b.amount),
        'payment_method': b.payment_method,
    } for b in bills])


# ─────────────────────────────────────────────
# Excel export — GET /admin/fuel/export/excel
# ─────────────────────────────────────────────
@fuel_bp.route('/admin/fuel/export/excel', methods=['GET'])
@login_required
def export_excel():
    """3-sheet workbook: บิล / ใบเบิก / Pivot — honors year/month/vehicle/driver filters."""
    _guard()

    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        flash('ไม่พบ openpyxl — รัน: pip install openpyxl', 'danger')
        return redirect(url_for('fuel.admin_fuel'))
    from flask import send_file

    today = get_bkk_time().date()
    f_year, f_month, f_veh, f_drv = _read_filters(request.args)

    bills = _filtered_bills_query(f_year, f_month, f_veh, f_drv).all()

    reimbursements = (FuelReimbursement.query
                      .filter(
                          (extract('year', FuelReimbursement.submitted_at) == f_year) |
                          (extract('year', FuelReimbursement.created_at) == f_year)
                      )
                      .order_by(FuelReimbursement.id.desc())
                      .all())

    pivot_rows = (db.session.query(
                      FuelBill.vehicle_id,
                      extract('month', FuelBill.bill_date).label('m'),
                      func.sum(FuelBill.amount).label('total'))
                  .filter(extract('year', FuelBill.bill_date) == f_year)
                  .group_by(FuelBill.vehicle_id, 'm')
                  .all())
    pivot = {}  # {vehicle_id: {month: total}}
    for vid, m, total in pivot_rows:
        pivot.setdefault(vid, {})[int(m)] = float(total or 0)
    vehicles = {v.id: v for v in Vehicle.query.order_by(Vehicle.license_plate).all()}

    # Workbook style helpers
    hdr_fill = PatternFill('solid', fgColor='4F46E5')
    hdr_font = Font(bold=True, color='FFFFFF', name='Sarabun')
    bold     = Font(bold=True, name='Sarabun')
    thin     = Side(style='thin', color='E4E4E7')
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)
    money_fmt = '#,##0.00'

    def write_header(ws, headers, widths):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = hdr_font; c.fill = hdr_fill
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = border
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[1].height = 22

    wb = openpyxl.Workbook()

    # ── Sheet 1: บิล ───────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'บิล'
    h1 = ['วันที่','รถ','ผู้เติม','จำนวน(฿)','ช่องทาง','ไมล์','สถานะ','เลขใบเบิก']
    write_header(ws1, h1, [12, 22, 22, 14, 12, 12, 12, 18])
    ri = 2
    bills_total = 0.0
    for b in bills:
        veh = b.vehicle
        veh_str = f"{veh.license_plate} {veh.brand} {veh.model}".strip() if veh else '-'
        drv_str = b.driver.name if b.driver else '-'
        rb_no = b.reimbursement.reimbursement_no if b.reimbursement else ''
        amt = float(b.amount or 0)
        bills_total += amt
        row = [
            b.bill_date.strftime('%d/%m/%Y') if b.bill_date else '',
            veh_str, drv_str, amt,
            PAYMENT_LABEL_TH.get(b.payment_method, b.payment_method or '-'),
            b.mileage if b.mileage is not None else '',
            _bill_status(b),
            rb_no,
        ]
        for ci, val in enumerate(row, 1):
            c = ws1.cell(row=ri, column=ci, value=val)
            c.border = border
            c.alignment = Alignment(horizontal='center' if ci in (1, 4, 5, 6, 7) else 'left')
            if ci == 4: c.number_format = money_fmt
            if ri % 2 == 0: c.fill = PatternFill('solid', fgColor='FAFAFA')
        ri += 1
    # totals
    if bills:
        c = ws1.cell(row=ri, column=3, value='รวม'); c.font = bold; c.alignment = Alignment(horizontal='right')
        c = ws1.cell(row=ri, column=4, value=round(bills_total, 2)); c.font = bold; c.number_format = money_fmt

    # ── Sheet 2: ใบเบิก ─────────────────────────────────────────────
    ws2 = wb.create_sheet('ใบเบิก')
    h2 = ['เลขใบเบิก','แหล่งเบิก','วันส่ง','วันได้เงิน','จำนวนบิล','รวมเงิน(฿)']
    write_header(ws2, h2, [18, 22, 12, 12, 12, 16])
    ri = 2
    rb_total = 0.0
    for rb in reimbursements:
        rb_bills = rb.bills if hasattr(rb, 'bills') else []
        rb_sum = sum((float(b.amount or 0) for b in rb_bills), 0.0)
        rb_total += rb_sum
        row = [
            rb.reimbursement_no or '',
            rb.source or '-',
            rb.submitted_at.strftime('%d/%m/%Y') if rb.submitted_at else '-',
            rb.received_at.strftime('%d/%m/%Y')  if rb.received_at  else 'รอเงิน',
            len(rb_bills),
            rb_sum,
        ]
        for ci, val in enumerate(row, 1):
            c = ws2.cell(row=ri, column=ci, value=val)
            c.border = border
            c.alignment = Alignment(horizontal='center' if ci in (1, 3, 4, 5, 6) else 'left')
            if ci == 6: c.number_format = money_fmt
            if ri % 2 == 0: c.fill = PatternFill('solid', fgColor='FAFAFA')
        ri += 1
    if reimbursements:
        c = ws2.cell(row=ri, column=5, value='รวม'); c.font = bold; c.alignment = Alignment(horizontal='right')
        c = ws2.cell(row=ri, column=6, value=round(rb_total, 2)); c.font = bold; c.number_format = money_fmt

    # ── Sheet 3: Pivot ─────────────────────────────────────────────
    ws3 = wb.create_sheet('Pivot')
    h3 = ['รถ'] + MONTH_LABEL_TH + ['รวมทั้งปี']
    write_header(ws3, h3, [22] + [12]*12 + [16])
    ri = 2
    month_totals = [0.0]*12
    grand_total = 0.0
    # sort vehicles by license plate for stable order
    sorted_vids = sorted(pivot.keys(), key=lambda vid: (vehicles[vid].license_plate if vid in vehicles else ''))
    for vid in sorted_vids:
        veh = vehicles.get(vid)
        veh_str = f"{veh.license_plate} {veh.brand} {veh.model}".strip() if veh else f"#{vid}"
        row_total = 0.0
        c = ws3.cell(row=ri, column=1, value=veh_str)
        c.border = border; c.alignment = Alignment(horizontal='left')
        if ri % 2 == 0: c.fill = PatternFill('solid', fgColor='FAFAFA')
        for mi in range(1, 13):
            v = pivot[vid].get(mi, 0.0)
            row_total += v
            month_totals[mi-1] += v
            c = ws3.cell(row=ri, column=1+mi, value=(v if v else None))
            c.border = border; c.alignment = Alignment(horizontal='right')
            c.number_format = money_fmt
            if ri % 2 == 0: c.fill = PatternFill('solid', fgColor='FAFAFA')
        grand_total += row_total
        c = ws3.cell(row=ri, column=14, value=row_total)
        c.border = border; c.alignment = Alignment(horizontal='right')
        c.number_format = money_fmt; c.font = bold
        if ri % 2 == 0: c.fill = PatternFill('solid', fgColor='FAFAFA')
        ri += 1
    # footer: รวมต่อเดือน
    if pivot:
        c = ws3.cell(row=ri, column=1, value='รวมต่อเดือน')
        c.font = bold; c.border = border; c.alignment = Alignment(horizontal='right')
        for mi in range(1, 13):
            c = ws3.cell(row=ri, column=1+mi, value=month_totals[mi-1] or None)
            c.font = bold; c.border = border; c.alignment = Alignment(horizontal='right')
            c.number_format = money_fmt
        c = ws3.cell(row=ri, column=14, value=grand_total)
        c.font = bold; c.border = border; c.alignment = Alignment(horizontal='right')
        c.number_format = money_fmt

    # Filename: fuel_<year>_<month>_<vehicle>_<date>.xlsx
    month_part = f"{f_month:02d}" if f_month else 'all'
    if f_veh and f_veh in vehicles:
        veh_part = ''.join(ch for ch in vehicles[f_veh].license_plate if ch.isalnum()) or 'veh'
    else:
        veh_part = 'all'
    fname = f"fuel_{f_year}_{month_part}_{veh_part}_{today.strftime('%Y%m%d')}.xlsx"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
