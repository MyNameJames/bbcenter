from flask import render_template, request, redirect, url_for, flash, jsonify, session
from flask_login import login_required, current_user
from models import (db, get_bkk_time, Vehicle, Driver, VehicleMileage, VehicleBooking,
                    VehicleBudgetLog, OTRateConfig, DriverOT, DriverOTSlot)
from sqlalchemy import extract, func
from datetime import datetime, date
from domain.vehicle.ot import build_slot, RATE_HOURLY, RATE_FLAT_DAY
import services.vehicle.budget_service as budget_svc
import services.vehicle.mileage_service as mileage_svc
from views.vehicle.vehicle_common import (
    vehicle_bp, adminfleet_bp, admincost_bp, driver_bp,
    is_vehicle_admin,
    EXPENSE_CATEGORIES, TH_MONTHS, _fmt_date_th, _build_budget_subs,
)


def _apply_budget_filter(q, f_budget_type, f_budget_sub):
    """กรอง DriverOT ตามงบ (derive จาก booking ที่ผูก) — join VehicleBooking.
    standalone OT (booking_id=None) จะหลุดเมื่อ filter งบ active เพราะไม่มีงบ"""
    if f_budget_type not in ('central', 'department', 'personal'):
        return q
    q = q.join(VehicleBooking, DriverOT.booking_id == VehicleBooking.id) \
         .filter(VehicleBooking.expense_type == f_budget_type)
    if f_budget_sub and f_budget_type == 'central':
        q = q.filter(VehicleBooking.central_category == f_budget_sub)
    elif f_budget_sub and f_budget_type == 'department':
        q = q.filter(VehicleBooking.trip_department == f_budget_sub)
    return q


def _parse_ot_slots(form, driver_id=None, on_date=None, exclude_ot_id=None):
    """แปลง slot_cfg[]/slot_start[]/slot_end[] จากฟอร์ม modal → list[DriverOTSlot]
    (derive label/rate จาก OTRateConfig — snapshot ลง slot). ใช้ร่วม ot_create + ot_edit

    เงิน/ชั่วโมงคิดที่ domain/vehicle/ot.py::build_slot() ตัวเดียวกับที่ระบบใช้ตอนปิดทริป
    (2026-07-28) — ห้าม inline สูตรที่นี่ กติกาเดียวกับค่าน้ำมันที่ทุกทางเรียก calc_fuel_cost()

    driver_id/on_date (2026-08-07) = ใช้เช็กว่า band เหมาจ่าย (flat_day) ถูกเก็บเงินไปแล้ว
    ในวันนั้นหรือยัง — แอดมินกรอกเองก็ต้องอยู่ใต้กฎ "เหมาจ่ายต่อวัน" เหมือนทางอัตโนมัติ
    (กติกา 6) ไม่ส่งมา = ไม่เช็ก (คิดเต็มทุก slot เหมือนเดิม)
    """
    claimed = (mileage_svc.claimed_flat_configs(driver_id, on_date, exclude_ot_id)
               if driver_id and on_date else frozenset())
    slot_cfgids = form.getlist('slot_cfg[]')
    slot_starts = form.getlist('slot_start[]')
    slot_ends   = form.getlist('slot_end[]')
    slots = []
    for i, cfg_raw in enumerate(slot_cfgids):
        try:
            start  = slot_starts[i]; end = slot_ends[i]
            cfg_id = int(cfg_raw) if cfg_raw else None
            cfg    = OTRateConfig.query.get(cfg_id) if cfg_id else None
            if not cfg or not start or not end:
                continue
            charge = not (cfg.rate_type == RATE_FLAT_DAY and cfg_id in claimed)
            spec = build_slot(cfg.label, start, end, float(cfg.rate), cfg_id,
                              rate_type=cfg.rate_type, charge=charge)
            if not spec:
                continue
            slots.append(DriverOTSlot(
                rate_config_id=spec['config_id'], slot_label=spec['label'],
                start_time=spec['start_time'], end_time=spec['end_time'],
                hours=spec['hours'], rate=spec['rate'], amount=spec['amount'],
            ))
        except (ValueError, IndexError):
            continue
    return slots


def _wants_json():
    """row action ถูกเรียกผ่าน AJAX (vehicle_ot.js) → ตอบ JSON ไม่ flash/redirect"""
    return request.headers.get('X-Requested-With') == 'fetch'


@admincost_bp.route('/vehicle/mileage/override-fuel', methods=['POST'])
@login_required
def override_fuel():
    if not is_vehicle_admin():
        flash('คุณไม่มีสิทธิ์', 'danger')
        return redirect(url_for('vehicle.index'))
    booking_id = int(request.form.get('booking_id'))
    fuel_cost  = float(request.form.get('fuel_cost', 0))
    mileage = VehicleMileage.query.filter_by(booking_id=booking_id).first()
    if not mileage:
        mileage = VehicleMileage(booking_id=booking_id, noted_by=current_user.id)
        db.session.add(mileage)

    mileage_svc.override_fuel_cost(mileage, fuel_cost, actor_username=current_user.username)

    db.session.commit()
    flash(f'Override ค่าน้ำมัน #{booking_id} เป็น {fuel_cost:,.2f} บาท เรียบร้อย', 'success')
    return redirect(request.referrer or url_for('admincost.cost_summary'))



# central_category code → Thai label (สำหรับ col งบ)
_CENTRAL_LABELS = {c['key']: c['label'] for c in EXPENSE_CATEGORIES.get('central', [])}


def _ot_budget_label(booking):
    """คืน (label, sub) สำหรับ col งบ — อิง booking.expense_type (เหมือนหน้า mileage)"""
    if not booking:
        return ('—', '')
    et = booking.expense_type or ''
    if et == 'central':
        return ('ส่วนกลาง', _CENTRAL_LABELS.get(booking.central_category, booking.central_category or ''))
    if et == 'department':
        return ('ส่วนกอง', booking.trip_department or '')
    if et == 'personal':
        return ('จ่ายเอง', '')
    return ('—', '')


def _ot_receipt_dict(ot):
    """ot → dict ให้ buildReceiptPage()/list renderer ฝั่ง JS ใช้ (vehicle_ot.js) — โครง field
    เดียวกับ otCostData JSON blob บนหน้า (ยังไม่รวมเป็นจุดเดียวกับ Jinja blob เดิม — คนละ endpoint
    ใช้กันคนละที่ ผลจึงต้องตรงกันเป๊ะ ถ้าจะแก้ field ต้องแก้คู่กับ block scripts ในเทมเพลตด้วย)"""
    d = ot.driver
    return {
        'id': ot.id, 'ot_number': ot.ot_number, 'driver_id': ot.driver_id,
        'driver_name': d.name if d else '',
        'driver_phone': d.phone if d else '',
        'driver_national_id': d.national_id if d else '',
        'driver_addr_line': d.addr_line if d else '',
        'driver_addr_subdistrict': d.addr_subdistrict if d else '',
        'driver_addr_district': d.addr_district if d else '',
        'driver_addr_province': d.addr_province if d else '',
        'driver_addr_postal': d.addr_postal if d else '',
        'driver_id_card_image': (url_for('static', filename='uploads/driver/' + d.id_card_image)
                                  if d and d.id_card_image else ''),
        'date': ot.date.strftime('%Y-%m-%d'),
        'date_display': f"{ot.date.day:02d} {TH_MONTHS[ot.date.month]} {ot.date.year + 543}",
        'destination': ot.booking.destination if ot.booking else '',
        'total_hours': float(ot.total_hours), 'total_amount': float(ot.total_amount),
        'status': ot.status, 'no_receipt': bool(ot.no_receipt), 'note': ot.note or '',
        'slots': [{'label': s.slot_label, 'start': s.start_time, 'end': s.end_time,
                   'rate': float(s.rate), 'hours': float(s.hours), 'amount': float(s.amount),
                   'cfg_id': s.rate_config_id} for s in ot.slots],
    }


def _build_ot_pivot(from_year):
    pivot_rows = (db.session.query(
        DriverOT.driver_id,
        extract('month', DriverOT.date).label('month'),
        func.sum(DriverOT.total_hours).label('hours'),
        func.sum(DriverOT.total_amount).label('amount'),
    ).filter(
        DriverOT.is_deleted == False,
        extract('year', DriverOT.date) == from_year,
    ).group_by(DriverOT.driver_id, extract('month', DriverOT.date))
    .all())

    ot_pivot        = {}
    ot_pivot_labels = {}
    for row in pivot_rows:
        did = row.driver_id
        m   = int(row.month)
        if did not in ot_pivot:
            ot_pivot[did] = {}
            drv = Driver.query.get(did)
            ot_pivot_labels[did] = drv.name if drv else str(did)
        ot_pivot[did][m] = {'hours': float(row.hours or 0), 'amount': float(row.amount or 0)}

    row_totals = {
        did: {'hours': sum(v['hours'] for v in months.values()),
              'amount': sum(v['amount'] for v in months.values())}
        for did, months in ot_pivot.items()
    }
    col_totals = {
        m: {'hours':  sum(ot_pivot[did].get(m, {}).get('hours',  0) for did in ot_pivot),
            'amount': sum(ot_pivot[did].get(m, {}).get('amount', 0) for did in ot_pivot)}
        for m in range(1, 13)
    }
    grand_hours  = sum(t['hours']  for t in row_totals.values())
    grand_amount = sum(t['amount'] for t in row_totals.values())
    return ot_pivot, ot_pivot_labels, row_totals, col_totals, grand_hours, grand_amount


@admincost_bp.route('/admin/cost', methods=['GET'])
@login_required
def cost_summary():
    if not is_vehicle_admin():
        flash('คุณไม่มีสิทธิ์เข้าหน้านี้', 'danger')
        return redirect(url_for('vehicle.index'))

    now        = get_bkk_time()
    from_month = int(request.args.get('from_month', now.month))
    from_year  = int(request.args.get('from_year',  now.year))
    to_month   = int(request.args.get('to_month',   now.month))
    to_year    = int(request.args.get('to_year',    now.year))

    # แท็บ "ทั้งหมด" — OT ที่ไม่ถูกลบทั้งหมด ไม่มี filter เดือน/คนขับ/งบ/สถานะ (ตัดออก 2026-08-08
    # ตามคำขอ — KPI strip/แถบอัตรา OT/ตารางแยกประเภทงาน ตัดออกด้วย เก็บไว้แค่ปุ่ม Excel ซึ่งยัง
    # scope ตามเดือนปัจจุบัน (from_month/from_year ด้านบน) ไม่ได้ตามรายการที่เห็นทั้งหมด — ตั้งใจ
    # ไม่แตะ cost_export ตอนนี้)
    ots = (DriverOT.query
           .filter(DriverOT.is_deleted.is_(False))
           .order_by(DriverOT.date.desc())
           .all())
    for o in ots:
        o.budget_label, o.budget_sub = _ot_budget_label(o.booking)

    drivers      = Driver.query.order_by(Driver.name).all()
    rate_configs = OTRateConfig.query.filter_by(is_active=True).order_by(OTRateConfig.sort_order).all()

    # แท็บ "ผู้ใช้จ่ายเอง" — รายการทั้งหมดไม่มี filter (ตามคำขอ) คนละชุดกับ ots ด้านบน
    # ไม่ต้อง enrich budget_label (ตารางไม่มีคอลัมน์ งบ)
    no_receipt_ots = (DriverOT.query
                       .filter(DriverOT.no_receipt.is_(True), DriverOT.is_deleted.is_(False))
                       .order_by(DriverOT.date.desc())
                       .all())

    ot_pivot, ot_pivot_labels, row_totals, col_totals, grand_hours, grand_amount = \
        _build_ot_pivot(from_year)

    return render_template('vehicle/admin/vehicle_cost.html',
        ots=ots, drivers=drivers, rate_configs=rate_configs, no_receipt_ots=no_receipt_ots,
        from_month=from_month, from_year=from_year,
        to_month=to_month, to_year=to_year, now=now,
        ot_pivot=ot_pivot, ot_pivot_labels=ot_pivot_labels,
        ot_pivot_row_totals=row_totals, ot_pivot_col_totals=col_totals,
        ot_grand_hours=grand_hours, ot_grand_amount=grand_amount,
    )



@admincost_bp.route('/admin/ot/slip', methods=['GET'])
@login_required
def ot_slip_data():
    """OT ของคนขับคนเดียว เดือนเดียว → JSON ให้แท็บ "ใบจ่ายจริง" render list + receipt preview
    (page contract redesign, 2026-08-08 — view-only, ไม่มี batch-payment entity)"""
    if not is_vehicle_admin():
        return jsonify(ok=False, msg='คุณไม่มีสิทธิ์'), 403
    driver_id = request.args.get('driver_id', type=int)
    year  = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    if not driver_id or not year or not month or not (1 <= month <= 12):
        return jsonify(ok=False, msg='ข้อมูลไม่ครบ'), 400

    from_date = date(year, month, 1)
    to_date   = date(year + 1, 1, 1) if month == 12 else date(year, month + 1, 1)
    ots = (DriverOT.query
           .filter(DriverOT.driver_id == driver_id,
                   DriverOT.date >= from_date, DriverOT.date < to_date,
                   DriverOT.is_deleted.is_(False))
           .order_by(DriverOT.date).all())
    return jsonify(ok=True, items=[_ot_receipt_dict(ot) for ot in ots])


@admincost_bp.route('/admin/ot/<int:ot_id>/mark_paid', methods=['POST'])
@login_required
def ot_mark_paid(ot_id):
    """Toggle จ่าย/ยังไม่จ่าย — ไม่ต้อง approve ก่อน (2026-06-08 เลิกใช้ step อนุมัติ)"""
    if not is_vehicle_admin():
        flash('คุณไม่มีสิทธิ์', 'danger')
        return redirect(url_for('vehicle.index'))
    ot = DriverOT.query.get_or_404(ot_id)
    if ot.status == 'paid':
        ot.status     = 'unpaid'
        ot.paid_by_id = None
        ot.paid_at    = None
        flash(f'ยกเลิกการจ่าย {ot.ot_number} แล้ว', 'success')
    else:
        ot.status     = 'paid'
        ot.paid_by_id = current_user.id
        ot.paid_at    = get_bkk_time()
        flash(f'บันทึกการจ่าย {ot.ot_number} เรียบร้อย', 'success')
    db.session.commit()
    if _wants_json():
        return jsonify(ok=True)
    return redirect(request.referrer or url_for('admincost.cost_summary'))



@admincost_bp.route('/admin/ot/<int:ot_id>/toggle_no_receipt', methods=['POST'])
@login_required
def ot_toggle_no_receipt(ot_id):
    """Toggle 'ผู้ใช้จ่ายเอง' = OT ที่ไม่ต้องออกใบเสร็จ"""
    if not is_vehicle_admin():
        flash('คุณไม่มีสิทธิ์', 'danger')
        return redirect(url_for('vehicle.index'))
    ot = DriverOT.query.get_or_404(ot_id)
    ot.no_receipt = not ot.no_receipt
    db.session.commit()
    if _wants_json():
        return jsonify(ok=True)
    flash(f'{ot.ot_number}: {"ย้ายไปผู้ใช้จ่ายเอง" if ot.no_receipt else "นำกลับเข้ารายการออกใบ"}', 'success')
    return redirect(request.referrer or url_for('admincost.cost_summary'))



@admincost_bp.route('/admin/ot/create', methods=['POST'])
@login_required
def ot_create():
    """Manual standalone OT — เพิ่มเองผ่านปุ่ม 'เพิ่ม OT' (booking_id=None, ไม่ผูกงบ, 2026-06-09)"""
    if not is_vehicle_admin():
        flash('คุณไม่มีสิทธิ์', 'danger')
        return redirect(url_for('vehicle.index'))

    try:
        driver_id = int(request.form.get('driver_id'))
        ot_date   = datetime.strptime(request.form.get('date'), '%Y-%m-%d').date()
    except (TypeError, ValueError):
        if _wants_json():
            return jsonify(ok=False, msg='กรุณาเลือกคนขับและวันที่'), 400
        flash('กรุณาเลือกคนขับและวันที่', 'danger')
        return redirect(request.referrer or url_for('admincost.cost_summary'))

    new_slots = _parse_ot_slots(request.form, driver_id=driver_id, on_date=ot_date)
    if not new_slots:
        if _wants_json():
            return jsonify(ok=False, msg='ต้องมีช่วงเวลา OT อย่างน้อย 1 ช่วงที่ถูกต้อง'), 400
        flash('ต้องมีช่วงเวลา OT อย่างน้อย 1 ช่วงที่ถูกต้อง', 'danger')
        return redirect(request.referrer or url_for('admincost.cost_summary'))

    ot = DriverOT(
        booking_id   =None,                      # standalone — ไม่ผูก booking/งบ
        driver_id    =driver_id,
        ot_number    =mileage_svc.next_ot_number(ot_date.year),
        date         =ot_date,
        note         =request.form.get('note', '').strip() or None,
        status       ='unpaid',
        is_manual    =True,                      # แอดมินกรอกเอง — sync_ot_for_trip() ห้ามทับ (2026-07-27)
        total_hours  =round(sum(float(s.hours)  for s in new_slots), 2),
        total_amount =round(sum(float(s.amount) for s in new_slots), 2),
        created_at   =get_bkk_time(),
        created_by_id=current_user.id,
    )
    ot.slots = new_slots
    db.session.add(ot)
    db.session.commit()
    if _wants_json():
        return jsonify(ok=True)
    flash(f'เพิ่ม {ot.ot_number} เรียบร้อย', 'success')
    return redirect(request.referrer or url_for('admincost.cost_summary'))


@admincost_bp.route('/admin/ot/<int:ot_id>/edit', methods=['POST'])
@login_required
def ot_edit(ot_id):
    if not is_vehicle_admin():
        flash('คุณไม่มีสิทธิ์', 'danger')
        return redirect(url_for('vehicle.index'))
    ot = DriverOT.query.get_or_404(ot_id)

    ot.driver_id = int(request.form.get('driver_id', ot.driver_id))
    ot.date      = datetime.strptime(request.form.get('date'), '%Y-%m-%d').date()
    ot.note      = request.form.get('note', '').strip() or None

    new_slots       = _parse_ot_slots(request.form, driver_id=ot.driver_id,
                                      on_date=ot.date, exclude_ot_id=ot.id)
    ot.slots        = new_slots
    ot.total_hours  = round(sum(float(s.hours)  for s in new_slots), 2)
    ot.total_amount = round(sum(float(s.amount) for s in new_slots), 2)
    ot.is_manual    = True   # แก้ด้วยมือแล้ว — sync_ot_for_trip() ห้ามคำนวณทับ (2026-07-27)
    db.session.commit()
    if _wants_json():
        return jsonify(ok=True)
    flash(f'แก้ไข {ot.ot_number} เรียบร้อย', 'success')
    return redirect(request.referrer or url_for('admincost.cost_summary'))



@admincost_bp.route('/admin/ot/<int:ot_id>/delete', methods=['POST'])
@login_required
def ot_delete(ot_id):
    """Soft delete → tab 'ลบ' (2026-06-08) — กู้คืนได้ผ่าน ot_restore"""
    if not is_vehicle_admin():
        flash('คุณไม่มีสิทธิ์', 'danger')
        return redirect(url_for('vehicle.index'))
    ot = DriverOT.query.get_or_404(ot_id)
    ot.is_deleted = True
    ot.deleted_at = get_bkk_time()
    db.session.commit()
    if _wants_json():
        return jsonify(ok=True)
    flash(f'ลบ {ot.ot_number} แล้ว (กู้คืนได้ใน tab "ลบ")', 'success')
    return redirect(request.referrer or url_for('admincost.cost_summary'))



@admincost_bp.route('/admin/ot/<int:ot_id>/restore', methods=['POST'])
@login_required
def ot_restore(ot_id):
    """กู้คืน OT จาก tab 'ลบ'"""
    if not is_vehicle_admin():
        flash('คุณไม่มีสิทธิ์', 'danger')
        return redirect(url_for('vehicle.index'))
    ot = DriverOT.query.get_or_404(ot_id)
    ot.is_deleted = False
    ot.deleted_at = None
    db.session.commit()
    if _wants_json():
        return jsonify(ok=True)
    flash(f'กู้คืน {ot.ot_number} เรียบร้อย', 'success')
    return redirect(request.referrer or url_for('admincost.cost_summary'))



def _reject_midnight_crossing(start, end):
    """band ที่ end <= start (ข้ามเที่ยงคืน) คิดเงินไม่ได้ — build_ot_specs() จะได้ overlap ติดลบ
    แล้วข้ามแถวนั้นเงียบๆ กลายเป็น OT 0 บาท (bug B2, 2026-08-07) จึงต้องกันตั้งแต่ตอนบันทึก
    '24:00' = ปลายวัน ไม่ใช่ 00:00 ของวันถัดไป จึงเทียบเป็น 1440 นาที"""
    to_min = lambda hm: 1440 if hm == '24:00' else int(hm[:2]) * 60 + int(hm[3:5])
    try:
        return to_min(end) <= to_min(start)
    except (ValueError, IndexError):
        return True


@admincost_bp.route('/admin/ot/rate_config/update', methods=['POST'])
@login_required
def ot_rate_config_update():
    """บันทึกอัตรา OT ทั้งชุด (แท็บ "ตั้งค่า OT") — แถวเดิมอัปเดต, แถวใหม่ (cfg_id ว่าง) สร้าง,
    cfg_delete[] = soft-delete (is_active=False, JS ปิด input ก่อน submit แถวนั้นจึงไม่มีใน
    cfg_id[]/cfg_label[]/... ที่เหลือ ไม่ถูกแตะซ้ำในลูปอัปเดตด้านล่าง) — pattern เดียวกับ
    rateConfigModal เดิม (2026-08-08: เปลี่ยนกลับจาก toggle is_active ตามคำขอ — ลบแล้วลบเลย
    ไม่มี UI เปิดกลับ ต้องแก้ DB มือถ้าจะเอาคืน เหมือนของเดิม)"""
    if not is_vehicle_admin():
        if _wants_json():
            return jsonify(ok=False, msg='คุณไม่มีสิทธิ์'), 403
        flash('คุณไม่มีสิทธิ์', 'danger')
        return redirect(url_for('vehicle.index'))

    for did in request.form.getlist('cfg_delete[]'):
        if did:
            cfg = OTRateConfig.query.get(int(did))
            if cfg:
                cfg.is_active = False

    rows = list(zip(
        request.form.getlist('cfg_id[]'),
        request.form.getlist('cfg_label[]'),
        request.form.getlist('cfg_start[]'),
        request.form.getlist('cfg_end[]'),
        request.form.getlist('cfg_rate[]'),
        request.form.getlist('cfg_day[]'),
        request.form.getlist('cfg_rate_type[]'),
    ))

    bad = [label for _, label, start, end, rate, _, _ in rows
           if label and start and end and rate != '' and _reject_midnight_crossing(start, end)]
    if bad:
        msg = f'ช่วงข้ามเที่ยงคืนใช้ไม่ได้ ต้องแยกเป็น 2 ท่อน: {", ".join(bad)}'
        if _wants_json():
            return jsonify(ok=False, msg=msg), 400
        flash(msg, 'danger')
        return redirect(request.referrer or url_for('admincost.cost_summary'))

    max_order = db.session.query(db.func.coalesce(db.func.max(OTRateConfig.sort_order), 0)).scalar()
    for cfg_id, label, start, end, rate, day, rtype in rows:
        if not label or not start or not end or rate == '':
            continue
        day_val   = int(day) if day not in ('', None) else None
        # ค่าที่ไม่รู้จักจาก form → hourly เสมอ (พฤติกรรมเดิม) ห้ามเชื่อ input ตรงๆ เพราะคุมเงิน
        type_val  = RATE_FLAT_DAY if rtype == RATE_FLAT_DAY else RATE_HOURLY
        if cfg_id:
            cfg = OTRateConfig.query.get(int(cfg_id))
            if cfg:
                cfg.label = label; cfg.start_time = start
                cfg.end_time = end; cfg.rate = float(rate)
                cfg.day_of_week = day_val
                cfg.rate_type = type_val
        else:
            max_order += 10
            db.session.add(OTRateConfig(
                label=label, start_time=start, end_time=end,
                rate=float(rate), is_active=True, sort_order=max_order,
                day_of_week=day_val, rate_type=type_val,
            ))
    db.session.commit()
    if _wants_json():
        return jsonify(ok=True, msg='อัปเดตอัตรา OT เรียบร้อย')
    flash('อัปเดตอัตรา OT เรียบร้อย', 'success')
    return redirect(request.referrer or url_for('admincost.cost_summary'))

# ─────────────────────────────────────────────
# Driver View — หน้าคนขับ (mobile-friendly)
# ─────────────────────────────────────────────

@admincost_bp.route('/admin/cost/export')
@login_required
def cost_export():
    if not is_vehicle_admin():
        flash('คุณไม่มีสิทธิ์', 'danger')
        return redirect(url_for('vehicle.index'))

    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        flash('ไม่พบ openpyxl — รัน: pip install openpyxl', 'danger')
        return redirect(url_for('admincost.cost_summary'))

    from flask import send_file
    now        = get_bkk_time()
    from_month = int(request.args.get('from_month', now.month))
    from_year  = int(request.args.get('from_year',  now.year))
    to_month   = int(request.args.get('to_month',   now.month))
    to_year    = int(request.args.get('to_year',    now.year))
    sel_driver = request.args.get('driver_id', type=int)
    sel_status = request.args.get('status', '')
    f_budget_type = request.args.get('budget_type', '').strip()
    f_budget_sub  = request.args.get('budget_sub', '').strip()

    from_date = date(from_year, from_month, 1)
    to_date   = date(to_year + 1, 1, 1) if to_month == 12 else date(to_year, to_month + 1, 1)

    q = DriverOT.query.filter(DriverOT.date >= from_date, DriverOT.date < to_date)
    if sel_driver:
        q = q.filter(DriverOT.driver_id == sel_driver)
    q = _apply_budget_filter(q, f_budget_type, f_budget_sub)
    if sel_status == 'deleted':
        q = q.filter(DriverOT.is_deleted.is_(True))
    else:
        q = q.filter(DriverOT.is_deleted.is_(False))
        if sel_status == 'unpaid':
            q = q.filter(DriverOT.no_receipt.is_(False), DriverOT.status == 'unpaid')
        elif sel_status == 'paid':
            q = q.filter(DriverOT.no_receipt.is_(False), DriverOT.status == 'paid')
        elif sel_status == 'self_paid':
            q = q.filter(DriverOT.no_receipt.is_(True))
    ots = q.order_by(DriverOT.date).all()

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = f"OT {TH_MONTHS[from_month]}{from_year+543}"

    headers  = ['เลขที่','วันที่','คนขับ','Booking','สถานที่','ช่วงเวลา OT','ชม.','ยอด(฿)','สถานะ','หมายเหตุ']
    hdr_fill = PatternFill('solid', fgColor='4F46E5')
    hdr_font = Font(bold=True, color='FFFFFF', name='Sarabun')
    thin     = Side(style='thin', color='E4E4E7')
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)
    ST_LABEL = {'unpaid':'ยังไม่จ่าย','paid':'จ่ายแล้ว','pending':'ยังไม่จ่าย','approved':'ยังไม่จ่าย'}

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = border

    total_hrs = total_amt = 0
    for ri, ot in enumerate(ots, 2):
        slot_str = ' | '.join(
            f"{s.slot_label} {s.start_time}–{s.end_time} ฿{s.rate}" for s in ot.slots
        )
        row_data = [
            ot.ot_number,
            ot.date.strftime('%d/%m/%Y'),
            ot.driver.name if ot.driver else '-',
            ot.booking.id if ot.booking else '-',
            ot.booking.destination if ot.booking else '-',
            slot_str,
            float(ot.total_hours),
            float(ot.total_amount),
            ('ผู้ใช้จ่ายเอง' if ot.no_receipt else ST_LABEL.get(ot.status, ot.status)),
            ot.note or '',
        ]
        total_hrs += float(ot.total_hours)
        total_amt += float(ot.total_amount)
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border    = border
            cell.alignment = Alignment(horizontal='center' if ci in [1,2,7,8,9] else 'left')
            if ri % 2 == 0:
                cell.fill = PatternFill('solid', fgColor='F7F7F8')

    tr = len(ots) + 2
    ws.cell(row=tr, column=6, value='รวม').font = Font(bold=True)
    ws.cell(row=tr, column=7, value=round(total_hrs, 2)).font = Font(bold=True)
    ws.cell(row=tr, column=8, value=round(total_amt, 2)).font = Font(bold=True)

    for ci, w in enumerate([14,12,18,10,24,36,8,12,12,20], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 22

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"driver_ot_{from_year}_{from_month:02d}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ══════════════════════════════════════════════════════
# Feature 6: Vehicle Service/Tax Date Update
# ══════════════════════════════════════════════════════
